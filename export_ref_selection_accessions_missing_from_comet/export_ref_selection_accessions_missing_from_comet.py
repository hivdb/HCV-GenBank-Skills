#!/usr/bin/env python3
"""Export Ref-selection raw sequences whose accessions are absent from COMET CSVs."""
from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
KNOWN_QUASISPECIES = {"19", "31", "32", "34", "70", "81", "115", "262", "1044", "2043", "2071", "2129", "2139", "2175", "2195", "2212", "2216", "2225", "2324"}
JOBS = {
    "NS3": ("___IncludedNS3Refs.xlsx", "NS3_PtGT0_Check"),
    "NS5A": ("___IncludedNS5ARefs.xlsx", "NS5A_PtGT0_Check"),
    "NS5B": ("___IncludedNS5BRefs.xlsx", "NS5B_PtGT0_Check"),
}


def selected_refids(workbook: Path, sheet_name: str) -> tuple[set[str], list[str], dict[str, dict[str, object]]]:
    wb = load_workbook(workbook, read_only=True, data_only=True); ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True)); headers = {str(x or '').strip(): i for i, x in enumerate(rows[0])}
    column_names = [str(value or '').strip() for value in rows[0]]
    selected: dict[str, dict[str, object]] = {}
    for row in rows[1:]:
        refid = str(row[headers['RefID']] or '').strip()
        patient_count = str(row[headers['Num Pts']] or '').strip().casefold()
        if not refid or patient_count in {'exclude', 'check'} or refid in KNOWN_QUASISPECIES:
            continue
        selected[refid] = {column_names[index]: row[index] if index < len(row) else None for index in range(len(column_names))}
    wb.close(); return set(selected), column_names, selected


def fasta_records(path: Path):
    header = None; parts: list[str] = []
    for line in path.read_text(encoding='utf-8').splitlines():
        if line.startswith('>'):
            if header is not None: yield header, ''.join(parts)
            header, parts = line[1:].split()[0].split('.')[0], []
        else: parts.append(line.strip())
    if header is not None: yield header, ''.join(parts)


def main() -> None:
    fasta_dir = ROOT / 'HCVData' / 'FASTA'; comet_dir = ROOT / 'HCVData' / 'Comet Subtyping'; output_dir = ROOT / 'outputs' / 'comet'
    output_dir.mkdir(parents=True, exist_ok=True)
    for gene, (workbook_name, sheet_name) in JOBS.items():
        refids, selection_headers, selection_rows = selected_refids(ROOT / 'HCVData' / 'Ref-selection' / workbook_name, sheet_name)
        all_records = [record for fasta in fasta_dir.glob('*.fasta') if fasta.name.split('_', 1)[0] in refids for record in fasta_records(fasta)]
        with (comet_dir / f'{gene}.csv').open(newline='', encoding='utf-8-sig') as handle:
            comet_accessions = {row['name'].strip().split('.')[0] for row in csv.DictReader(handle) if row.get('name', '').strip()}
        unique_records = dict(all_records)
        missing = [(accession, sequence) for accession, sequence in unique_records.items() if accession not in comet_accessions]
        output = output_dir / f'{gene}_RefSelection_Accessions_NotIn_Comet.fasta'
        with output.open('w', encoding='utf-8') as handle:
            for accession, sequence in missing:
                handle.write(f'>{accession}\n{sequence}\n')
        by_refid: dict[str, set[str]] = defaultdict(set)
        present_by_refid: dict[str, set[str]] = defaultdict(set)
        for fasta in fasta_dir.glob('*.fasta'):
            refid = fasta.name.split('_', 1)[0]
            if refid not in refids:
                continue
            for accession, _sequence in fasta_records(fasta):
                by_refid[refid].add(accession)
                if accession in comet_accessions:
                    present_by_refid[refid].add(accession)
        report = output_dir / f'{gene}_RefSelection_RefIDs_With_Accessions_NotIn_Comet.csv'
        with report.open('w', newline='', encoding='utf-8') as handle:
            writer = csv.writer(handle)
            writer.writerow([*selection_headers, 'FastaAccessionsTotal', 'AccessionsInComet', 'AccessionsNotInComet'])
            for refid in sorted(by_refid, key=lambda value: int(value) if value.isdigit() else value):
                total = len(by_refid[refid]); present = len(present_by_refid[refid]); missing_count = total - present
                if missing_count:
                    writer.writerow([selection_rows[refid].get(header) for header in selection_headers] + [total, present, missing_count])
        print(f'{gene}: {len(missing)} of {len(unique_records)} unique Ref-selection FASTA accessions are absent from {gene}.csv; {output}')
        print(f'{gene}: {sum(1 for refid in by_refid if len(by_refid[refid]) > len(present_by_refid[refid]))} RefIDs have missing accessions; {report}')


if __name__ == '__main__': main()
