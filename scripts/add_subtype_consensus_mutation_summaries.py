#!/usr/bin/env python3
"""Add reference-to-consensus mutation summaries to subtype comparison workbooks."""
from __future__ import annotations
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font

VARIANT = re.compile(r"([A-Z*])(\d+(?:\.\d+)?)")

def profiles(path: Path) -> dict[tuple[str,str,int,str],str]:
    wb=load_workbook(path,read_only=True,data_only=True); ws=wb.active; result={}
    header=[c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
    for row in ws.iter_rows(min_row=2,values_only=True):
        label=str(row[0] or ''); m=re.match(r"GT(\d+)_(\S+) \(",label)
        if not m: continue
        for column,value in enumerate(row[1:],1):
            if column>=len(header) or not str(header[column] or '').startswith('P'): continue
            pos=int(str(header[column])[1:])
            for aa,pct in VARIANT.findall(str(value or '')): result[(m.group(1),m.group(2).lower(),pos,aa)]=pct
    wb.close(); return result

def main() -> None:
    root=Path('Reference_seqs'); comet=Path('outputs/comet')
    readme: dict[str,dict[str,set[str]]] = {}
    for gene,token in [('NS3','NS3'),('NS5A_NTD','NS5A'),('NS5B','NS5B')]:
        path=root/f'HCV_Subtype_Ref_vs_Comet_Subtype_Consensus_Aligned_{gene}.xlsx'
        values=profiles(comet/f'{token}_Subtype_RAS_Profiles.xlsx')
        wb=load_workbook(path); ws=wb.active
        col=next((cell.column for cell in ws[2] if cell.value=='Mutations'), ws.max_column+1)
        for row in range(1,ws.max_row+1):
            if ws.cell(row,1).value=='Gene': ws.cell(row,col).value='Mutations'; ws.cell(row,col).font=Font(bold=True)
            if ws.cell(row,4).value!='Consensus': continue
            refrow=row-1; gt=str(ws.cell(row,2).value or '').removeprefix('GT'); subtype=str(ws.cell(row,3).value or '').lower(); parts=[]; plain=[]
            for c in range(5,col):
                cons=ws.cell(row,c).value
                if not cons: continue
                ref=ws.cell(refrow,c).value; pos=ws.cell(2,c).value
                if not ref or not pos: continue
                mutation=f'{ref}{pos}{cons}'; pct=values.get((gt,subtype,int(pos),str(cons)))
                parts.extend([mutation, TextBlock(InlineFont(vertAlign='subscript'), pct or 'NA'), ', '])
                plain.append(f'{mutation} ({pct or "NA"}%)')
            if parts:
                parts.pop(); ws.cell(row,col).value=CellRichText(*parts)
                readme.setdefault(gene,{}).setdefault(f'GT{gt}_{subtype}',set()).update(plain)
        ws.column_dimensions[ws.cell(1,col).column_letter].width=50; wb.save(path)
    lines=['# Subtype reference-to-COMET-consensus mutations', '', 'Only subtypes with one or more RAS-position differences are listed. Percentages are from the final combined profile; `NA` means the mutation was not displayed there.', '']
    for gene in ('NS3','NS5A_NTD','NS5B'):
        lines.extend([f'## {gene}', ''])
        for subtype, mutations in sorted(readme.get(gene,{}).items(), key=lambda item: (int(item[0][2]), item[0])):
            lines.append(f'- {subtype}: {", ".join(sorted(mutations))}')
        lines.append('')
    (root/'README_Subtype_Consensus_Mutations.md').write_text('\n'.join(lines),encoding='utf-8')

if __name__=='__main__': main()
