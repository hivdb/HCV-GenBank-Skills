#!/usr/bin/env python3
"""Summarize Num Pts=Check RefIDs and their workflow-specific filters."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


FILTERS = {
    "NS3": {
        "30": "source_isolate contains Day1", "2116": "source_collection_date before 2011",
        "943": "source_isolate contains Day 1", "2150": "source_isolate contains b",
        "600": "source_isolate does not contain failure", "884": "source_isolate contains Pre-TH",
        "499": "source_isolate contains HCC", "2227": "Accession membership list: temp/hcv-ns3-comet-build-workflow/run_ns3_pipeline/2227_Nguyen_(2015)_w_metadata_filtered.csv",
        "661": "source_isolation_source equals plasma", "2168": "source_isolate contains pre",
        "1356": "source_isolate does not contain IC", "192": "source_isolate contains day 1",
        "2138": "source_isolate contains Week 0", "346": "source_isolate contains baseline/D0",
    },
    "NS5A": {
        "29": "source_isolate contains SCRN", "600": "source_isolate does not contain failure",
        "535": "Accession membership list: temp/hcv-ns5a-comet-build-workflow/run_ns5a_pipeline/535.csv",
        "661": "source_isolation_source equals plasma", "123": "source_isolate does not contain TF",
        "50": "source_isolate contains week 0", "192": "source_isolate contains day1",
        "142": "source_isolate contains baseline", "17": "Accession membership list: temp/hcv-ns5a-comet-build-workflow/run_ns5a_pipeline/17.csv",
        "288": "source_isolate contains pre", "346": "source_isolate contains baseline/D0",
    },
    "NS5B": {
        "891": "source_isolate contains token Ha01 through Ha97", "30": "source_isolate contains day1",
        "943": "source_isolate contains day 1", "1051": "source_isolate contains token 1a through 51a",
        "192": "source_isolate contains day1", "17": "Accession membership list: temp/hcv-ns5b-comet-build-workflow/run_ns5b_pipeline/17.csv",
        "346": "source_isolate contains baseline",
    },
}

WORKFLOW_TEMP_ROOTS = {
    "NS3": Path("temp/hcv-ns3-comet-build-workflow/run_ns3_pipeline"),
    "NS5A": Path("temp/hcv-ns5a-comet-build-workflow/run_ns5a_pipeline"),
    "NS5B": Path("temp/hcv-ns5b-comet-build-workflow/run_ns5b_pipeline"),
}


def filter_result_creation(gene: str, refid: str, information: str) -> str:
    if refid not in FILTERS[gene]:
        return "No FilterResultFile is created because this RefID has no configured RefID-specific filter."
    return (
        f"Step 5 reads rows with RefID {refid} from included_accessions_metadata.csv, "
        f"keeps rows where {information}, and writes the retained full metadata rows to FilterResultFile."
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--output-csv", required=True)
    args = parser.parse_args()
    workbook_path = Path(args.workbook)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    for gene in ("NS3", "NS5A", "NS5B"):
        sheet_name = f"{gene}_PtGT0_Check"
        sheet = workbook[sheet_name]
        header = [str(value or "") for value in next(sheet.iter_rows(values_only=True))]
        refid_index, num_pts_index = header.index("RefID"), header.index("Num Pts")
        for values in sheet.iter_rows(min_row=2, values_only=True):
            num_pts = str(values[num_pts_index] or "").strip()
            if num_pts.casefold() != "check":
                continue
            refid = str(values[refid_index] or "").strip()
            filter_information = FILTERS[gene].get(refid, "No RefID-specific filter configured in the current workflow")
            result_path = WORKFLOW_TEMP_ROOTS[gene] / "refid_metadata" / f"RefID_{refid}_metadata.csv"
            rows.append({
                "OldFileName": workbook_path.name,
                "SheetName": sheet_name,
                "GeneName": gene,
                "RefID": refid,
                "NumPtsValue": num_pts,
                "FilterInformation": filter_information,
                "FilterResultFile": str(result_path) if refid in FILTERS[gene] and result_path.is_file() else "",
                "FilterResultCreation": filter_result_creation(gene, refid, filter_information),
            })
    workbook.close()
    output = Path(args.output_csv)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["OldFileName", "SheetName", "GeneName", "RefID", "NumPtsValue", "FilterInformation", "FilterResultFile", "FilterResultCreation"], lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    print(f"output_csv={output.resolve()}\ncheck_refid_count={len(rows)}")


if __name__ == "__main__":
    main()
