#!/usr/bin/env python3
"""Build Comet profile-input CSVs and compare them to local alignment."""
from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook

GENES = ("NS3", "NS5A", "NS5B")


def profile_rows(source: Path) -> dict[str, tuple[str, str]]:
    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(v) if v is not None else "" for v in next(ws.iter_rows(values_only=True))]
    idx = {name: i for i, name in enumerate(header)}
    required = {"AccessionID", "ClosestGT", "ClosestSubtype", "StartAAPosition", "EndAAPosition", "AASequence"}
    if missing := required.difference(idx):
        raise RuntimeError(f"Columns missing from {source}: {', '.join(sorted(missing))}")
    rows: dict[str, tuple[str, str]] = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        accession = str(values[idx["AccessionID"]] or "").strip()
        genotype = str(values[idx["ClosestGT"]] or "").strip()
        subtype = str(values[idx["ClosestSubtype"]] or "").strip()
        if not accession or not values[idx["AASequence"]] or values[idx["StartAAPosition"]] in (None, "") or values[idx["EndAAPosition"]] in (None, ""):
            continue
        if genotype.casefold().startswith("unassign") or subtype.casefold().startswith("unassign"):
            continue
        rows[accession] = (genotype, subtype)
    wb.close()
    return rows


def read_csv(path: Path) -> dict[str, tuple[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return {r["accession"].strip(): (r["genotype"].strip(), r["subtype"].strip()) for r in csv.DictReader(handle) if r.get("accession", "").strip()}


def write_rows(path: Path, rows: dict[str, tuple[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["accession", "genotype", "subtype"])
        writer.writerows((a, *rows[a]) for a in sorted(rows))


def write_comparison(gene: str, comet: dict[str, tuple[str, str]], local: dict[str, tuple[str, str]], out: Path, status: str) -> None:
    shared = sorted(set(comet) & set(local))
    same_gt = sum(comet[a][0] == local[a][0] for a in shared)
    same_subtype = sum(comet[a][1] == local[a][1] for a in shared)
    diff_gt, diff_subtype = len(shared) - same_gt, len(shared) - same_subtype
    summary = [("status", status), ("comet_profile_accession_count", len(comet)), ("local_profile_accession_count", len(local)), ("shared_accession_count", len(shared)), ("same_genotype_count", same_gt), ("different_genotype_count", diff_gt), ("different_genotype_percent", f"{100 * diff_gt / len(shared) if shared else 0:.2f}"), ("same_subtype_count", same_subtype), ("different_subtype_count", diff_subtype), ("different_subtype_percent", f"{100 * diff_subtype / len(shared) if shared else 0:.2f}")]
    with (out / f"{gene}_Profile_Input_Assignment_Comparison.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle); writer.writerow(["metric", "value"]); writer.writerows(summary)
    with (out / f"{gene}_Profile_Input_Assignment_Differences.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle); writer.writerow(["accession", "comet_genotype", "local_genotype", "comet_subtype", "local_subtype"])
        for accession in shared:
            if comet[accession] != local[accession]:
                writer.writerow([accession, comet[accession][0], local[accession][0], comet[accession][1], local[accession][1]])
    for metric, value in summary[4:]: print(f"{gene.lower()}_{metric}={value}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--repo-root", type=Path, default=Path.cwd())
    root = parser.parse_args().repo_root.resolve()
    comet_dir, local_dir = root / "outputs" / "comet", root / "outputs" / "local_alignment"
    for gene in GENES:
        source = comet_dir / f"{gene}_Profile_Input_Source.xlsx"
        if not source.is_file():
            print(f"{gene.lower()}_status=comet_profile_input_source_missing"); continue
        comet = profile_rows(source)
        write_rows(comet_dir / f"{gene}_Profile_Input_Accessions.csv", comet)
        local_path = local_dir / f"{gene}_Profile_Input_Accessions.csv"
        local = read_csv(local_path) if local_path.is_file() else {}
        write_comparison(gene, comet, local, comet_dir, "ok" if local_path.is_file() else "local_profile_input_csv_missing")
        print(f"{gene.lower()}_profile_input_accession_count={len(comet)}")
    return 0


if __name__ == "__main__": raise SystemExit(main())
