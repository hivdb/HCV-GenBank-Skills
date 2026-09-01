#!/usr/bin/env python3
"""Report per-subtype non-X coverage at COMET RAS positions."""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

COMBINED_SUBTYPE_RE = re.compile(r"^GT(?P<gt>\d+)_(?P<subtype>\S+) \((?P<count>\d+),")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--gene", required=True)
    p.add_argument("--combined-profile-workbook", required=True)
    p.add_argument("--profile-input-workbook", required=True)
    p.add_argument("--profile-accessions-csv", required=True)
    p.add_argument("--output-xlsx", required=True)
    p.add_argument("--threshold", type=float, default=95.0)
    return p.parse_args()


def ranges(positions: list[int]) -> str:
    if not positions:
        return "none"
    result, start, previous = [], positions[0], positions[0]
    for position in positions[1:]:
        if position != previous + 1:
            result.append(str(start) if start == previous else f"{start}-{previous}")
            start = position
        previous = position
    result.append(str(start) if start == previous else f"{start}-{previous}")
    return ", ".join(result)


def main() -> int:
    a = parse_args()
    combined = load_workbook(
        a.combined_profile_workbook, read_only=True, data_only=True
    )
    combined_sheet = combined.active
    positions = [
        int(str(c.value)[1:])
        for c in combined_sheet[1]
        if re.fullmatch(r"P\d+", str(c.value or ""))
    ]
    combined_subtypes = {
        (match.group("gt"), match.group("subtype").lower()): int(match.group("count"))
        for (label,) in combined_sheet.iter_rows(min_row=2, max_col=1, values_only=True)
        if (match := COMBINED_SUBTYPE_RE.match(str(label or "")))
    }
    combined.close()
    with open(a.profile_accessions_csv, newline="", encoding="utf-8-sig") as f:
        allowed = {
            (
                r["accession"].strip(),
                r["genotype"].strip(),
                r["subtype"].strip().lower(),
            )
            for r in csv.DictReader(f)
            if r.get("accession", "").strip()
        }

    wb = load_workbook(a.profile_input_workbook, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ix = {h: i for i, h in enumerate(headers)}
    groups: dict[tuple[str, str], list[tuple[int, str]]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        accession = str(row[ix["AccessionID"]] or "").strip()
        gt = str(row[ix["ClosestGT"]] or "").strip().removeprefix("GT")
        subtype = str(row[ix["ClosestSubtype"]] or "").strip().lower()
        if (accession, gt, subtype) not in allowed:
            continue
        if (gt, subtype) not in combined_subtypes:
            continue
        if (
            "AlignmentQCStatus" in ix
            and str(row[ix["AlignmentQCStatus"]] or "").strip() != "PASS"
        ):
            continue
        start, sequence = (
            row[ix["StartAAPosition"]],
            str(row[ix["AASequence"]] or "").strip().upper(),
        )
        if start and sequence:
            groups[(gt, subtype)].append((int(start), sequence))
    wb.close()

    out = Workbook()
    summary = out.active
    summary.title = "Subtype_Summary"
    summary.append(
        [
            "Gene",
            "Genotype",
            "Subtype",
            "ProfileSequences",
            "RASPositionCount",
            "FullyCoveredPositions",
            "PartiallyCoveredPositions",
            "Conclusion",
        ]
    )
    detail = out.create_sheet("Position_Coverage")
    detail.append(
        [
            "Gene",
            "Genotype",
            "Subtype",
            "Position",
            "ProfileSequences",
            "NonXCalls",
            "NoCallOrX",
            "CoveragePercent",
            "CoverageStatus",
        ]
    )
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for sheet in (summary, detail):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        sheet.freeze_panes = "A2"
    for (gt, subtype), sequences in sorted(
        groups.items(), key=lambda x: (int(x[0][0]), x[0][1])
    ):
        total = combined_subtypes[(gt, subtype)]
        covered, partial, critically_low = [], [], []
        for position in positions:
            non_x = sum(
                0 <= position - start < len(sequence)
                and sequence[position - start] != "X"
                for start, sequence in sequences
            )
            pct = 100 * non_x / total if total else 0
            status = "Well covered" if pct >= a.threshold else "Partial"
            (covered if status == "Well covered" else partial).append(position)
            if pct < 50:
                critically_low.append((position, non_x, pct))
            detail.append(
                [
                    a.gene,
                    f"GT{gt}",
                    subtype,
                    position,
                    total,
                    non_x,
                    total - non_x,
                    pct,
                    status,
                ]
            )
        if not partial:
            conclusion = ""
        else:
            conclusion = f"Poorly covered (<{a.threshold:g}%): P{ranges(partial)}."
            if critically_low:
                low_text = ", ".join(
                    f"P{position} ({count}/{total}; {pct:.1f}%)"
                    for position, count, pct in critically_low
                )
                conclusion += f" Below 50%: {low_text}."
        summary.append(
            [
                a.gene,
                f"GT{gt}",
                subtype,
                total,
                len(positions),
                f"P{ranges(covered)}",
                f"P{ranges(partial)}",
                conclusion,
            ]
        )
    for row in detail.iter_rows(min_row=2, min_col=8, max_col=8):
        row[0].number_format = "0.0"
    for sheet in (summary, detail):
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = min(
                max(max(len(str(c.value or "")) for c in col) + 2, 12), 80
            )
    Path(a.output_xlsx).parent.mkdir(parents=True, exist_ok=True)
    out.save(a.output_xlsx)
    print(f"Wrote {a.output_xlsx} ({len(groups)} subtypes)")


if __name__ == "__main__":
    main()
