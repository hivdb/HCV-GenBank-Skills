#!/usr/bin/env python3
"""Create an accession-level NS3 coverage report for one genotype/subtype profile."""

from __future__ import annotations

import argparse
import csv
from collections import Counter, defaultdict
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


VALID_AAS = set("ACDEFGHIKLMNPQRSTVWY")
DEFAULT_RAS_POSITIONS = (
    36,
    41,
    43,
    54,
    55,
    56,
    80,
    122,
    155,
    156,
    158,
    166,
    168,
    170,
    175,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--profile-input-workbook", required=True)
    parser.add_argument("--profile-accessions-csv", required=True)
    parser.add_argument("--genotype", required=True)
    parser.add_argument("--subtype", required=True)
    parser.add_argument("--range-start", type=int, default=36)
    parser.add_argument("--range-end", type=int, default=175)
    parser.add_argument(
        "--ras-positions", default=",".join(map(str, DEFAULT_RAS_POSITIONS))
    )
    parser.add_argument("--output-xlsx", required=True)
    parser.add_argument("--position-coverage-csv", required=True)
    parser.add_argument("--position-coverage-png", required=True)
    return parser.parse_args()


def load_profile_accessions(path: Path, genotype: str, subtype: str) -> set[str]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return {
            str(row.get("accession", "")).strip()
            for row in csv.DictReader(handle)
            if str(row.get("accession", "")).strip()
            and str(row.get("genotype", "")).strip().removeprefix("GT") == genotype
            and str(row.get("subtype", "")).strip().lower() == subtype.lower()
        }


def load_source_rows(
    path: Path, accessions: set[str], genotype: str, subtype: str
) -> dict[str, list[dict[str, object]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    header = [str(value or "") for value in next(worksheet.iter_rows(values_only=True))]
    index = {name: number for number, name in enumerate(header)}
    required = {
        "AccessionID",
        "ClosestGT",
        "ClosestSubtype",
        "StartAAPosition",
        "EndAAPosition",
        "AASequence",
    }
    missing = required - index.keys()
    if missing:
        raise RuntimeError(f"Missing columns in {path}: {', '.join(sorted(missing))}")

    rows_by_accession: dict[str, list[dict[str, object]]] = defaultdict(list)
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        accession = str(values[index["AccessionID"]] or "").strip()
        row_genotype = str(values[index["ClosestGT"]] or "").strip().removeprefix("GT")
        row_subtype = str(values[index["ClosestSubtype"]] or "").strip().lower()
        if (
            accession not in accessions
            or row_genotype != genotype
            or row_subtype != subtype.lower()
        ):
            continue
        rows_by_accession[accession].append(
            {name: values[position] for name, position in index.items()}
        )
    workbook.close()
    return rows_by_accession


def calls_by_position(row: dict[str, object]) -> dict[int, str]:
    start = row.get("StartAAPosition")
    sequence = str(row.get("AASequence") or "").strip().upper()
    if start in (None, ""):
        return {}
    return {int(start) + offset: aa for offset, aa in enumerate(sequence)}


def coverage_row(
    accession: str,
    source_rows: list[dict[str, object]],
    range_positions: tuple[int, ...],
    ras_positions: tuple[int, ...],
) -> dict[str, object]:
    candidates = []
    for row in source_rows:
        calls = calls_by_position(row)
        ras_valid = sum(calls.get(position) in VALID_AAS for position in ras_positions)
        range_valid = sum(
            calls.get(position) in VALID_AAS for position in range_positions
        )
        candidates.append((ras_valid, range_valid, len(calls), row, calls))
    if not candidates:
        return {
            "Accession": accession,
            "SourceRecordCount": 0,
            "CoverageStatus": "MISSING_FROM_QC_WORKBOOK",
            "RASPositionsCovered": 0,
            "RangePositionsCovered": 0,
            "_calls": {},
        }

    # One accession can be present in multiple source studies.  Select the row
    # with the greatest RAS coverage, then range coverage, for an unambiguous
    # accession-level report while retaining the source-record count for audit.
    ras_valid, range_valid, _, row, calls = max(candidates, key=lambda item: item[:3])
    covered_ras = [
        position for position in ras_positions if calls.get(position) in VALID_AAS
    ]
    missing_ras = [
        position for position in ras_positions if position not in covered_ras
    ]
    start = row.get("StartAAPosition")
    end = row.get("EndAAPosition")
    qc_status = str(row.get("AlignmentQCStatus") or "")
    return {
        "Accession": accession,
        "SourceRecordCount": len(source_rows),
        "SelectedRefID": row.get("RefID"),
        "SelectedRefName": row.get("RefName"),
        "StartAAPosition": start,
        "EndAAPosition": end,
        "AlignmentQCStatus": qc_status,
        "CoverageStatus": "COMPLETE_RAS_AND_RANGE"
        if ras_valid == len(ras_positions) and range_valid == len(range_positions)
        else "COMPLETE_RAS"
        if ras_valid == len(ras_positions)
        else "PARTIAL_RAS",
        "RASPositionsCovered": ras_valid,
        "RASPositionsMissing": len(missing_ras),
        "RASCovered": ",".join(f"P{position}" for position in covered_ras),
        "RASMissing": ",".join(f"P{position}" for position in missing_ras),
        "CompleteRASCoverage": "Yes" if ras_valid == len(ras_positions) else "No",
        "RangePositionsCovered": range_valid,
        "RangePositionCount": len(range_positions),
        "RangeCoveragePercent": 100 * range_valid / len(range_positions),
        "CompleteRangeCoverage": "Yes" if range_valid == len(range_positions) else "No",
        **{
            f"P{position}": calls.get(position, "")
            if calls.get(position) in VALID_AAS
            else ""
            for position in ras_positions
        },
        "_calls": calls,
    }


def write_workbook(
    path: Path,
    rows: list[dict[str, object]],
    genotype: str,
    subtype: str,
    range_positions: tuple[int, ...],
    ras_positions: tuple[int, ...],
) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    summary.append(["Profile", f"GT{genotype}_{subtype}"])
    summary.append(["Profile accessions", len(rows)])
    summary.append(["RAS positions", ",".join(map(str, ras_positions))])
    summary.append(["Range", f"{range_positions[0]}-{range_positions[-1]}"])
    summary.append(
        [
            "Complete RAS coverage",
            sum(row.get("CompleteRASCoverage") == "Yes" for row in rows),
        ]
    )
    summary.append(
        [
            "Complete range coverage",
            sum(row.get("CompleteRangeCoverage") == "Yes" for row in rows),
        ]
    )
    summary.append(
        [
            "Missing from QC workbook",
            sum(
                row.get("CoverageStatus") == "MISSING_FROM_QC_WORKBOOK" for row in rows
            ),
        ]
    )
    summary.append([])
    summary.append(["Coverage status", "Accession count"])
    for status, count in sorted(
        Counter(str(row.get("CoverageStatus", "")) for row in rows).items()
    ):
        summary.append([status, count])

    detail = workbook.create_sheet("Accession_Coverage")
    columns = [
        "Accession",
        "SourceRecordCount",
        "SelectedRefID",
        "SelectedRefName",
        "StartAAPosition",
        "EndAAPosition",
        "AlignmentQCStatus",
        "CoverageStatus",
        "RASPositionsCovered",
        "RASPositionsMissing",
        "RASCovered",
        "RASMissing",
        "CompleteRASCoverage",
        "RangePositionsCovered",
        "RangePositionCount",
        "RangeCoveragePercent",
        "CompleteRangeCoverage",
        *[f"P{position}" for position in ras_positions],
    ]
    detail.append(columns)
    for row in sorted(
        rows,
        key=lambda item: (
            str(item.get("CoverageStatus")) != "COMPLETE_RAS_AND_RANGE",
            str(item["Accession"]),
        ),
    ):
        detail.append([row.get(column, "") for column in columns])
    for cell in detail[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
    for row in detail.iter_rows(min_row=2):
        if row[7].value != "COMPLETE_RAS_AND_RANGE":
            for cell in row:
                cell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = detail.dimensions
    for column, width in {
        "A": 16,
        "B": 18,
        "C": 14,
        "D": 28,
        "E": 17,
        "F": 15,
        "G": 20,
        "H": 27,
        "I": 19,
        "J": 19,
        "K": 34,
        "L": 34,
        "M": 21,
        "N": 22,
        "O": 18,
        "P": 22,
        "Q": 23,
    }.items():
        detail.column_dimensions[column].width = width
    for number in range(18, detail.max_column + 1):
        detail.column_dimensions[get_column_letter(number)].width = 8
    for row in detail.iter_rows(min_row=2, min_col=16, max_col=16):
        row[0].number_format = "0.0"

    for sheet in workbook.worksheets:
        sheet.freeze_panes = sheet.freeze_panes or "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_position_coverage_csv(
    path: Path,
    rows: list[dict[str, object]],
    range_positions: tuple[int, ...],
    ras_positions: tuple[int, ...],
) -> None:
    """Write raw and unambiguous accession coverage at every requested position."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=["Pos", "IsRASPosition", "#Seq", "#NonMixture"]
        )
        writer.writeheader()
        for position in range_positions:
            calls = [dict(row.get("_calls", {})).get(position, "") for row in rows]
            writer.writerow(
                {
                    "Pos": position,
                    "IsRASPosition": "Yes" if position in ras_positions else "",
                    "#Seq": sum(bool(call) for call in calls),
                    "#NonMixture": sum(call in VALID_AAS for call in calls),
                }
            )


def write_position_coverage_png(
    path: Path,
    rows: list[dict[str, object]],
    range_positions: tuple[int, ...],
    ras_positions: tuple[int, ...],
    genotype: str,
    subtype: str,
) -> None:
    """Plot raw positional coverage, retaining ambiguous and stop AA calls."""
    coverage = [
        sum(bool(dict(row.get("_calls", {})).get(position, "")) for row in rows)
        for position in range_positions
    ]
    colors = [
        "#C0504D" if position in ras_positions else "#5B9BD5"
        for position in range_positions
    ]
    figure, axis = plt.subplots(
        figsize=(max(22, len(range_positions) * 0.08), 7), constrained_layout=True
    )
    axis.bar(range_positions, coverage, color=colors, width=0.9, linewidth=0)
    axis.set_title(
        f"NS3 GT{genotype}_{subtype}: sequence coverage by amino-acid position"
    )
    axis.set_xlabel("NS3 amino-acid position")
    axis.set_ylabel("Sequences with any AA call")
    axis.set_xlim(range_positions[0] - 1, range_positions[-1] + 1)
    axis.set_ylim(0, max(coverage, default=0) + 3)
    axis.set_xticks(range_positions)
    axis.tick_params(axis="x", labelrotation=90, labelsize=5)
    axis.grid(axis="y", alpha=0.3)
    for position in ras_positions:
        axis.axvline(position, color="#C0504D", alpha=0.15, linewidth=0.8)
    axis.bar([], [], color="#5B9BD5", label="Any AA call (including X and stop)")
    axis.bar([], [], color="#C0504D", label="RAS position")
    axis.legend(loc="lower left")
    path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(path, dpi=200)
    plt.close(figure)


def main() -> int:
    args = parse_args()
    if args.range_start > args.range_end:
        raise ValueError("--range-start must be no greater than --range-end")
    ras_positions = tuple(
        int(value) for value in args.ras_positions.split(",") if value.strip()
    )
    range_positions = tuple(range(args.range_start, args.range_end + 1))
    accessions = load_profile_accessions(
        Path(args.profile_accessions_csv), args.genotype, args.subtype
    )
    source_rows = load_source_rows(
        Path(args.profile_input_workbook), accessions, args.genotype, args.subtype
    )
    rows = [
        coverage_row(
            accession, source_rows.get(accession, []), range_positions, ras_positions
        )
        for accession in accessions
    ]
    write_workbook(
        Path(args.output_xlsx),
        rows,
        args.genotype,
        args.subtype,
        range_positions,
        ras_positions,
    )
    write_position_coverage_csv(
        Path(args.position_coverage_csv), rows, range_positions, ras_positions
    )
    write_position_coverage_png(
        Path(args.position_coverage_png),
        rows,
        range_positions,
        ras_positions,
        args.genotype,
        args.subtype,
    )
    print(f"Wrote {len(rows)} accession rows to {Path(args.output_xlsx).resolve()}")
    print(
        f"Wrote position coverage CSV to {Path(args.position_coverage_csv).resolve()}"
    )
    print(
        f"Wrote position coverage chart to {Path(args.position_coverage_png).resolve()}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
