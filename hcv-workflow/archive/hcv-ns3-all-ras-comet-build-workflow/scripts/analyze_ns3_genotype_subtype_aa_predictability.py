#!/usr/bin/env python3
"""Measure genotype and subtype information about amino acids at RAS sites."""

from __future__ import annotations

import argparse
import math
import re
from collections import defaultdict
from pathlib import Path
from typing import TypeAlias

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


DEFAULT_RESISTANCE_POSITIONS = (
    36,
    41,
    43,
    54,
    55,
    56,
    80,
    122,
    155,
    156,
    158,
    166,
    168,
    170,
    175,
)
EXCLUDED_AAS = {"*", "X"}
Counts: TypeAlias = dict[str, int]
SubtypeCounts: TypeAlias = dict[str, dict[int, Counts]]
Coverage: TypeAlias = dict[str, dict[int, int]]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--subtype-profile-workbook", required=True)
    parser.add_argument("--output-xlsx", required=True)
    parser.add_argument("--gene", default="NS3")
    parser.add_argument("--position-column", default="NS3Position")
    parser.add_argument(
        "--positions",
        default=",".join(str(position) for position in DEFAULT_RESISTANCE_POSITIONS),
        help="Comma-separated RAS amino-acid positions.",
    )
    parser.add_argument(
        "--min-subtype-sequences",
        type=int,
        default=1,
        help="Minimum covered sequences for a subtype at a position (default: 1).",
    )
    return parser.parse_args()


def entropy(counts: Counts) -> float:
    total = sum(counts.values())
    if total == 0:
        return 0.0
    return -sum(
        (count / total) * math.log2(count / total)
        for count in counts.values()
        if count > 0
    )


def genotype_for_subtype(subtype: str) -> str:
    match = re.match(r"^(\d+)", subtype.strip())
    if not match:
        raise ValueError(f"Cannot derive genotype from subtype {subtype!r}")
    return f"GT{match.group(1)}"


def genotype_sort_key(genotype: str) -> tuple[int, str]:
    match = re.match(r"^GT(\d+)$", genotype)
    return (int(match.group(1)), genotype) if match else (999, genotype)


def subtype_sort_key(subtype: str) -> tuple[int, str]:
    match = re.match(r"^(\d+)(.*)$", subtype)
    return (int(match.group(1)), match.group(2)) if match else (999, subtype)


def load_profile_counts(
    path: Path, position_column: str, positions: set[int]
) -> tuple[SubtypeCounts, Coverage]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    header = [
        str(value or "").strip() for value in next(sheet.iter_rows(values_only=True))
    ]
    index = {name: position for position, name in enumerate(header)}
    required = {
        "Subtype",
        position_column,
        "NumSeqsIncludingPosition",
        "AminoAcid",
        "CountWithAA",
    }
    missing = sorted(required - index.keys())
    if missing:
        raise RuntimeError(f"{path} is missing columns: {', '.join(missing)}")

    counts: SubtypeCounts = defaultdict(lambda: defaultdict(dict))
    coverage: Coverage = defaultdict(dict)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        subtype = str(row[index["Subtype"]] or "").strip()
        if not subtype:
            continue
        position = int(row[index[position_column]])
        if position not in positions:
            continue
        covered = int(row[index["NumSeqsIncludingPosition"]] or 0)
        previous_coverage = coverage[subtype].get(position)
        if previous_coverage is not None and previous_coverage != covered:
            raise RuntimeError(
                f"Inconsistent coverage for subtype {subtype}, position {position}: "
                f"{previous_coverage} and {covered}"
            )
        coverage[subtype][position] = covered
        amino_acid = str(row[index["AminoAcid"]] or "").strip().upper()
        if amino_acid in EXCLUDED_AAS or not amino_acid:
            continue
        count = int(row[index["CountWithAA"]] or 0)
        if count > 0:
            position_counts = counts[subtype][position]
            position_counts[amino_acid] = position_counts.get(amino_acid, 0) + count
    workbook.close()
    return counts, coverage


def weighted_entropy(groups: list[tuple[int, Counts]]) -> float:
    total = sum(weight for weight, _ in groups)
    if total == 0:
        return 0.0
    return sum((weight / total) * entropy(counts) for weight, counts in groups)


def summarize_position(
    position: int,
    counts: SubtypeCounts,
    coverage: Coverage,
    min_subtype_sequences: int,
) -> tuple[dict[str, int | float | None], list[dict[str, int | str | bool]]]:
    all_covered = 0
    eligible_groups: dict[str, list[tuple[str, int, Counts]]] = defaultdict(list)
    eligibility_rows: list[dict[str, int | str | bool]] = []

    for subtype in sorted(coverage, key=subtype_sort_key):
        covered = coverage[subtype].get(position, 0)
        valid_counts = counts[subtype].get(position, {})
        analyzed = sum(valid_counts.values())
        eligible = covered >= min_subtype_sequences and analyzed > 0
        genotype = genotype_for_subtype(subtype)
        all_covered += covered
        eligibility_rows.append(
            {
                "Position": position,
                "Genotype": genotype,
                "Subtype": subtype,
                "CoveredSequences": covered,
                "AnalyzedSequences": analyzed,
                "Eligible": eligible,
                "ExclusionReason": ""
                if eligible
                else (
                    "below minimum covered sequences"
                    if covered < min_subtype_sequences
                    else "no non-ambiguous amino acids"
                ),
            }
        )
        if eligible:
            eligible_groups[genotype].append((subtype, analyzed, valid_counts))

    all_counts: Counts = defaultdict(int)
    genotype_groups: list[tuple[int, Counts]] = []
    subtype_groups: list[tuple[int, Counts]] = []
    for genotype in sorted(eligible_groups, key=genotype_sort_key):
        genotype_counts: Counts = defaultdict(int)
        genotype_weight = 0
        for _, subtype_weight, subtype_counts in eligible_groups[genotype]:
            genotype_weight += subtype_weight
            subtype_groups.append((subtype_weight, subtype_counts))
            for amino_acid, count in subtype_counts.items():
                genotype_counts[amino_acid] += count
                all_counts[amino_acid] += count
        genotype_groups.append((genotype_weight, genotype_counts))

    overall_entropy = entropy(all_counts)
    genotype_entropy = weighted_entropy(genotype_groups)
    genotype_subtype_entropy = weighted_entropy(subtype_groups)
    genotype_information = overall_entropy - genotype_entropy
    subtype_information = genotype_entropy - genotype_subtype_entropy
    percent_resolved = (
        100 * subtype_information / genotype_entropy if genotype_entropy > 0 else None
    )
    summary = {
        "Position": position,
        "CoveredSequencesAllSubtypes": all_covered,
        "AnalyzedSequencesEligibleSubtypes": sum(all_counts.values()),
        "EligibleGenotypes": len(eligible_groups),
        "EligibleSubtypes": sum(len(groups) for groups in eligible_groups.values()),
        "OverallEntropyBits": overall_entropy,
        "EntropyGivenGenotypeBits": genotype_entropy,
        "EntropyGivenGenotypeAndSubtypeBits": genotype_subtype_entropy,
        "GenotypeInformationBits": genotype_information,
        "AdditionalSubtypeInformationBits": subtype_information,
        "GenotypeConditionedUncertaintyResolvedBySubtypePct": percent_resolved,
    }
    return summary, eligibility_rows


def style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 48
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    for column in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[column[0].column_letter].width = min(width, 42)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000000"


def significant_decimal_number_format(value: float, digits: int = 2) -> str:
    """Return an Excel format with significant digits after the decimal point."""
    fractional_part = abs(value) % 1
    if fractional_part == 0:
        decimal_places = digits
    else:
        decimal_places = max(
            digits,
            digits - 1 - math.floor(math.log10(fractional_part)),
        )
    return "0." + "0" * decimal_places


def format_summary_metrics(sheet) -> None:
    """Format entropy and information values without changing stored precision."""
    for row in sheet.iter_rows(min_row=2, min_col=6, max_col=10):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = significant_decimal_number_format(cell.value)
    for cell in sheet.iter_cols(min_row=2, min_col=11, max_col=11):
        for value_cell in cell:
            if isinstance(value_cell.value, float):
                value_cell.number_format = significant_decimal_number_format(
                    value_cell.value
                )


def write_report(
    output_path: Path,
    source_path: Path,
    gene: str,
    min_subtype_sequences: int,
    summaries: list[dict[str, int | float | None]],
    eligibility_rows: list[dict[str, int | str | bool]],
) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Position_Summary"
    summary_headers = [
        ("Position", "RAS position"),
        ("CoveredSequencesAllSubtypes", "Coverage (all subtypes, sequences)"),
        (
            "AnalyzedSequencesEligibleSubtypes",
            "Analyzed sequences (non-ambiguous amino acid)",
        ),
        ("EligibleGenotypes", "Genotypes included (count)"),
        ("EligibleSubtypes", "Subtypes included (count)"),
        ("OverallEntropyBits", "Entropy (amino-acid uncertainty, bits)"),
        (
            "EntropyGivenGenotypeBits",
            "Entropy after genotype (remaining uncertainty, bits)",
        ),
        (
            "EntropyGivenGenotypeAndSubtypeBits",
            "Entropy after genotype + subtype (remaining uncertainty, bits)",
        ),
        (
            "GenotypeInformationBits",
            "Information from genotype (uncertainty reduced, bits)",
        ),
        (
            "AdditionalSubtypeInformationBits",
            "Additional information from subtype (after genotype, bits)",
        ),
        (
            "GenotypeConditionedUncertaintyResolvedBySubtypePct",
            "Uncertainty resolved by subtype (after genotype, %)",
        ),
    ]
    summary_sheet.append([label for _, label in summary_headers])
    for summary in summaries:
        summary_sheet.append([summary[column] for column, _ in summary_headers])
    style_sheet(summary_sheet)
    format_summary_metrics(summary_sheet)

    eligibility_sheet = workbook.create_sheet("Subtype_Eligibility")
    eligibility_headers = [
        ("Position", "RAS position"),
        ("Genotype", "Genotype"),
        ("Subtype", "Subtype"),
        ("CoveredSequences", "Coverage (sequences)"),
        ("AnalyzedSequences", "Analyzed sequences (non-ambiguous amino acid)"),
        ("Eligible", "Included in analysis"),
        ("ExclusionReason", "Reason not included"),
    ]
    eligibility_sheet.append([label for _, label in eligibility_headers])
    for row in eligibility_rows:
        eligibility_sheet.append([row[column] for column, _ in eligibility_headers])
    style_sheet(eligibility_sheet)

    metadata_sheet = workbook.create_sheet("Metadata")
    metadata_rows = [
        ("source_workbook", str(source_path.resolve())),
        ("gene", gene),
        ("analysis_unit", f"One {gene} RAS position"),
        ("population", "Final subtype complete-profile accessions"),
        (
            "primary_estimand",
            "Random accession; subtypes weighted by analyzed accession count",
        ),
        ("minimum_subtype_covered_sequences", min_subtype_sequences),
        ("excluded_amino_acids", ",".join(sorted(EXCLUDED_AAS))),
        ("metric_display_precision", "Two significant decimal digits"),
        ("overall_entropy", "H(A) = -sum_a P(A=a) log2 P(A=a)"),
        ("genotype_entropy", "H(A|G) = sum_g P(G=g) H(A|G=g)"),
        ("genotype_subtype_entropy", "H(A|G,S) = sum_g,s P(G=g,S=s) H(A|G=g,S=s)"),
        ("genotype_information", "I(A;G) = H(A) - H(A|G)"),
        ("additional_subtype_information", "I(A;S|G) = H(A|G) - H(A|G,S)"),
        ("subtype_percent_resolved", "100 * I(A;S|G) / H(A|G); blank when H(A|G)=0"),
    ]
    for row in metadata_rows:
        metadata_sheet.append(row)
    style_sheet(metadata_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> int:
    args = parse_args()
    if args.min_subtype_sequences < 1:
        raise ValueError("--min-subtype-sequences must be at least 1")
    source_path = Path(args.subtype_profile_workbook)
    output_path = Path(args.output_xlsx)
    positions = tuple(
        int(value) for value in args.positions.split(",") if value.strip()
    )
    if not positions:
        raise ValueError("--positions must contain at least one position")
    counts, coverage = load_profile_counts(
        source_path, args.position_column, set(positions)
    )
    summaries: list[dict[str, int | float | None]] = []
    eligibility_rows: list[dict[str, int | str | bool]] = []
    for position in positions:
        summary, rows = summarize_position(
            position, counts, coverage, args.min_subtype_sequences
        )
        summaries.append(summary)
        eligibility_rows.extend(rows)
    write_report(
        output_path,
        source_path,
        args.gene,
        args.min_subtype_sequences,
        summaries,
        eligibility_rows,
    )
    print(f"output_xlsx={output_path.resolve()}")
    print(f"minimum_subtype_covered_sequences={args.min_subtype_sequences}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
