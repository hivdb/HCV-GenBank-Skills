#!/usr/bin/env python3
"""Compare NS5A subtype profile consensus to translated subtype genome references."""

from __future__ import annotations
import argparse, json, re
from pathlib import Path
from Bio import Align, Seq
from openpyxl import Workbook, load_workbook

GENE_START_NA = 6258
VALID = set("ACDEFGHIKLMNPQRSTVWY*")
RAS_POSITIONS = (24, 26, 28, 29, 30, 31, 32, 38, 58, 62, 92, 93)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--subtype-profile-workbook", required=True)
    p.add_argument("--subtype-json", required=True)
    p.add_argument("--output-xlsx", required=True)
    p.add_argument("--positions", default=",".join(map(str, RAS_POSITIONS)))
    a = p.parse_args()
    positions = tuple(
        sorted({int(pos) for pos in a.positions.split(",") if pos.strip()})
    )
    refs = {}
    for r in json.loads(Path(a.subtype_json).read_text()):
        m = re.search(r"Genotype\s*(\d+[A-Za-z0-9]*)", str(r.get("genotypeName", "")))
        if m and m.group(1).lower() not in refs:
            offset = GENE_START_NA - int(r["firstNA"])
            nt = re.sub(r"[^ACGTRYSWKMBDHVN]", "N", str(r["sequence"])[offset:].upper())
            refs[m.group(1).lower()] = (
                r["accession"],
                str(Seq.Seq(nt[: len(nt) // 3 * 3]).translate()),
            )
    wb = load_workbook(a.subtype_profile_workbook, read_only=True, data_only=True)
    out = Workbook()
    o = out.active
    o.title = "subtype_reference_distance"
    o.append(
        [
            "Genotype",
            "Subtype",
            "ReferenceAccession",
            "ComparedAA",
            "Differences",
            "DifferencePositions",
            "Distance",
            "Status",
        ]
    )
    for sheet in wb.sheetnames:
        gt = sheet.removeprefix("GT")
        w = wb[sheet]
        h = [str(c.value or "") for c in next(w.iter_rows(max_row=1))]
        i = {v: n for n, v in enumerate(h)}
        pc = [x for x in h if x.endswith("Position")][0]
        calls = {}
        for r in w.iter_rows(min_row=2, values_only=True):
            st = str(r[i["Subtype"]] or "").lower()
            pos = r[i[pc]]
            aa = str(r[i["AminoAcid"]] or "")
            pct = float(r[i["PctWithAA"]] or 0)
            if (
                st
                and pos
                and (st, pos) not in calls
                or (st and pos and pct > calls[(st, pos)][0])
            ):
                calls[(st, pos)] = (pct, aa)
        for st in sorted({x[0] for x in calls}):
            if st not in refs:
                o.append([gt, st, "", "", "", "", "", "reference_not_found"])
                continue
            acc, ref = refs[st]
            q = "".join(calls.get((st, x), (0, "X"))[1] for x in range(1, len(ref) + 1))
            al = Align.PairwiseAligner(mode="global")
            z = al.align(ref, q)[0]
            rp = 0
            pairs = []
            for x, y in zip(str(z[0]), str(z[1])):
                if x != "-":
                    rp += 1
                if x != "-" and rp in positions and x in VALID and y in VALID:
                    pairs.append((rp, x, y))
            mismatches = [
                f"{pos}:{refaa_char}>{consensus_char}"
                for pos, refaa_char, consensus_char in pairs
                if refaa_char != consensus_char
            ]
            d = len(mismatches)
            o.append(
                [
                    gt,
                    st,
                    acc,
                    len(pairs),
                    d,
                    ";".join(mismatches),
                    d / len(pairs) if pairs else "",
                    "ok" if pairs else "no_comparable_positions",
                ]
            )
    wb.close()
    m = out.create_sheet("metadata")
    m.append(["aa_positions", ",".join(map(str, positions))])
    m.append(["alignment", "global AA alignment"])
    m.append(
        [
            "distance_definition",
            "AA differences / comparable aligned NS5A RAS positions",
        ]
    )
    Path(a.output_xlsx).parent.mkdir(parents=True, exist_ok=True)
    out.save(a.output_xlsx)


if __name__ == "__main__":
    main()
