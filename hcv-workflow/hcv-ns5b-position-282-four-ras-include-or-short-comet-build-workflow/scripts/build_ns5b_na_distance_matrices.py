#!/usr/bin/env python3
"""Build NS5A profile-accession paired AA or NA distance matrices."""

from __future__ import annotations
import argparse, csv
from collections import Counter, defaultdict
from pathlib import Path
from statistics import median
from openpyxl import Workbook, load_workbook

VALID_AA = set("ACDEFGHIKLMNPQRSTVWY")
VALID_NA = set("ACGT")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input-workbook", required=True)
    p.add_argument("--profile-accessions-csv", required=True)
    p.add_argument("--sequence-type", choices=("aa", "na"), required=True)
    p.add_argument("--positions", required=True)
    p.add_argument("--gt-output-xlsx", required=True)
    p.add_argument("--subtype-output-xlsx", required=True)
    p.add_argument("--min-subtype-sequences", type=int, default=10)
    a = p.parse_args()
    pos = tuple(map(int, a.positions.split(",")))
    valid = VALID_AA if a.sequence_type == "aa" else VALID_NA
    with open(a.profile_accessions_csv, encoding="utf-8-sig", newline="") as f:
        allowed = {
            r["accession"].strip()
            for r in csv.DictReader(f)
            if r.get("accession", "").strip()
        }
    wb = load_workbook(a.input_workbook, read_only=True, data_only=True)
    ws = wb.active
    h = [str(c.value or "") for c in next(ws.iter_rows(max_row=1))]
    ix = {v: i for i, v in enumerate(h)}
    col = "AASequence" if a.sequence_type == "aa" else "NASequence"
    groups = {"gt": defaultdict(list), "subtype": defaultdict(list)}
    excluded = Counter()
    for r in ws.iter_rows(min_row=2, values_only=True):
        ac = str(r[ix["AccessionID"]] or "").strip()
        gt = str(r[ix["ClosestGT"]] or "").strip()
        st = str(r[ix["ClosestSubtype"]] or "").strip().lower()
        start = r[ix["StartAAPosition"]]
        seq = str(r[ix[col]] or "").upper()
        if ac not in allowed:
            excluded["not_in_profile_accession_set"] += 1
            continue
        if not gt or not st:
            excluded["missing_genotype_or_subtype_assignment"] += 1
            continue
        if not start or not seq:
            excluded["missing_sequence_or_start_position"] += 1
            continue
        scale = 3 if a.sequence_type == "na" else 1
        offset = (int(start) - 1) * scale
        indices = tuple((x - 1) * scale + k for x in pos for k in range(scale))
        if offset > min(indices) or offset + len(seq) <= max(indices):
            excluded["does_not_cover_all_positions"] += 1
            continue
        values = tuple(seq[i - offset] for i in indices)
        if any(x not in valid for x in values):
            excluded["missing_or_ambiguous_call_at_position"] += 1
            continue
        groups["gt"][f"GT{gt}"].append(values)
        groups["subtype"][f"GT{gt}_{st}"].append(values)
    wb.close()

    def dist(x, y):
        d = n = 0
        for j in range(len(x[0])):
            cx = Counter(v[j] for v in x)
            cy = Counter(v[j] for v in y)
            total = len(x) * (len(x) - 1) // 2 if x is y else len(x) * len(y)
            same = (
                sum(v * (v - 1) // 2 for v in cx.values())
                if x is y
                else sum(cx[b] * cy[b] for b in valid)
            )
            d += total - same
            n += total
        return d / n if n else None

    def save(path, kind):
        out = Workbook()
        out.remove(out.active)
        blocks = (
            {"distance_matrix": groups[kind]} if kind == "gt" else defaultdict(dict)
        )
        if kind == "subtype":
            for lab, seqs in groups[kind].items():
                gt, _, st = lab.partition("_")
                blocks[gt][st] = seqs
        for name, block in sorted(
            blocks.items(),
            key=lambda item: (
                (
                    0,
                    int(item[0][2:]),
                )
                if item[0].startswith("GT") and item[0][2:].isdigit()
                else (1, item[0])
            ),
        ):
            keep = {
                k: v
                for k, v in block.items()
                if kind == "gt" or len(v) >= a.min_subtype_sequences
            }
            if keep:
                s = out.create_sheet(name)
                labels = sorted(keep)
                s.append(["Group", *labels])
                same_genotype, different_genotype = [], []
                for row_index, label_a in enumerate(labels):
                    row = [label_a]
                    for column_index, label_b in enumerate(labels):
                        value = dist(keep[label_a], keep[label_b])
                        if column_index < row_index:
                            row.append(None)
                        else:
                            row.append(value)
                            if value is not None:
                                if kind == "subtype":
                                    (same_genotype if row_index == column_index else different_genotype).append(value)
                                elif kind == "gt":
                                    (same_genotype if row_index == column_index else different_genotype).append(value)
                    s.append(row)
                if kind == "gt":
                    s.append([])
                    s.append(["Distance comparison", "Mean", "Median", "Pair count"])
                    for label, distances in (("Same genotype", same_genotype), ("Different genotype", different_genotype)):
                        s.append([label, sum(distances) / len(distances) if distances else None, median(distances) if distances else None, len(distances)])
                else:
                    s.append([])
                    s.append(["Distance comparison", "Mean", "Median", "Pair count"])
                    for label, distances in (("Same subtype", same_genotype), ("Different subtype", different_genotype)):
                        s.append([label, sum(distances) / len(distances) if distances else None, median(distances) if distances else None, len(distances)])
                for row in s.iter_rows(min_row=2, min_col=2):
                    for c in row:
                        c.number_format = "0.0%"
        for name, rows in [
            ("sequence_counts", [(k, len(v)) for k, v in sorted(groups[kind].items())]),
            (
                "excluded_sequences",
                list(sorted(excluded.items()))
                + [("total_excluded", sum(excluded.values()))],
            ),
        ]:
            s = out.create_sheet(name)
            s.append(
                ["Group" if name == "sequence_counts" else "Reason", "SequenceCount"]
            )
            [s.append(r) for r in rows]
        s = out.create_sheet("metadata")
        s.append(["sequence_type", a.sequence_type])
        s.append(["aa_positions", a.positions])
        s.append(
            [
                "distance_definition",
                "mean paired differences across complete comparison positions",
            ]
        )
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        out.save(path)

    save(a.gt_output_xlsx, "gt")
    save(a.subtype_output_xlsx, "subtype")


if __name__ == "__main__":
    main()
