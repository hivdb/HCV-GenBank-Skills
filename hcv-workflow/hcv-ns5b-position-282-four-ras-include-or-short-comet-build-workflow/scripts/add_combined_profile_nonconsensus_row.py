#!/usr/bin/env python3
"""Create an annotated combined profile with MeanDiff and PositionDiff values."""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

NORMAL_AAS = set("ACDEFGHIKLMNPQRSTVWY")
SUBTYPE_RE = re.compile(r"^GT(?P<gt>\d+)_(?P<subtype>\S+) \(")
GT_RE = re.compile(r"^GT(?P<gt>\d+)(?:\s|_)")
VARIANT_RE = re.compile(r"([A-Z*])(\d+(?:\.\d+)?)")


def fasta(path: Path) -> dict[str, str]:
    result: dict[str, str] = {}
    header = None
    parts: list[str] = []
    for raw in path.read_text().splitlines():
        line = raw.strip()
        if line.startswith(">"):
            if header:
                result[header] = "".join(parts)
            header, parts = line[1:].split()[0].removeprefix("GT"), []
        elif line:
            parts.append(line.upper())
    if header:
        result[header] = "".join(parts)
    return result


def most_frequent_variant(value: object) -> tuple[str, str] | None:
    variants = VARIANT_RE.findall(str(value or ""))
    return max(variants, key=lambda item: float(item[1])) if variants else None


def add_mean_diff_column(sheet) -> None:
    """Add MeanDiff for subtype rows using the preceding genotype consensus row."""
    for column in range(sheet.max_column, 0, -1):
        if sheet.cell(1, column).value == "MeanDiff":
            sheet.delete_cols(column)
    mean_diff_column = sheet.max_column + 1
    sheet.cell(1, mean_diff_column, "MeanDiff")
    genotype_variants: list[tuple[str, str] | None] | None = None
    for row in range(2, sheet.max_row + 1):
        label = str(sheet.cell(row, 1).value or "")
        if GT_RE.match(label) and not SUBTYPE_RE.match(label):
            genotype_variants = [
                most_frequent_variant(sheet.cell(row, column).value)
                for column in range(2, mean_diff_column)
            ]
        elif SUBTYPE_RE.match(label) and genotype_variants is not None:
            displayed_percent = sum(
                float(frequency)
                for column, gt_variant in zip(
                    range(2, mean_diff_column), genotype_variants
                )
                for amino_acid, frequency in VARIANT_RE.findall(
                    str(sheet.cell(row, column).value or "")
                )
                if gt_variant is None or amino_acid != gt_variant[0]
            )
            cell = sheet.cell(row, mean_diff_column, displayed_percent / 100.0)
            cell.number_format = "0.0"


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--combined-profile-workbook", required=True)
    p.add_argument("--output-workbook", required=True)
    p.add_argument("--profile-input-workbook", required=True)
    p.add_argument("--profile-accessions-csv", required=True)
    p.add_argument("--genotype-consensus-fasta", required=True)
    a = p.parse_args()
    with open(a.profile_accessions_csv, newline="", encoding="utf-8-sig") as f:
        allowed = {
            (
                r["accession"].strip(),
                r["genotype"].strip(),
                r["subtype"].strip().lower(),
            )
            for r in csv.DictReader(f)
            if r.get("accession", "").strip()
        }
    out = load_workbook(a.combined_profile_workbook, rich_text=True)
    sheet = out.active
    add_mean_diff_column(sheet)
    positions = [
        int(str(c.value)[1:])
        for c in sheet[1]
        if re.fullmatch(r"P\d+", str(c.value or ""))
    ]
    included = {
        (m.group("gt"), m.group("subtype").lower())
        for (value,) in sheet.iter_rows(min_row=2, max_col=1, values_only=True)
        if (m := SUBTYPE_RE.match(str(value or "")))
    }
    references = fasta(Path(a.genotype_consensus_fasta))
    numerator, denominator = defaultdict(int), defaultdict(int)
    source_wb = load_workbook(a.profile_input_workbook, read_only=True, data_only=True)
    source = source_wb.active
    headers = [str(c.value or "") for c in next(source.iter_rows(min_row=1, max_row=1))]
    ix = {v: i for i, v in enumerate(headers)}
    for row in source.iter_rows(min_row=2, values_only=True):
        accession = str(row[ix["AccessionID"]] or "").strip()
        gt = str(row[ix["ClosestGT"]] or "").strip().removeprefix("GT")
        subtype = str(row[ix["ClosestSubtype"]] or "").strip().lower()
        if (accession, gt, subtype) not in allowed or (gt, subtype) not in included:
            continue
        if (
            "AlignmentQCStatus" in ix
            and str(row[ix["AlignmentQCStatus"]] or "").strip() != "PASS"
        ):
            continue
        start, sequence = (
            row[ix["StartAAPosition"]],
            str(row[ix["AASequence"]] or "").strip().upper(),
        )
        if not start or gt not in references:
            continue
        for pos in positions:
            offset = pos - int(start)
            if not 0 <= offset < len(sequence):
                continue
            aa = sequence[offset]
            if aa not in NORMAL_AAS:
                continue
            denominator[pos] += 1
            if aa != references[gt][pos - 1]:
                numerator[pos] += 1
    source_wb.close()
    for row in range(sheet.max_row, 1, -1):
        if sheet.cell(row, 1).value in {"NonConsensusAAFraction", "PositionDiff"}:
            sheet.delete_rows(row)
    sheet.append(
        [
            "PositionDiff",
            *[
                numerator[pos] / denominator[pos] if denominator[pos] else None
                for pos in positions
            ],
            None,
        ]
    )
    for cell in sheet[sheet.max_row]:
        cell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        cell.font = Font(bold=True)
        cell.number_format = "0.0%"
    output_path = Path(a.output_workbook)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    out.save(output_path)
    print(f"annotated_combined_profile={output_path}")


if __name__ == "__main__":
    main()
