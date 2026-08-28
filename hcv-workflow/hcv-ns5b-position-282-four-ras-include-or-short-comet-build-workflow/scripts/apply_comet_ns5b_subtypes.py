#!/usr/bin/env python3
"""Replace NS5B subtype assignments in a workbook with authoritative Comet calls."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--subtype-workbook", required=True)
    parser.add_argument("--comet-subtype-csv", required=True)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    assignments: dict[str, str] = {}
    with Path(args.comet_subtype_csv).open(encoding="utf-8", newline="") as handle:
        for row in csv.DictReader(handle):
            if row.get("accession") and row.get("subtype"):
                assignments[row["accession"]] = row["subtype"]

    workbook = load_workbook(args.subtype_workbook)
    worksheet = workbook.active
    headers = {cell.value: cell.column for cell in worksheet[1]}
    required = {
        "AccessionID",
        "ClosestSubtype",
        "ClosestSubtypeAssignmentSource",
        "ClosestSubtypeMetadataColumn",
    }
    if not required <= set(headers):
        raise RuntimeError(
            f"Missing subtype columns: {sorted(required - set(headers))}"
        )
    updated = missing = 0
    for row in range(2, worksheet.max_row + 1):
        accession = str(worksheet.cell(row, headers["AccessionID"]).value or "")
        subtype = assignments.get(accession)
        if not subtype:
            missing += 1
            continue
        worksheet.cell(row, headers["ClosestSubtype"]).value = subtype
        worksheet.cell(row, headers["ClosestSubtypeAssignmentSource"]).value = "Comet"
        worksheet.cell(
            row, headers["ClosestSubtypeMetadataColumn"]
        ).value = "Comet NS5B"
        updated += 1
    workbook.save(args.subtype_workbook)
    print(f"comet_subtype_assignments_applied={updated}")
    print(f"comet_subtype_assignments_missing={missing}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
