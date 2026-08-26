#!/usr/bin/env python3
"""Merge NS5A subtype complete-profile worksheets into one simplified table."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook


COLUMNS = ["Subtype", "NS5APosition", "NumSeqsIncludingPosition", "AminoAcid", "CountWithAA", "PctWithAA"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-workbook", required=True)
    parser.add_argument("--output-workbook", required=True)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    source = load_workbook(args.input_workbook, read_only=True, data_only=True)
    source_sheet_count = len(source.sheetnames)
    output = Workbook()
    sheet = output.active
    sheet.title = "NS5A_Subtype_CompleteProfiles"
    sheet.append(COLUMNS)
    merged_rows = 0
    for source_sheet in source.worksheets:
        header = [str(value or "") for value in next(source_sheet.iter_rows(values_only=True))]
        index = {name: position for position, name in enumerate(header)}
        missing = [name for name in COLUMNS if name not in index]
        if missing:
            raise RuntimeError(f"Columns missing from worksheet {source_sheet.title}: {', '.join(missing)}")
        for values in source_sheet.iter_rows(min_row=2, values_only=True):
            sheet.append([values[index[name]] for name in COLUMNS])
            merged_rows += 1
    source.close()
    output_path = Path(args.output_workbook)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output.save(output_path)
    print(json.dumps({"output_workbook": str(output_path.resolve()), "source_sheet_count": source_sheet_count, "merged_row_count": merged_rows}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
