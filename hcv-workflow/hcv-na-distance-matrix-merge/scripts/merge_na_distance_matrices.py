#!/usr/bin/env python3
"""Merge selected active COMET NA distance matrices into paper-ready workbooks."""

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_OUTPUT_DIR = REPO_ROOT / "outputs/hcv-na-distance-matrix-merge"
SOURCES = (
    (
        "NS3 One RAS",
        REPO_ROOT
        / "outputs/comet-NS3-one-ras/29_build-paired-distance-matrices/NS3_GT_NA_Distance_RAS.xlsx",
        REPO_ROOT
        / "outputs/comet-NS3-one-ras/29_build-paired-distance-matrices/NS3_Subtype_NA_Distance_RAS.xlsx",
    ),
    (
        "NS5A One RAS",
        REPO_ROOT
        / "outputs/comet-NS5A-one-ras/29_build-paired-distance-matrices/NS5A_GT_NA_Distance_RAS.xlsx",
        REPO_ROOT
        / "outputs/comet-NS5A-one-ras/29_build-paired-distance-matrices/NS5A_Subtype_NA_Distance_RAS.xlsx",
    ),
    (
        "NS5B Position 282 Four RAS",
        REPO_ROOT
        / "outputs/comet-NS5B-position-282-four-ras/29_build-paired-distance-matrices/NS5B_GT_NA_Distance_RAS.xlsx",
        REPO_ROOT
        / "outputs/comet-NS5B-position-282-four-ras/29_build-paired-distance-matrices/NS5B_Subtype_NA_Distance_RAS.xlsx",
    ),
)


def copy_block(source, destination, start_column: int, title: str) -> int:
    """Copy a worksheet block, including cell formatting, and return its ending column."""
    title_cell = destination.cell(1, start_column, title)
    title_font = copy(source["A1"].font)
    title_font.bold = True
    title_cell.font = title_font
    for row in source.iter_rows():
        for cell in row:
            target = destination.cell(cell.row + 1, start_column + cell.column - 1, cell.value)
            if cell.has_style:
                target._style = copy(cell._style)
            if cell.number_format:
                target.number_format = cell.number_format
    for column, dimension in source.column_dimensions.items():
        destination.column_dimensions[
            destination.cell(1, start_column + source[column][0].column - 1).column_letter
        ].width = dimension.width
    return start_column + source.max_column - 1


def relevant_subtype_sheets(workbook) -> list[str]:
    return [
        name
        for name in workbook.sheetnames
        if name.startswith("GT") and not name.endswith("_counts")
    ]


def merge_gt(output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "GT_NA_Distance"
    column = 1
    for title, gt_path, _ in SOURCES:
        source_book = load_workbook(gt_path, data_only=False)
        try:
            column = copy_block(source_book["distance_matrix"], worksheet, column, title) + 2
        finally:
            source_book.close()
    workbook.save(output_path)


def merge_subtypes(output_path: Path) -> None:
    source_books = [(title, load_workbook(path, data_only=False)) for title, _, path in SOURCES]
    try:
        genotypes = sorted({name for _, book in source_books for name in relevant_subtype_sheets(book)})
        workbook = Workbook()
        workbook.remove(workbook.active)
        for genotype in genotypes:
            worksheet = workbook.create_sheet(genotype)
            column = 1
            for title, source_book in source_books:
                if genotype in source_book.sheetnames:
                    column = copy_block(source_book[genotype], worksheet, column, title) + 2
        workbook.save(output_path)
    finally:
        for _, workbook in source_books:
            workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    gt_output = args.output_dir / "GT_NA_Distance_Merged.xlsx"
    subtype_output = args.output_dir / "Subtype_NA_Distance_Merged.xlsx"
    merge_gt(gt_output)
    merge_subtypes(subtype_output)
    print(f"gt_output={gt_output.resolve()}")
    print(f"subtype_output={subtype_output.resolve()}")


if __name__ == "__main__":
    main()
