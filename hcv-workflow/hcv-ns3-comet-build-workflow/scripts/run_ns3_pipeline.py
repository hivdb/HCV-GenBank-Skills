#!/usr/bin/env python3
"""Run named stages of the HCV NS3 COMET build workflow.

Examples:
  scripts/run_ns3_pipeline.py --list-steps
  scripts/run_ns3_pipeline.py
  scripts/run_ns3_pipeline.py --step discover-refid-fastas
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import subprocess
import sys
import tomllib
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path


SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parents[2]
SKILL_NAME = "hcv-ns3-comet-build-workflow"
RAS_POSITIONS = "36,41,43,54,55,56,80,122,155,156,158,166,168,170,175"
RANGE_POSITIONS = ",".join(str(position) for position in range(36, 176))
STEP_NAMES = (
    "prepare-workdirs", "discover-refid-fastas", "stage-refid-fastas",
    "prepare-comet-assignments", "select-noncomet-priority-assignments",
    "filter-accession-metadata", "split-refid-metadata", "filter-refid-fastas",
    "build-genotype-workbook", "add-genotype-counts",
    "build-subtype-workbook", "extract-profile-aa", "validate-profile-alignment",
    "summarize-qc-mutation-burden", "report-profile-input-counts", "build-complete-profiles",
    "merge-subtype-complete-profiles",
    "export-noncomet-priority-accessions", "export-consensus-fastas",
    "align-subtype-consensus-to-gt1a", "compare-reference-consensus",
    "build-genotype-ras-profile", "build-subtype-ras-profile", "build-combined-ras-reports",
    "add-nonconsensus-row", "summarize-subtype-ras-differences",
    "build-genotype-aa-consensus-distance", "build-subtype-aa-consensus-distance",
    "build-paired-distance-matrices", "build-subtype-profile-coverage", "build-ras-entropy",
    "publish-ictv-report",
    "audit-gt7-gt8-sequences",
    "compare-gt7-gt8-local-assignments",
)
STEP_ORDER = {name: number for number, name in enumerate(STEP_NAMES, start=1)}


@dataclass(frozen=True)
class Step:
    name: str
    description: str
    action: Callable[[], None]

    @property
    def order(self) -> int:
        return STEP_ORDER[self.name]


def parse_dotenv(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.is_file():
        return values
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[7:].lstrip()
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
            value = value[1:-1]
        if key:
            values[key] = value
    return values


def load_config(config_path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if config_path.is_file():
        with config_path.open("rb") as handle:
            config = tomllib.load(handle)
        for section_name in ("common", "ns3_comet"):
            section = config.get(section_name, {})
            if isinstance(section, dict):
                values.update({key.upper(): str(value) for key, value in section.items()})
    return values


def path_value(value: str, *, default: Path | None = None) -> Path:
    path = Path(value).expanduser() if value else default
    if path is None:
        raise RuntimeError("A required path value is missing")
    if path.is_absolute():
        return path
    repository_path = REPO_ROOT / path
    hcvdata_path = REPO_ROOT / "HCVData" / path
    # The shared TOML stores reference-data filenames without an HCVData/ prefix.
    # Prefer an explicit repository-relative path, then resolve those data files under HCVData/.
    if not repository_path.exists() and hcvdata_path.exists():
        return hcvdata_path
    return repository_path


def executable_value(value: str) -> str:
    candidate = Path(value).expanduser()
    if candidate.is_absolute() or "/" in value:
        return str(candidate if candidate.is_absolute() else REPO_ROOT / candidate)
    return value


def repository_relative(value: str | Path) -> str:
    path = Path(value).expanduser()
    if not path.is_absolute():
        return str(path)
    try:
        return str(path.relative_to(REPO_ROOT))
    except ValueError:
        return str(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--list-steps", action="store_true", help="Print available step names and exit")
    parser.add_argument("--step", action="append", help="Run only this named step; repeat to run multiple steps")
    parser.add_argument("--config", type=Path, default=REPO_ROOT / "pipeline.local.toml")
    parser.add_argument("--excel-file")
    parser.add_argument("--sheet-name")
    parser.add_argument("--fasta-pool")
    parser.add_argument("--genbank-dir")
    parser.add_argument("--output-dir")
    parser.add_argument("--reference-fasta")
    parser.add_argument("--subtype-json")
    parser.add_argument("--gt-aa-json")
    parser.add_argument("--accessions-metadata-csv")
    parser.add_argument("--comet-subtyping-csv")
    parser.add_argument("--noncomet-coverage-csv")
    parser.add_argument("--temp-root")
    parser.add_argument("--python-bin")
    return parser.parse_args()


class Pipeline:
    def __init__(self, args: argparse.Namespace) -> None:
        dotenv = parse_dotenv(REPO_ROOT / ".env")
        config = load_config(args.config.expanduser())
        values = {**config, **dotenv, **os.environ}

        def value(argument: str, environment: str, default: str = "") -> str:
            supplied = getattr(args, argument)
            return supplied if supplied is not None else values.get(environment, default)

        self.python_bin = executable_value(value("python_bin", "PYTHON_BIN", str(REPO_ROOT / ".venv/bin/python")))
        self.excel_file = path_value(value("excel_file", "EXCEL_FILE"))
        self.sheet_name = value("sheet_name", "SHEET_NAME")
        self.fasta_pool = path_value(value("fasta_pool", "FASTA_POOL"))
        self.genbank_dir = path_value(value("genbank_dir", "GENBANK_DIR"))
        self.output_dir = path_value(value("output_dir", "OUTPUT_DIR", "outputs/comet-NS3"))
        self.reference_fasta = path_value(value("reference_fasta", "REFERENCE_FASTA", "HCVData/HCV_GT_RefSeqs.fasta"))
        self.subtype_json = path_value(value("subtype_json", "SUBTYPE_JSON", "HCVData/HCV_Subtype_Refs_By_Genome_NA.json"))
        self.gt_aa_json = path_value(value("gt_aa_json", "GT_AA_JSON", "HCVData/HCV_GT_Refs_By_Gene_AA.json"))
        self.metadata_csv = path_value(value("accessions_metadata_csv", "ACCESSIONS_METADATA_CSV", "HCVData/Accessions_metadata.csv"))
        self.comet_csv = path_value(value("comet_subtyping_csv", "COMET_SUBTYPING_CSV", "HCVData/HCV-all-seq-subtype/all_comet_subtype.csv"))
        self.noncomet_coverage_csv = path_value(value("noncomet_coverage_csv", "NONCOMET_COVERAGE_CSV", "HCVData/HCV-all-seq-subtype/NS3_AllSeq_NonComet_Coverage.csv"))
        self.reference_subtypes_csv = REPO_ROOT / "HCVData/Reference_seqs/HCV_Subtype_Refs_AA_Accession_Subtype.csv"
        self.temp_root = path_value(value("temp_root", "TEMP_ROOT", str(self.step_dir("prepare-comet-assignments"))))
        self.discovery_tmp = self.step_dir("discover-refid-fastas")
        self.matched_txt = self.discovery_tmp / "NS3_matched_fasta_files.txt"
        self.staged_fasta_dir = self.step_dir("stage-refid-fastas") / "included_refid_fastas"
        self.metadata_dir = self.step_dir("filter-accession-metadata")
        self.refid_metadata_dir = self.step_dir("split-refid-metadata") / "refid_metadata"
        self.filtered_fasta_dir = self.step_dir("filter-refid-fastas") / "included_refid_fastas"
        self.kept_accessions_csv = self.step_dir("filter-refid-fastas") / "kept_accessions.csv"
        self.included_fasta_dir = self.step_dir("prepare-comet-assignments") / "included_refid_fastas"
        self.comet_subtype_assignments_csv = self.temp_root / "comet_ns3_subtype_assignments.csv"
        self.priority_assignments_csv = self.step_dir("select-noncomet-priority-assignments") / "NS3_NonComet_Priority_Assignments.csv"
        self.aa_workbook = self.step_dir("validate-profile-alignment") / "NS3_Profile_Input_Alignment_QC.xlsx"
        self.profile_accessions_csv = self.step_dir("build-complete-profiles") / "NS3_Profile_Accessions_QC_Pass.csv"

    def step_dir(self, name: str) -> Path:
        return self.output_dir / f"{STEP_ORDER[name]:02d}_{name}"

    def ensure_summary_directories(self) -> None:
        self.output_dir.mkdir(parents=True, exist_ok=True)
        for name in STEP_NAMES:
            self.step_dir(name).mkdir(parents=True, exist_ok=True)

    def run(
        self,
        script: str,
        *arguments: str,
        stdout_path: Path | None = None,
        stream_output: bool = False,
    ) -> str:
        command = [self.python_bin, repository_relative(SCRIPT_DIR / script), *(repository_relative(argument) for argument in arguments)]
        environment = {**os.environ, "NS3_STEP_OUTPUT_DIR": str(self.current_step_dir)}
        if stdout_path is None:
            subprocess.run(command, check=True, cwd=REPO_ROOT, env=environment)
            return ""
        stdout_path.parent.mkdir(parents=True, exist_ok=True)
        if stream_output:
            output_lines: list[str] = []
            with stdout_path.open("w", encoding="utf-8") as handle:
                process = subprocess.Popen(command, cwd=REPO_ROOT, env=environment, text=True, stdout=subprocess.PIPE)
                assert process.stdout is not None
                for line in process.stdout:
                    print(line, end="")
                    handle.write(line)
                    output_lines.append(line)
                if process.wait() != 0:
                    raise subprocess.CalledProcessError(process.returncode, command)
            return "".join(output_lines)
        try:
            result = subprocess.run(command, check=True, text=True, capture_output=True, cwd=REPO_ROOT, env=environment)
        except subprocess.CalledProcessError as error:
            if error.stdout:
                print(error.stdout, end="", file=sys.stderr)
            if error.stderr:
                print(error.stderr, end="", file=sys.stderr)
            raise
        stdout_path.write_text(result.stdout, encoding="utf-8")
        if result.stderr:
            print(result.stderr, end="", file=sys.stderr)
        return result.stdout

    def announce(self, order: int, name: str, description: str) -> None:
        print(f"\n## Step {order}: {name}\n{description}")

    @staticmethod
    def fasta_accessions(directory: Path) -> set[str]:
        accessions: set[str] = set()
        if directory.is_dir():
            for path in directory.glob("*.fasta"):
                for line in path.read_text(encoding="utf-8").splitlines():
                    if line.startswith(">") and (accession := line[1:].strip().split(maxsplit=1)[0]):
                        accessions.add(accession)
        return accessions

    def print_accession_counts(self, input_accessions: set[str]) -> None:
        included_accessions: set[str] = set()
        if self.current_step_order >= STEP_ORDER["build-complete-profiles"] and self.profile_accessions_csv.is_file():
            with self.profile_accessions_csv.open(encoding="utf-8", newline="") as handle:
                for row in csv.DictReader(handle):
                    accession = (row.get("accession") or row.get("AccessionID") or row.get("Accession") or "").strip()
                    if accession:
                        included_accessions.add(accession)
        if not included_accessions:
            output_dir = self.filtered_fasta_dir if self.current_step_order >= STEP_ORDER["filter-refid-fastas"] else self.included_fasta_dir
            included_accessions = self.fasta_accessions(output_dir)
        if not included_accessions:
            included_accessions = input_accessions
        print(f"Input accessions: {len(input_accessions)}")
        print(f"Excluded accessions: {len(input_accessions - included_accessions)}")
        print(f"Final included accessions: {len(included_accessions)}")

    def print_gt7_gt8_subtype_counts(self) -> None:
        """Report current GT7/GT8 subtype diversity and accession counts."""
        counts: dict[str, tuple[set[str], int]] = {"7": (set(), 0), "8": (set(), 0)}
        stage = ""
        subtype_workbook = self.step_dir("build-subtype-workbook") / "NS3_Subtype_AllStudies_WSeqs.xlsx"
        if self.current_step_order >= STEP_ORDER["build-complete-profiles"] and self.profile_accessions_csv.is_file():
            stage = "complete-profile eligibility"
            with self.profile_accessions_csv.open(encoding="utf-8", newline="") as handle:
                for row in csv.DictReader(handle):
                    genotype = (row.get("genotype") or "").strip().removeprefix("GT")
                    subtype = (row.get("subtype") or "").strip().lower()
                    if genotype in counts and subtype:
                        counts[genotype][0].add(subtype)
                        counts[genotype] = (counts[genotype][0], counts[genotype][1] + 1)
        elif self.current_step_order >= STEP_ORDER["validate-profile-alignment"] and self.aa_workbook.is_file():
            stage = "alignment QC"
            from openpyxl import load_workbook
            workbook = load_workbook(self.aa_workbook, read_only=True, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            header = [str(value or "") for value in next(sheet.iter_rows(values_only=True))]
            index = {name: position for position, name in enumerate(header)}
            for values in sheet.iter_rows(min_row=2, values_only=True):
                if str(values[index["AlignmentQCStatus"]] or "").strip() != "PASS":
                    continue
                genotype = str(values[index["ClosestGT"]] or "").strip().removeprefix("GT")
                subtype = str(values[index["ClosestSubtype"]] or "").strip().lower()
                if genotype in counts and subtype:
                    counts[genotype][0].add(subtype)
                    counts[genotype] = (counts[genotype][0], counts[genotype][1] + 1)
            workbook.close()
        elif self.current_step_order >= STEP_ORDER["extract-profile-aa"]:
            extraction_workbook = self.step_dir("extract-profile-aa") / "NS3_Profile_Input_Source.xlsx"
            if extraction_workbook.is_file():
                stage = "amino-acid extraction"
                from openpyxl import load_workbook
                workbook = load_workbook(extraction_workbook, read_only=True, data_only=True)
                sheet = workbook[workbook.sheetnames[0]]
                header = [str(value or "") for value in next(sheet.iter_rows(values_only=True))]
                index = {name: position for position, name in enumerate(header)}
                for values in sheet.iter_rows(min_row=2, values_only=True):
                    genotype = str(values[index["ClosestGT"]] or "").strip().removeprefix("GT")
                    subtype = str(values[index["ClosestSubtype"]] or "").strip().lower()
                    if genotype in counts and subtype:
                        counts[genotype][0].add(subtype)
                        counts[genotype] = (counts[genotype][0], counts[genotype][1] + 1)
                workbook.close()
        elif subtype_workbook.is_file():
            stage = "COMET and priority assignments"
            from openpyxl import load_workbook

            workbook = load_workbook(subtype_workbook, read_only=True, data_only=True)
            sheet = workbook[workbook.sheetnames[0]]
            header = [str(value or "") for value in next(sheet.iter_rows(values_only=True))]
            index = {name: position for position, name in enumerate(header)}
            for values in sheet.iter_rows(min_row=2, values_only=True):
                genotype = str(values[index["ClosestGT"]] or "").strip().removeprefix("GT")
                subtype = str(values[index["ClosestSubtype"]] or "").strip().lower()
                if genotype in counts and subtype:
                    counts[genotype][0].add(subtype)
                    counts[genotype] = (counts[genotype][0], counts[genotype][1] + 1)
            workbook.close()
        elif self.comet_subtype_assignments_csv.is_file():
            stage = "COMET and priority assignments"
            comet_assignments: dict[str, tuple[str, str]] = {}
            with self.comet_subtype_assignments_csv.open(encoding="utf-8-sig", newline="") as handle:
                for row in csv.DictReader(handle):
                    accession = (row.get("accession") or "").strip()
                    if accession:
                        comet_assignments[accession.split(".", 1)[0]] = (
                            (row.get("genotype") or "").strip().removeprefix("GT"),
                            (row.get("subtype") or "").strip().lower(),
                        )
            priority_assignments: dict[str, tuple[str, str]] = {}
            if self.priority_assignments_csv.is_file():
                with self.priority_assignments_csv.open(encoding="utf-8-sig", newline="") as handle:
                    for row in csv.DictReader(handle):
                        accession = (row.get("Accession") or "").strip()
                        if accession:
                            priority_assignments[accession.split(".", 1)[0]] = (
                                (row.get("ClosestGenotype") or "").strip().removeprefix("GT"),
                                (row.get("ClosestSubtype") or "").strip().lower(),
                            )
            fasta_dir = self.filtered_fasta_dir if self.current_step_order >= STEP_ORDER["filter-refid-fastas"] else self.included_fasta_dir
            seen_accessions: set[str] = set()
            for accession in self.fasta_accessions(fasta_dir):
                accession_key = accession.split(".", 1)[0]
                assignment = priority_assignments.get(accession_key) or comet_assignments.get(accession_key)
                if assignment is None:
                    continue
                genotype, subtype = assignment
                if genotype in counts and subtype:
                    counts[genotype][0].add(subtype)
                    counts[genotype] = (counts[genotype][0], counts[genotype][1] + 1)
                seen_accessions.add(accession_key)
            for accession_key, (genotype, subtype) in priority_assignments.items():
                if accession_key not in seen_accessions and genotype in counts and subtype:
                    counts[genotype][0].add(subtype)
                    counts[genotype] = (counts[genotype][0], counts[genotype][1] + 1)
        for genotype in ("7", "8"):
            subtypes, accessions = counts[genotype]
            print(f"GT{genotype} after {stage or 'assignment'}: {len(subtypes)} subtypes ({accessions} accessions)")

    def steps(self) -> list[Step]:
        summary = lambda step, filename="last_run_summary.json": self.step_dir(step) / filename
        gt_workbook = self.step_dir("build-genotype-workbook") / "NS3_GT_AllStudies.xlsx"
        subtype_workbook = self.step_dir("build-subtype-workbook") / "NS3_Subtype_AllStudies_WSeqs.xlsx"
        gt_profile = self.step_dir("build-complete-profiles") / "NS3_GT_CompleteProfiles_TabsPerGT.xlsx"
        subtype_profile = self.step_dir("build-complete-profiles") / "NS3_Subtype_CompleteProfiles_TabsPerGT.xlsx"
        gt_consensus = self.step_dir("export-consensus-fastas") / "NS3_GT_Consensus.fasta"
        subtype_consensus = self.step_dir("export-consensus-fastas") / "NS3_Subtype_Consensus.fasta"
        gt_ras = self.step_dir("build-genotype-ras-profile") / "NS3_GT_RAS_Profiles.xlsx"
        subtype_ras = self.step_dir("build-subtype-ras-profile") / "NS3_Subtype_RAS_Profiles.xlsx"
        explicit_subtype_ras = self.step_dir("build-subtype-ras-profile") / "NS3_Subtype_RAS_Profiles_Explicit_AA.xlsx"
        combined_ras = self.step_dir("build-combined-ras-reports") / "NS3_Combined_RAS_Profiles.xlsx"
        annotated_combined_ras = self.step_dir("add-nonconsensus-row") / "NS3_Combined_RAS_Profiles_Annotated.xlsx"
        comet_gt_csv = self.temp_root / "comet_ns3_genotype_assignments.csv"
        comet_subtype_csv = self.comet_subtype_assignments_csv

        def prepare() -> None:
            self.run("prepare_ns3_pipeline_workdirs.py", "--clean-dir", self.output_dir)

        def discover() -> None:
            output = self.run("find_refid_fastas.py", "--excel-file", self.excel_file, "--sheet", self.sheet_name, "--fasta-dir", self.fasta_pool, "--output-dir", self.discovery_tmp, stdout_path=self.discovery_tmp / "discovery_ns3.json", stream_output=True)
            del output
            matches = sorted(self.discovery_tmp.glob("refid_fasta_*/matched_fasta_files.txt"))
            if len(matches) != 1:
                raise RuntimeError(f"Expected one matched FASTA file list under {self.discovery_tmp}, found {len(matches)}")
            self.matched_txt.write_text(matches[0].read_text(encoding="utf-8"), encoding="utf-8")

        def prepare_comet() -> None:
            shutil.copytree(self.staged_fasta_dir, self.included_fasta_dir, dirs_exist_ok=True)
            self.run("prepare_comet_ns3_assignments.py", "--comet-csv", self.comet_csv, "--fasta-dir", self.included_fasta_dir, "--genotype-output-csv", comet_gt_csv, "--subtype-output-csv", comet_subtype_csv, "--not-found-output-csv", self.temp_root / "comet_ns3_not_found_or_unassigned.csv", "--not-found-fasta-output", self.temp_root / "comet_ns3_not_found_or_unassigned.fasta", "--remove-unassigned")

        def extract_aa() -> None:
            step_dir = self.step_dir("extract-profile-aa")
            self.run("build_ns3_subtype_with_gt_aa.py", "--subtype-workbook", subtype_workbook, "--fasta-dir", self.fasta_pool, "--gt-aa-json", self.gt_aa_json, "--output-dir", step_dir, "--output-workbook", step_dir / "NS3_Profile_Input_Source.xlsx", stdout_path=summary("extract-profile-aa"), stream_output=True)

        def report_profile_counts() -> None:
            output = self.run("build_ns3_completeprofiles_tabspergt.py", "--input-workbook", self.aa_workbook, "--report-only", stdout_path=self.step_dir("report-profile-input-counts") / "profile_input_counts.json")
            print(output, end="")

        def paired_distances() -> None:
            for sequence_type in ("aa", "na"):
                upper = sequence_type.upper()
                script = f"build_ns3_{sequence_type}_distance_matrices.py"
                for suffix, range_arguments in (("RAS", ()), ("Pos36_175", ("--start", "36", "--end", "175"))):
                    step_dir = self.step_dir("build-paired-distance-matrices")
                    self.run(script, "--input-workbook", self.aa_workbook, "--profile-accessions-csv", self.profile_accessions_csv, "--gt-output-xlsx", step_dir / f"NS3_GT_{upper}_Distance_{suffix}.xlsx", "--subtype-output-xlsx", step_dir / f"NS3_Subtype_{upper}_Distance_{suffix}.xlsx", "--min-subtype-sequences", "10", *range_arguments)

        return [
            Step("prepare-workdirs", "recreate the NS3 COMET output and run directories", prepare),
            Step("discover-refid-fastas", "discover selected RefID FASTA files", discover),
            Step("stage-refid-fastas", "copy discovered FASTA files into the staging directory", lambda: (self.staged_fasta_dir.mkdir(parents=True, exist_ok=True), self.run("stage_matched_refid_fastas.py", "--matched-files", self.matched_txt, "--output-dir", self.staged_fasta_dir))),
            Step("prepare-comet-assignments", "create COMET calls and remove missing or unassigned records", prepare_comet),
            Step("select-noncomet-priority-assignments", "select non-COMET and reference assignments that override or supplement COMET", lambda: self.run("select_noncomet_priority_assignments.py", "--noncomet-coverage-csv", self.noncomet_coverage_csv, "--reference-subtypes-csv", self.reference_subtypes_csv, "--fasta-dir", self.fasta_pool, "--output-csv", self.priority_assignments_csv, stdout_path=summary("select-noncomet-priority-assignments"))),
            Step("filter-accession-metadata", "filter master metadata to COMET-filtered FASTA accessions", lambda: self.run("filter_accessions_metadata_by_fasta.py", "--fasta-dir", self.included_fasta_dir, "--metadata-csv", self.metadata_csv, "--output-dir", self.metadata_dir)),
            Step("split-refid-metadata", "create per-RefID metadata filters", lambda: (self.run("prepare_ns3_pipeline_workdirs.py", "--clean-dir", self.refid_metadata_dir), self.run("split_refid_metadata_csv.py", "--input-csv", self.metadata_dir / "included_accessions_metadata.csv", "--output-dir", self.refid_metadata_dir, "--source-fasta-dir", self.staged_fasta_dir, "--comet-fasta-dir", self.included_fasta_dir))),
            Step("filter-refid-fastas", "filter COMET-filtered FASTA records by RefID metadata", lambda: (shutil.copytree(self.included_fasta_dir, self.filtered_fasta_dir, dirs_exist_ok=True), self.run("filter_refid_fastas_by_metadata.py", "--metadata-dir", self.refid_metadata_dir, "--fasta-dir", self.filtered_fasta_dir, "--kept-accessions-output", self.kept_accessions_csv))),
            Step("build-genotype-workbook", "build the COMET genotype workbook", lambda: self.run("build_ns3_comet_gt_allstudies.py", "--fasta-dir", self.filtered_fasta_dir, "--comet-genotype-csv", comet_gt_csv, "--priority-assignments-csv", self.priority_assignments_csv, "--output-dir", self.step_dir("build-genotype-workbook"), stdout_path=summary("build-genotype-workbook"))),
            Step("add-genotype-counts", "add the genotype-count worksheet", lambda: self.run("add_gt_counts_sheet.py", "--workbook", gt_workbook)),
            Step("build-subtype-workbook", "build the COMET subtype workbook", lambda: self.run("build_ns3_comet_subtype_allstudies.py", "--genotype-workbook", gt_workbook, "--comet-subtype-csv", comet_subtype_csv, "--priority-assignments-csv", self.priority_assignments_csv, "--output-dir", self.step_dir("build-subtype-workbook"), stdout_path=summary("build-subtype-workbook"))),
            Step("extract-profile-aa", "extract genotype-position amino-acid sequences", extract_aa),
            Step("validate-profile-alignment", "write alignment QC columns and flagged-accession report", lambda: self.run("validate_ns3_profile_alignment.py", "--input-workbook", self.step_dir("extract-profile-aa") / "NS3_Profile_Input_Source.xlsx", "--gt-aa-json", self.gt_aa_json, "--output-workbook", self.aa_workbook, "--flagged-accessions-csv", self.step_dir("validate-profile-alignment") / "NS3_Profile_Alignment_QC_Flagged_Accessions.csv", stdout_path=summary("validate-profile-alignment"))),
            Step("summarize-qc-mutation-burden", "summarize QC-passed genotype mutation burden", lambda: self.run("build_qc_passed_genotype_mutation_burden_summary.py", "--input-workbook", self.aa_workbook, "--output-csv", self.step_dir("summarize-qc-mutation-burden") / "NS3_QC_Passed_Genotype_Mutation_Burden_Summary.csv")),
            Step("report-profile-input-counts", "report profile-input inclusion counts", report_profile_counts),
            Step("build-complete-profiles", "build genotype and subtype complete profiles", lambda: self.run("build_ns3_completeprofiles_tabspergt.py", "--input-workbook", self.aa_workbook, "--output-dir", self.step_dir("build-complete-profiles"), "--profile-accessions-csv", self.profile_accessions_csv, stdout_path=summary("build-complete-profiles"))),
            Step("merge-subtype-complete-profiles", "merge subtype complete-profile worksheets into one table", lambda: self.run("merge_ns3_subtype_completeprofiles.py", "--input-workbook", subtype_profile, "--output-workbook", self.step_dir("merge-subtype-complete-profiles") / "NS3_Subtype_CompleteProfiles_Merged.xlsx", stdout_path=summary("merge-subtype-complete-profiles"))),
            Step("export-noncomet-priority-accessions", "export non-COMET priority profile accessions", lambda: self.run("export_noncomet_priority_profile_accessions.py", "--profile-accessions-csv", self.profile_accessions_csv, "--comet-subtype-csv", self.comet_csv, "--noncomet-coverage-csv", self.noncomet_coverage_csv, "--output-csv", self.step_dir("export-noncomet-priority-accessions") / "NS3_NonComet_Priority_Profile_Accessions.csv")),
            Step("export-consensus-fastas", "export genotype and subtype consensus FASTA files", lambda: self.run("export_ns3_consensus_fasta.py", "--gt-profile-workbook", gt_profile, "--subtype-profile-workbook", subtype_profile, "--output-dir", self.step_dir("export-consensus-fastas"))),
            Step("align-subtype-consensus-to-gt1a", "align subtype consensuses to the fixed GT1_1a coordinate system", lambda: self.run("align_ns3_subtype_consensuses_to_gt1a.py", "--input-fasta", subtype_consensus, "--output-fasta", self.step_dir("align-subtype-consensus-to-gt1a") / "NS3_Subtype_Consensus_Aligned_to_GT1_1a.fasta")),
            Step("compare-reference-consensus", "compare GT1_1a-aligned COMET consensuses to pairwise-aligned subtype references", lambda: (self.run("export_gt_reference_consensus_differences.py", "--gene", "NS3", "--reference-fasta", REPO_ROOT / "HCVData/HCV_GT_Refs_NS3_NS5A_NTD_NS5B_AA.fasta", "--consensus-dir", self.step_dir("export-consensus-fastas"), "--output-dir", self.step_dir("compare-reference-consensus"), "--subtype-reference-fasta", REPO_ROOT / "HCVData/Reference_seqs/HCV_Subtype_Refs_NS3_AA_Pairwise_Aligned.fasta", "--subtype-consensus-fasta", self.step_dir("align-subtype-consensus-to-gt1a") / "NS3_Subtype_Consensus_Aligned_to_GT1_1a.fasta", "--subtype-length-csv", self.step_dir("compare-reference-consensus") / "NS3_Subtype_Reference_Consensus_AA_Lengths.csv"), self.run("build_ns3_subtype_consensus_reference_distance.py", "--subtype-profile-workbook", subtype_profile, "--subtype-json", self.subtype_json, "--output-xlsx", self.step_dir("compare-reference-consensus") / "NS3_Subtype_Consensus_Reference_AA_Distance_RAS.xlsx"))),
            Step("build-genotype-ras-profile", "build genotype RAS profile", lambda: self.run("build_ns3_gt_ras_profiles.py", "--gt-profile-workbook", gt_profile, "--gt-aa-json", self.gt_aa_json, "--output-dir", self.step_dir("build-genotype-ras-profile"), stdout_path=summary("build-genotype-ras-profile"))),
            Step("build-subtype-ras-profile", "build subtype RAS profile", lambda: (self.run("build_ns3_subtype_ras_profiles.py", "--subtype-profile-workbook", subtype_profile, "--gt-ras-profile-workbook", gt_ras, "--gt-aa-json", self.gt_aa_json, "--output-dir", self.step_dir("build-subtype-ras-profile"), stdout_path=summary("build-subtype-ras-profile")), self.run("replace_comet_profile_coverage_range_with_mean_diff.py", "--combined-profile-workbook", subtype_ras, "--profile-input-workbook", self.aa_workbook, "--profile-accessions-csv", self.profile_accessions_csv))),
            Step("build-combined-ras-reports", "build combined RAS, coverage, and sequence-audit reports", lambda: (self.run("build_ns3_combined_ras_profiles.py", "--gt-ras-profile-workbook", gt_ras, "--subtype-ras-profile-workbook", subtype_ras, "--output-xlsx", combined_ras, stdout_path=summary("build-combined-ras-reports")), self.run("replace_comet_profile_coverage_range_with_mean_diff.py", "--combined-profile-workbook", combined_ras, "--profile-input-workbook", self.aa_workbook, "--profile-accessions-csv", self.profile_accessions_csv), self.run("build_comet_subtype_ras_coverage_report.py", "--gene", "NS3", "--combined-profile-workbook", combined_ras, "--profile-input-workbook", self.aa_workbook, "--profile-accessions-csv", self.profile_accessions_csv, "--output-xlsx", self.step_dir("build-combined-ras-reports") / "NS3_Subtype_RAS_Coverage_Report.xlsx"), self.run("build_comet_workflow_sequence_audit.py", "--gene", "NS3", "--selection-workbook", self.excel_file, "--selection-sheet", self.sheet_name, "--fasta-dir", self.fasta_pool, "--metadata-csv", self.metadata_csv, "--comet-csv", self.comet_csv, "--qc-workbook", self.aa_workbook, "--profile-accessions-csv", self.profile_accessions_csv, "--combined-profile-workbook", combined_ras, "--combined-subtype-cutoff", "10", "--output-xlsx", self.step_dir("build-combined-ras-reports") / "NS3_Workflow_Sequence_Audit.xlsx"))),
            Step("add-nonconsensus-row", "create annotated combined profile with MeanDiff and PositionDiff", lambda: self.run("add_combined_profile_nonconsensus_row.py", "--combined-profile-workbook", combined_ras, "--output-workbook", annotated_combined_ras, "--profile-input-workbook", self.aa_workbook, "--profile-accessions-csv", self.profile_accessions_csv, "--genotype-consensus-fasta", gt_consensus)),
            Step("summarize-subtype-ras-differences", "summarize subtype RAS differences from genotype consensus", lambda: self.run("build_ns3_subtype_ras_consensus_difference_summary.py", "--combined-profile-workbook", combined_ras, "--profile-input-workbook", self.aa_workbook, "--profile-accessions-csv", self.profile_accessions_csv, "--genotype-consensus-fasta", gt_consensus, "--output-xlsx", self.step_dir("summarize-subtype-ras-differences") / "NS3_Subtype_RAS_Consensus_Difference_Summary.xlsx", "--output-png", self.step_dir("summarize-subtype-ras-differences") / "NS3_Subtype_RAS_Consensus_Difference_Trend.png", stdout_path=summary("summarize-subtype-ras-differences"))),
            Step("build-genotype-aa-consensus-distance", "build genotype consensus amino-acid distance matrix", lambda: self.run("build_ns3_gt_aa_distance_matrix.py", "--input-fasta", gt_consensus, "--aligned-fasta", self.step_dir("build-genotype-aa-consensus-distance") / "NS3_GT_Consensus_aligned.fasta", "--output-xlsx", self.step_dir("build-genotype-aa-consensus-distance") / "NS3_GT_Consensus_AA_Distance_Pos36_175.xlsx", "--details-xlsx", summary("build-genotype-aa-consensus-distance", "NS3_GT_Consensus_AA_Distance_Pos36_175_details.xlsx"), "--start", "36", "--end", "175", stdout_path=summary("build-genotype-aa-consensus-distance", "last_run_summary.txt"))),
            Step("build-subtype-aa-consensus-distance", "build subtype consensus amino-acid distance matrices", lambda: self.run("build_ns3_subtype_aa_distance_matrices.py", "--input-fasta", subtype_consensus, "--subtype-profile-workbook", subtype_profile, "--min-subtype-sequences", "10", "--output-xlsx", self.step_dir("build-subtype-aa-consensus-distance") / "NS3_Subtype_Consensus_AA_Distance_Pos36_175.xlsx", "--temp-dir", self.step_dir("build-subtype-aa-consensus-distance"), "--start", "36", "--end", "175", stdout_path=summary("build-subtype-aa-consensus-distance", "last_run_summary.txt"))),
            Step("build-paired-distance-matrices", "build paired AA and NA RAS/range distance matrices", paired_distances),
            Step("build-subtype-profile-coverage", "build the NS3 genotype 5 subtype 5a coverage audit", lambda: self.run("build_ns3_subtype_profile_coverage_report.py", "--profile-input-workbook", self.aa_workbook, "--profile-accessions-csv", self.profile_accessions_csv, "--genotype", "5", "--subtype", "5a", "--range-start", "1", "--range-end", "631", "--output-xlsx", self.step_dir("build-subtype-profile-coverage") / "NS3_GT5_5a_Profile_Coverage.xlsx", "--position-coverage-csv", self.step_dir("build-subtype-profile-coverage") / "NS3_GT5_5a_Profile_Position_Coverage.csv", "--position-coverage-png", self.step_dir("build-subtype-profile-coverage") / "NS3_GT5_5a_Profile_Position_Coverage.png")),
            Step("build-ras-entropy", "build genotype and subtype RAS entropy reports", lambda: self.run("build_ns3_ras_entropy.py", "--gt-profile-workbook", gt_profile, "--subtype-profile-workbook", subtype_profile, "--gt-output-xlsx", self.step_dir("build-ras-entropy") / "NS3_GT_RAS_Entropy.xlsx", "--subtype-output-xlsx", self.step_dir("build-ras-entropy") / "NS3_Subtype_RAS_Entropy.xlsx", stdout_path=summary("build-ras-entropy", "last_run_summary.txt"))),
            Step("publish-ictv-report", "publish the NS3 ICTV comparison report", lambda: self.run(str(REPO_ROOT / "hcv-workflow" / "hcv-ns5a-comet-build-workflow" / "scripts" / "add_subtype_consensus_mutation_summaries.py"), "--gene", "NS3", "--comparison-workbook", self.step_dir("compare-reference-consensus") / "HCV_Subtype_Ref_vs_Comet_Subtype_Consensus_Aligned_NS3.xlsx", "--ras-workbook", explicit_subtype_ras, "--combined-profile-workbook", combined_ras, "--comparison-output-dir", self.step_dir("compare-reference-consensus"), "--ns3-output-dir", self.step_dir("build-subtype-ras-profile"), "--shared-report-dir", self.step_dir("publish-ictv-report") / "shared_report")),
            Step("audit-gt7-gt8-sequences", "audit GT7 and GT8 kept/excluded sequences against each previous workflow step", lambda: self.run("build_ns3_gt7_gt8_step_audit.py", "--pipeline-output-dir", self.output_dir, "--output-csv", self.step_dir("audit-gt7-gt8-sequences") / "NS3_GT7_GT8_Step_Sequence_Audit.csv", "--accessions-csv", self.step_dir("audit-gt7-gt8-sequences") / "NS3_GT7_GT8_Step_Sequence_Audit_Accessions.csv", "--summary-xlsx", self.step_dir("audit-gt7-gt8-sequences") / "NS3_GT7_GT8_Step_Sequence_Audit_Summary.xlsx", "--summary-markdown", self.step_dir("audit-gt7-gt8-sequences") / "NS3_GT7_GT8_Step_Sequence_Audit_Summary.md")),
            Step("compare-gt7-gt8-local-assignments", "compare GT7 and GT8 workflow subtype calls with local NS3 assignments", lambda: self.run("compare_ns3_gt7_gt8_local_assignments.py", "--subtype-workbook", subtype_workbook, "--coverage-csv", self.noncomet_coverage_csv, "--output-xlsx", self.step_dir("compare-gt7-gt8-local-assignments") / "NS3_GT7_GT8_Local_Assignment_Comparison.xlsx", "--output-csv", self.step_dir("compare-gt7-gt8-local-assignments") / "NS3_GT7_GT8_Local_Assignment_Comparison.csv", stdout_path=summary("compare-gt7-gt8-local-assignments"))),
        ]


def main() -> int:
    args = parse_args()
    pipeline = Pipeline(args)
    steps = pipeline.steps()
    step_by_name = {step.name: step for step in steps}
    if args.list_steps:
        for step in steps:
            pipeline.announce(step.order, step.name, step.description)
        return 0

    requested = args.step or [step.name for step in steps]
    unknown = [name for name in requested if name not in step_by_name]
    if unknown:
        raise SystemExit(f"Unknown step name(s): {', '.join(unknown)}. Use --list-steps.")
    if not args.excel_file and not pipeline.excel_file.is_file() and "discover-refid-fastas" in requested:
        raise SystemExit(f"Excel file was not found: {pipeline.excel_file}")
    if not pipeline.sheet_name and "discover-refid-fastas" in requested:
        raise SystemExit("SHEET_NAME or --sheet-name is required for discover-refid-fastas")
    pipeline.ensure_summary_directories()
    for name in requested:
        step = step_by_name[name]
        pipeline.current_step_dir = pipeline.step_dir(step.name)
        pipeline.current_step_order = step.order
        pipeline.current_step_dir.mkdir(parents=True, exist_ok=True)
        pipeline.announce(step.order, step.name, step.description)
        input_fasta_dir = pipeline.staged_fasta_dir if step.order <= STEP_ORDER["stage-refid-fastas"] else (
            pipeline.included_fasta_dir if step.order <= STEP_ORDER["filter-refid-fastas"] else pipeline.filtered_fasta_dir
        )
        input_accessions = pipeline.fasta_accessions(input_fasta_dir)
        try:
            step.action()
        finally:
            pipeline.print_accession_counts(input_accessions)
            if step.order > STEP_ORDER["stage-refid-fastas"]:
                pipeline.print_gt7_gt8_subtype_counts()
    print("NS3 pipeline complete")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
