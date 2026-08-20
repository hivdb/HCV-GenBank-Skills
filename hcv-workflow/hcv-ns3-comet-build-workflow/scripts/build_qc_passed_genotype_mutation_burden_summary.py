#!/usr/bin/env python3
"""Summarize genotype mutation burden among QC-passed profile-input rows."""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from pathlib import Path
from statistics import median

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-workbook", required=True, help="Profile input alignment-QC workbook.")
    parser.add_argument("--output-csv", required=True)
    return parser.parse_args()


def percentile(values: list[float], percent: float) -> float:
    """Return the linearly interpolated percentile used for report summaries."""
    ordered = sorted(values)
    if not ordered:
        raise ValueError("Cannot calculate a percentile of no values")
    index = (len(ordered) - 1) * percent / 100
    lower = int(index)
    upper = min(lower + 1, len(ordered) - 1)
    return ordered[lower] + (ordered[upper] - ordered[lower]) * (index - lower)


def display_number(value: float) -> str | int:
    rounded = round(value, 1)
    return int(rounded) if rounded.is_integer() else rounded


def main() -> int:
    args = parse_args()
    input_path = Path(args.input_workbook)
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook.active
    header = [str(value or "") for value in next(worksheet.iter_rows(values_only=True))]
    index = {name: position for position, name in enumerate(header)}
    required = {"ClosestGT", "AlignmentQCStatus", "AlignmentQCComparedAA", "AlignmentQCMutationCount", "AlignmentQCMutationPercent"}
    missing = required - index.keys()
    if missing:
        raise RuntimeError(f"Missing columns in {input_path}: {', '.join(sorted(missing))}")

    by_genotype: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if str(row[index["AlignmentQCStatus"]] or "").strip() != "PASS":
            continue
        genotype = str(row[index["ClosestGT"]] or "").strip().removeprefix("GT")
        if not genotype:
            continue
        try:
            covered = float(row[index["AlignmentQCComparedAA"]])
            mutations = float(row[index["AlignmentQCMutationCount"]])
            mutation_percent = float(row[index["AlignmentQCMutationPercent"]])
        except (TypeError, ValueError):
            continue
        by_genotype[genotype]["covered"].append(covered)
        by_genotype[genotype]["mutations"].append(mutations)
        by_genotype[genotype]["mutation_percent"].append(mutation_percent)
    workbook.close()

    fields = [
        "genotype", "n", "median_covered_aa", "median_mutations", "p95_mutations", "p99_mutations", "max_mutations",
        "median_mutation_pct", "p95_mutation_pct", "p99_mutation_pct", "max_mutation_pct", "n_ge_10pct", "n_ge_15pct", "n_ge_20pct",
    ]
    output_rows: list[dict[str, str | int | float]] = []
    for genotype in sorted(by_genotype, key=lambda value: int(value) if value.isdigit() else 999):
        values = by_genotype[genotype]
        mutation_percent = values["mutation_percent"]
        output_rows.append({
            "genotype": f"GT{genotype}",
            "n": len(mutation_percent),
            "median_covered_aa": display_number(median(values["covered"])),
            "median_mutations": display_number(median(values["mutations"])),
            "p95_mutations": display_number(percentile(values["mutations"], 95)),
            "p99_mutations": display_number(percentile(values["mutations"], 99)),
            "max_mutations": display_number(max(values["mutations"])),
            "median_mutation_pct": display_number(median(mutation_percent)),
            "p95_mutation_pct": display_number(percentile(mutation_percent, 95)),
            "p99_mutation_pct": display_number(percentile(mutation_percent, 99)),
            "max_mutation_pct": display_number(max(mutation_percent)),
            "n_ge_10pct": sum(value >= 10 for value in mutation_percent),
            "n_ge_15pct": sum(value >= 15 for value in mutation_percent),
            "n_ge_20pct": sum(value >= 20 for value in mutation_percent),
        })
    output_path = Path(args.output_csv)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(output_rows)
    total_accessions = sum(int(row["n"]) for row in output_rows)
    distribution = sorted(output_rows, key=lambda row: (-int(row["n"]), str(row["genotype"])))
    distribution_text = ", ".join(
        f"{row['genotype']} ({int(row['n']) / total_accessions:.1%}, {int(row['n'])})"
        for row in distribution
    ) if total_accessions else ""
    print(f"QC-passed genotype distribution: {distribution_text}")
    print(f"Total accessions: {total_accessions}")
    print(f"Wrote {len(output_rows)} genotype summaries to {output_path.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
