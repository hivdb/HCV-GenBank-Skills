#!/usr/bin/env python3
"""Create the NS3 subtype workbook directly from Comet assignments."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--genotype-workbook", required=True)
    parser.add_argument("--comet-subtype-csv", required=True)
    parser.add_argument("--priority-assignments-csv", required=True)
    parser.add_argument("--output-dir", required=True)
    return parser.parse_args()


def load_subtypes(path: Path) -> dict[str, str]:
    assignments: dict[str, str] = {}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            accession = (row.get("accession") or "").strip()
            subtype = (row.get("subtype") or "").strip().lower()
            if accession and subtype:
                assignments[accession] = subtype
                assignments.setdefault(accession.split(".", 1)[0], subtype)
    return assignments


def load_priority_subtypes(path: Path) -> dict[str, tuple[str, str, str, str, str]]:
    assignments: dict[str, tuple[str, str, str, str, str]] = {}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            accession = str(row.get("Accession") or "").strip()
            genotype = str(row.get("ClosestGenotype") or "").strip().lower()
            subtype = str(row.get("ClosestSubtype") or "").strip().lower()
            if accession:
                assignments[accession.split(".", 1)[0]] = ("", "", accession, genotype, subtype)
    return assignments


def main() -> int:
    args = parse_args()
    assignments = load_subtypes(Path(args.comet_subtype_csv))
    noncomet_priority_subtypes = load_priority_subtypes(Path(args.priority_assignments_csv))
    workbook = load_workbook(args.genotype_workbook, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header = [str(value) if value is not None else "" for value in next(sheet.iter_rows(values_only=True))]
    index = {name: position for position, name in enumerate(header)}
    required = ["RefID", "RefName", "GenBankAccession", "BestGT"]
    missing = [name for name in required if name not in index]
    if missing:
        raise RuntimeError(f"Columns missing from {args.genotype_workbook}: {', '.join(missing)}")

    rows: list[tuple[str, str, str, str, str, str, str]] = []
    seen_accessions: set[str] = set()
    override_count = 0
    for values in sheet.iter_rows(min_row=2, values_only=True):
        accession = str(values[index["GenBankAccession"]]).strip()
        accession_key = accession.split(".", 1)[0]
        subtype = assignments.get(accession) or assignments.get(accession.split(".", 1)[0])
        if subtype:
            genotype = str(values[index["BestGT"]]).strip()
            if accession_key in noncomet_priority_subtypes:
                priority_assignment = noncomet_priority_subtypes[accession_key]
                genotype, subtype = priority_assignment[3], priority_assignment[4]
                source, metadata_column = "Non-Comet priority subtype override", "Non-Comet ClosestSubtype"
                override_count += 1
            else:
                source, metadata_column = "Comet", "Comet NS3"
            rows.append((
                str(values[index["RefID"]]).strip(),
                str(values[index["RefName"]]).strip(),
                accession,
                genotype,
                subtype,
                source,
                metadata_column,
            ))
            seen_accessions.add(accession_key)
    workbook.close()

    addition_count = 0
    for accession_key, (refid, refname, accession, genotype, subtype) in noncomet_priority_subtypes.items():
        if accession_key not in seen_accessions:
            rows.append((refid, refname, accession, genotype, subtype, "Non-Comet priority subtype addition", "Non-Comet ClosestSubtype"))
            addition_count += 1

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "NS3_Subtype_AllStudies_WSeqs.xlsx"
    output = Workbook()
    sheet = output.active
    sheet.title = "NS3_Subtype_Comet"
    sheet.append([
        "RefID", "RefName", "AccessionID", "ClosestGT", "ClosestSubtype",
        "ClosestSubtypeAssignmentSource", "ClosestSubtypeMetadataColumn",
    ])
    for row in rows:
        sheet.append(row)
    output.save(output_path)
    print(json.dumps({"output_workbook": str(output_path.resolve()), "row_count": len(rows), "noncomet_priority_subtype_override_count": override_count, "noncomet_priority_subtype_addition_count": addition_count}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
