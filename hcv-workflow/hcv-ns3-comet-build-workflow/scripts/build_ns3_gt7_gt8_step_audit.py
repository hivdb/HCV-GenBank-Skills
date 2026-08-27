#!/usr/bin/env python3
"""Audit GT7/GT8 retention at each NS3 COMET workflow step."""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--pipeline-output-dir", required=True, type=Path)
    parser.add_argument("--gene", choices=("NS3", "NS5A", "NS5B"), default="NS3")
    parser.add_argument("--output-csv", required=True, type=Path)
    parser.add_argument("--accessions-csv", required=True, type=Path)
    parser.add_argument("--summary-xlsx", required=True, type=Path)
    parser.add_argument("--summary-markdown", required=True, type=Path)
    return parser.parse_args()


def workflow_steps(root: Path) -> list[tuple[int, str]]:
    steps: list[tuple[int, str]] = []
    for path in root.iterdir() if root.is_dir() else ():
        match = re.fullmatch(r"(\d+)_([\w-]+)", path.name)
        if path.is_dir() and match:
            steps.append((int(match.group(1)), match.group(2)))
    return sorted(steps)


def accession_key(value: object) -> str:
    return str(value or "").strip().split(".", 1)[0]


def genotype(value: object) -> str:
    return str(value or "").strip().upper().removeprefix("GT")


def workbook_rows(path: Path) -> list[dict[str, object]]:
    if not path.is_file():
        return []
    book = load_workbook(path, read_only=True, data_only=True)
    sheet = book[book.sheetnames[0]]
    header = [str(value or "") for value in next(sheet.iter_rows(values_only=True), ())]
    rows = [dict(zip(header, values)) for values in sheet.iter_rows(min_row=2, values_only=True)]
    book.close()
    return rows


def csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        return []
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def rows_as_gt_accessions(rows: list[dict[str, object]], accession_column: str, genotype_column: str) -> dict[str, str]:
    return {
        accession_key(row.get(accession_column)): genotype(row.get(genotype_column))
        for row in rows
        if accession_key(row.get(accession_column)) and genotype(row.get(genotype_column)) in {"7", "8"}
    }


def assignment_map(path: Path, accession_column: str, genotype_column: str) -> dict[str, str]:
    """Return final assignment per accession, matching the workflow builders' overwrite behavior."""
    assignments: dict[str, str] = {}
    for row in csv_rows(path):
        accession = accession_key(row.get(accession_column))
        gt = genotype(row.get(genotype_column))
        if accession and gt:
            assignments[accession] = gt
    return assignments


def gt7_gt8_assignments(assignments: dict[str, str]) -> dict[str, str]:
    return {accession: gt for accession, gt in assignments.items() if gt in {"7", "8"}}


def fasta_accessions(directory: Path) -> set[str]:
    accessions: set[str] = set()
    if directory.is_dir():
        for path in directory.glob("*.fasta"):
            for line in path.read_text(encoding="utf-8").splitlines():
                if line.startswith(">"):
                    accessions.add(accession_key(line[1:].split(maxsplit=1)[0]))
    return accessions


def filtered_assignments(directory: Path, comet: dict[str, str], priority: dict[str, str]) -> dict[str, str]:
    retained = fasta_accessions(directory)
    result = {accession: priority.get(accession, comet.get(accession, "")) for accession in retained}
    # Priority rows deliberately supplement COMET calls, including accessions absent from COMET.
    result.update(priority)
    return gt7_gt8_assignments(result)


def qc_passes_and_reasons(path: Path) -> tuple[dict[str, str], dict[str, str]]:
    passes: dict[str, str] = {}
    reasons: dict[str, str] = {}
    for row in workbook_rows(path):
        accession = accession_key(row.get("AccessionID"))
        gt = genotype(row.get("ClosestGT"))
        if not accession or gt not in {"7", "8"}:
            continue
        status = str(row.get("AlignmentQCStatus") or "").strip()
        if status == "PASS":
            passes[accession] = gt
        else:
            reasons[accession] = str(row.get("AlignmentQCReasons") or status or "not_qc_passed").strip()
    return passes, reasons


def write_sheet(book: Workbook, title: str, fields: list[str], rows: list[dict[str, object]]) -> None:
    sheet = book.create_sheet(title)
    sheet.append(fields)
    for row in rows:
        sheet.append([row.get(field, "") for field in fields])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for column in range(1, len(fields) + 1):
        width = max(len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1))
        sheet.column_dimensions[get_column_letter(column)].width = min(width + 2, 60)


def write_summary_workbook(path: Path, summary: list[dict[str, object]], details: list[dict[str, object]], summary_fields: list[str], accession_fields: list[str]) -> None:
    book = Workbook()
    book.remove(book.active)
    key_changes: list[dict[str, object]] = []
    previous_counts: dict[str, int] = {}
    for row in summary:
        genotype_value = str(row["genotype"])
        current = int(row["kept_count"])
        prior = previous_counts.get(genotype_value)
        if prior is not None and (current != prior or int(row["excluded_since_previous_count"]) > 0):
            key_changes.append({
                "step": row["step"], "genotype": genotype_value, "previous_kept_count": prior,
                "kept_count": current, "net_change": current - prior,
                "excluded_since_previous_count": row["excluded_since_previous_count"],
                "exclusion_reason": row["exclusion_reason"],
            })
        previous_counts[genotype_value] = current
    write_sheet(book, "Key Changes", ["step", "genotype", "previous_kept_count", "kept_count", "net_change", "excluded_since_previous_count", "exclusion_reason"], key_changes)
    write_sheet(book, "Step Audit", summary_fields, summary)
    write_sheet(book, "Accession Audit", accession_fields, details)
    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)


def write_summary_markdown(path: Path, gene: str, summary: list[dict[str, object]], details: list[dict[str, object]]) -> None:
    previous_counts: dict[str, int] = {}
    key_changes: list[dict[str, object]] = []
    for row in summary:
        gt = str(row["genotype"])
        current = int(row["kept_count"])
        prior = previous_counts.get(gt)
        if prior is not None and (current != prior or int(row["excluded_since_previous_count"]) > 0):
            key_changes.append({**row, "previous_kept_count": prior, "net_change": current - prior})
        previous_counts[gt] = current
    lines = [f"# {gene} GT7/GT8 sequence-retention summary", "", "## Key changes", "", "| Step | Genotype | Previous kept | Kept | Net change | Excluded | Reason |", "| --- | --- | ---: | ---: | ---: | ---: | --- |"]
    lines.extend(f"| {row['step']} | GT{row['genotype']} | {row['previous_kept_count']} | {row['kept_count']} | {row['net_change']:+d} | {row['excluded_since_previous_count']} | {row['exclusion_reason'] or ''} |" for row in key_changes)
    excluded = [row for row in details if row["status"] == "excluded"]
    lines.extend(["", "## Excluded accessions", "", "| Step | Genotype | Accession | Reason |", "| --- | --- | --- | --- |"])
    lines.extend(f"| {row['step']} | GT{row['genotype']} | {row['accession']} | {row['exclusion_reason']} |" for row in excluded)
    final = {row["genotype"]: row["kept_count"] for row in summary if row["step"] == "audit-gt7-gt8-sequences"}
    lines.extend(["", "## Final retained sequences", "", f"- GT7: {final.get('7', 0)}", f"- GT8: {final.get('8', 0)}", ""])
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    args = parse_args()
    root = args.pipeline_output_dir
    gene = args.gene.upper()
    steps = workflow_steps(root)
    step_dirs = {name: root / f"{number:02d}_{name}" for number, name in steps}
    step = lambda name: step_dirs.get(name, root / "__missing_step__")
    comet_all = assignment_map(step("prepare-comet-assignments") / f"comet_{gene.lower()}_subtype_assignments.csv", "accession", "genotype")
    priority_all = assignment_map(step("select-noncomet-priority-assignments") / f"{gene}_NonComet_Priority_Assignments.csv", "Accession", "ClosestGenotype")
    comet = gt7_gt8_assignments(comet_all)
    prefilter_all = {**comet_all, **priority_all}
    prefilter = gt7_gt8_assignments(prefilter_all)
    priority_reassignments = {
        accession: f"reassigned_to_GT{prefilter_all[accession]}_by_priority_assignment"
        for accession in comet
        if accession in prefilter_all and prefilter_all[accession] not in {"7", "8"}
    }
    filtered = filtered_assignments(step("filter-refid-fastas") / "included_refid_fastas", comet_all, priority_all)
    genotype_rows = rows_as_gt_accessions(workbook_rows(step("build-genotype-workbook") / f"{gene}_GT_AllStudies.xlsx"), "GenBankAccession", "BestGT")
    subtype_rows = rows_as_gt_accessions(workbook_rows(step("build-subtype-workbook") / f"{gene}_Subtype_AllStudies_WSeqs.xlsx"), "AccessionID", "ClosestGT")
    extracted_rows = rows_as_gt_accessions(workbook_rows(step("extract-profile-aa") / f"{gene}_Profile_Input_Source.xlsx"), "AccessionID", "ClosestGT")
    qc_rows, qc_reasons = qc_passes_and_reasons(step("validate-profile-alignment") / f"{gene}_Profile_Input_Alignment_QC.xlsx")
    profile_rows = rows_as_gt_accessions(csv_rows(step("build-complete-profiles") / f"{gene}_Profile_Accessions_QC_Pass.csv"), "accession", "genotype")

    snapshots: dict[str, tuple[dict[str, str] | None, dict[str, str]]] = {
        "prepare-comet-assignments": (comet, {}),
        "select-noncomet-priority-assignments": (prefilter, priority_reassignments),
        "filter-accession-metadata": (prefilter, {}),
        "split-refid-metadata": (prefilter, {}),
        "filter-refid-fastas": (filtered, {accession: "excluded_by_RefID_metadata_filter_or_missing_from_filtered_FASTA" for accession in prefilter}),
        "build-genotype-workbook": (genotype_rows, {}),
        "add-genotype-counts": (genotype_rows, {}),
        "build-subtype-workbook": (subtype_rows, {accession: "missing_subtype_assignment" for accession in genotype_rows}),
        "extract-profile-aa": (extracted_rows, {accession: "not_available_in_profile_AA_extraction" for accession in subtype_rows}),
        "validate-profile-alignment": (qc_rows, qc_reasons),
        "build-complete-profiles": (profile_rows, {}),
    }

    summary_fields = ["step_number", "step", "genotype", "previous_step", "kept_count", "excluded_since_previous_count", "exclusion_reason", "comparison_note"]
    accession_fields = ["step_number", "step", "genotype", "accession", "status", "previous_step", "exclusion_reason"]
    summary: list[dict[str, object]] = []
    details: list[dict[str, object]] = []
    previous: dict[str, str] | None = None
    previous_step = ""
    last_snapshot: dict[str, str] | None = None
    for number, name in steps:
        snapshot, reasons = snapshots.get(name, (last_snapshot, {}))
        if snapshot is None:
            for gt in ("7", "8"):
                summary.append({"step_number": number, "step": name, "genotype": gt, "previous_step": "", "kept_count": 0, "excluded_since_previous_count": 0, "exclusion_reason": "", "comparison_note": "Genotype assignment is not available before COMET assignments."})
            continue
        if name in snapshots:
            last_snapshot = snapshot
        old = previous or {}
        excluded = set(old) - set(snapshot) if previous is not None else set()
        for gt in ("7", "8"):
            kept = sorted(accession for accession, value in snapshot.items() if value == gt)
            lost = sorted(accession for accession in excluded if old[accession] == gt)
            reason_counts: dict[str, int] = {}
            for accession in lost:
                reason = reasons.get(accession, "not_retained_by_this_step")
                reason_counts[reason] = reason_counts.get(reason, 0) + 1
                details.append({"step_number": number, "step": name, "genotype": gt, "accession": accession, "status": "excluded", "previous_step": previous_step, "exclusion_reason": reason})
            for accession in kept:
                details.append({"step_number": number, "step": name, "genotype": gt, "accession": accession, "status": "kept", "previous_step": previous_step, "exclusion_reason": ""})
            reason_text = "; ".join(f"{reason} ({count})" for reason, count in sorted(reason_counts.items()))
            note = "Initial GT7/GT8 assignment snapshot." if previous is None else ("No sequence-level change from the previous step." if not lost else "Compared with the previous step.")
            summary.append({"step_number": number, "step": name, "genotype": gt, "previous_step": previous_step, "kept_count": len(kept), "excluded_since_previous_count": len(lost), "exclusion_reason": reason_text, "comparison_note": note})
        previous, previous_step = snapshot, name

    for path, fields, rows in ((args.output_csv, summary_fields, summary), (args.accessions_csv, accession_fields, details)):
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fields)
            writer.writeheader()
            writer.writerows(rows)
    write_summary_workbook(args.summary_xlsx, summary, details, summary_fields, accession_fields)
    write_summary_markdown(args.summary_markdown, gene, summary, details)
    print(f"gt7_gt8_step_audit_csv={args.output_csv.resolve()}")
    print(f"gt7_gt8_step_audit_accessions_csv={args.accessions_csv.resolve()}")
    print(f"gt7_gt8_step_audit_summary_xlsx={args.summary_xlsx.resolve()}")
    print(f"gt7_gt8_step_audit_summary_markdown={args.summary_markdown.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
