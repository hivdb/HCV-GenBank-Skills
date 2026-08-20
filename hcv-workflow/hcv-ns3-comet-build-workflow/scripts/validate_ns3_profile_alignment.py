#!/usr/bin/env python3
"""Flag gene AA extractions whose linear coordinates cannot represent the alignment.

The profile builders use ``StartAAPosition + AASequence offset`` as a coordinate
map.  That map is invalid as soon as an extraction has an internal deletion or
insertion relative to its genotype reference.  This QC step preserves the input
rows, appends explicit QC columns, and emits a flag-only CSV before profiles are
constructed.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


RAS_POSITIONS = [36, 41, 43, 54, 55, 56, 80, 122, 155, 156, 158, 166, 168, 170, 175]
VALID_AAS = set("ACDEFGHIKLMNPQRSTVWY")
FLAG_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate NS3 profile coordinate mapping against genotype AA references."
    )
    parser.add_argument("--input-workbook", required=True)
    parser.add_argument("--gt-aa-json", required=True)
    parser.add_argument("--reference-gene", default="NS3", help="Reference gene label, e.g. NS3, NS5A_NTD, or NS5B.")
    parser.add_argument("--ras-positions", default=",".join(map(str, RAS_POSITIONS)))
    parser.add_argument("--output-workbook", required=True)
    parser.add_argument("--flagged-accessions-csv", required=True)
    parser.add_argument("--high-divergence-percent", type=float, default=15.0)
    parser.add_argument("--min-divergence-coverage", type=int, default=150)
    return parser.parse_args()


def load_gt_refs(path: Path, gene: str) -> dict[str, str]:
    refs: dict[str, str] = {}
    for row in json.loads(path.read_text(encoding="utf-8")):
        match = re.fullmatch(rf"HCV([1-8]){re.escape(gene)}", str(row.get("name", "")))
        if match:
            refs[match.group(1)] = str(row.get("refSequence", "")).strip().upper()
    return refs


def qc_row(
    row: dict[str, Any], refs: dict[str, str], high_divergence_percent: float, min_divergence_coverage: int, ras_positions: list[int]
) -> dict[str, str | int | float]:
    sequence = str(row.get("AASequence", "") or "").strip().upper()
    start_value = row.get("StartAAPosition")
    if not sequence:
        return {"AlignmentQCStatus": "NO_AA_SEQUENCE", "AlignmentQCReasons": "no_aa_sequence", "AlignmentQCAlignmentColumnsMissing": "", "AlignmentQCRASPositionsRequiringReview": "", "AlignmentQCCoordinateSpan": "", "AlignmentQCComparedAA": "", "AlignmentQCMutationCount": "", "AlignmentQCMutationPercent": ""}
    if start_value in (None, "") or row.get("EndAAPosition") in (None, ""):
        return {"AlignmentQCStatus": "MISSING_AA_COORDINATES", "AlignmentQCReasons": "missing_start_or_end_aa_position", "AlignmentQCAlignmentColumnsMissing": "", "AlignmentQCRASPositionsRequiringReview": "", "AlignmentQCCoordinateSpan": "", "AlignmentQCComparedAA": "", "AlignmentQCMutationCount": "", "AlignmentQCMutationPercent": ""}

    start = int(start_value)
    end = int(row["EndAAPosition"])
    coordinate_span = end - start + 1
    missing_columns = coordinate_span - len(sequence)
    needs_review = [position for position in ras_positions if start <= position <= end]
    reasons = ["coordinate_span_length_mismatch"] if missing_columns else []
    compared = mutations = 0
    reference = refs.get(str(row.get("ClosestGT", "")).strip(), "")
    if not reasons and reference:
        for offset, aa in enumerate(sequence):
            position = start + offset
            if position <= len(reference) and aa in VALID_AAS and reference[position - 1] in VALID_AAS:
                compared += 1
                mutations += aa != reference[position - 1]
    mutation_percent = 100.0 * mutations / compared if compared else 0.0
    status = "FLAGGED" if reasons else "PASS"
    if status == "PASS" and compared >= min_divergence_coverage and mutation_percent >= high_divergence_percent:
        status = "HIGH_DIVERGENCE"
        reasons.append(f"mutation_percent_gte_{high_divergence_percent:g}")
    return {
        "AlignmentQCStatus": status,
        "AlignmentQCReasons": ";".join(reasons),
        "AlignmentQCAlignmentColumnsMissing": missing_columns if missing_columns else "",
        "AlignmentQCRASPositionsRequiringReview": ";".join(f"P{position}" for position in needs_review) if reasons else "",
        "AlignmentQCCoordinateSpan": f"{start}-{end} ({coordinate_span} columns; {len(sequence)} AA)",
        "AlignmentQCComparedAA": compared,
        "AlignmentQCMutationCount": mutations,
        "AlignmentQCMutationPercent": round(mutation_percent, 1),
    }


def main() -> int:
    args = parse_args()
    input_path = Path(args.input_workbook)
    refs = load_gt_refs(Path(args.gt_aa_json), args.reference_gene)
    ras_positions = [int(value) for value in args.ras_positions.split(",") if value.strip()]
    wb = load_workbook(input_path, read_only=True, data_only=True)
    source = wb[wb.sheetnames[0]]
    header = [str(value) if value is not None else "" for value in next(source.iter_rows(values_only=True))]
    rows = [dict(zip(header, values)) for values in source.iter_rows(min_row=2, values_only=True)]
    wb.close()

    qc_columns = ["AlignmentQCStatus", "AlignmentQCReasons", "AlignmentQCAlignmentColumnsMissing", "AlignmentQCRASPositionsRequiringReview", "AlignmentQCCoordinateSpan", "AlignmentQCComparedAA", "AlignmentQCMutationCount", "AlignmentQCMutationPercent"]
    flagged: list[dict[str, Any]] = []
    for row in rows:
        row.update(qc_row(row, refs, args.high_divergence_percent, args.min_divergence_coverage, ras_positions))
        if row["AlignmentQCStatus"] != "PASS":
            flagged.append(row)

    output = Workbook()
    profile_sheet = output.active
    profile_sheet.title = "Profile_Input_QC"
    full_header = header + qc_columns
    profile_sheet.append(full_header)
    for row in rows:
        profile_sheet.append([row.get(column, "") for column in full_header])
        if row["AlignmentQCStatus"] != "PASS":
            for cell in profile_sheet[profile_sheet.max_row]:
                cell.fill = FLAG_FILL

    flagged_sheet = output.create_sheet("Flagged_Accessions")
    flagged_header = ["RefID", "RefName", "AccessionID", "ClosestGT", "ClosestSubtype"] + qc_columns
    flagged_sheet.append(flagged_header)
    for row in flagged:
        flagged_sheet.append([row.get(column, "") for column in flagged_header])
    for cell in flagged_sheet[1]:
        cell.font = Font(bold=True)

    summary_sheet = output.create_sheet("Summary")
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["Input rows", len(rows)])
    summary_sheet.append(["Non-pass rows", len(flagged)])
    for status, count in sorted(Counter(str(row["AlignmentQCStatus"]) for row in rows).items()):
        summary_sheet.append([f"{status} rows", count])
    for (gt, subtype), count in sorted(Counter((str(row.get("ClosestGT", "")), str(row.get("ClosestSubtype", ""))) for row in flagged).items()):
        summary_sheet.append([f"Flagged GT{gt}_{subtype}", count])
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)

    output_path = Path(args.output_workbook)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output.save(output_path)
    csv_path = Path(args.flagged_accessions_csv)
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=flagged_header)
        writer.writeheader()
        writer.writerows({column: row.get(column, "") for column in flagged_header} for row in flagged)

    print(json.dumps({"input_rows": len(rows), "status_counts": dict(sorted(Counter(str(row["AlignmentQCStatus"]) for row in rows).items())), "output_workbook": str(output_path.resolve()), "flagged_accessions_csv": str(csv_path.resolve()), "high_divergence_percent": args.high_divergence_percent, "min_divergence_coverage": args.min_divergence_coverage}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
