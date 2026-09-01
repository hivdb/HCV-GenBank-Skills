#!/usr/bin/env python3
"""Audit ambiguous amino-acid calls at RAS positions used in a COMET profile.

The profile source stores one character per amino-acid position.  ``B`` (D/N),
``Z`` (E/Q), and ``J`` (I/L) are amino-acid mixture codes; ``X`` is an
unknown call.  This audit reports them separately and never assigns a mixture
to one of its component amino acids.
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


MIXTURE_CODES = frozenset({"B", "Z", "J"})
OUTPUT_COLUMNS = [
    "Genotype",
    "Subtype",
    "Position",
    "IncludedAccessions",
    "NonXCoverage",
    "MixtureCount",
    "XCount",
    "B_DN_Count",
    "Z_EQ_Count",
    "J_IL_Count",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--profile-input-workbook", required=True)
    parser.add_argument("--profile-accessions-csv", required=True)
    parser.add_argument(
        "--positions", required=True, help="Comma-separated RAS positions"
    )
    parser.add_argument("--group-by", choices=("genotype", "subtype"), required=True)
    parser.add_argument("--output-csv", required=True)
    return parser.parse_args()


def positions(value: str) -> list[int]:
    result = sorted({int(token.strip()) for token in value.split(",") if token.strip()})
    if not result:
        raise RuntimeError("At least one RAS position is required.")
    return result


def load_assignments(path: Path) -> dict[str, tuple[str, str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        assignments = {}
        for row in csv.DictReader(handle):
            accession = str(row.get("accession") or "").strip()
            if accession:
                assignments[accession] = (
                    str(row.get("genotype") or "").strip(),
                    str(row.get("subtype") or "").strip(),
                )
    return assignments


def audit(args: argparse.Namespace) -> dict[str, object]:
    wanted_positions = positions(args.positions)
    assignments = load_assignments(Path(args.profile_accessions_csv))
    workbook = load_workbook(
        args.profile_input_workbook, read_only=True, data_only=True
    )
    worksheet = workbook[workbook.sheetnames[0]]
    header = [str(value or "") for value in next(worksheet.iter_rows(values_only=True))]
    index = {name: number for number, name in enumerate(header)}
    required = ["AccessionID", "StartAAPosition", "AASequence"]
    missing = [name for name in required if name not in index]
    if missing:
        raise RuntimeError(f"Missing columns in profile input: {', '.join(missing)}")

    groups: dict[tuple[str, str], set[str]] = defaultdict(set)
    calls: dict[tuple[str, str, int], Counter[str]] = defaultdict(Counter)
    wanted = set(wanted_positions)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        accession = str(row[index["AccessionID"]] or "").strip()
        assignment = assignments.get(accession)
        if assignment is None:
            continue
        if (
            "AlignmentQCStatus" in index
            and str(row[index["AlignmentQCStatus"]] or "").strip() != "PASS"
        ):
            continue
        start = row[index["StartAAPosition"]]
        sequence = str(row[index["AASequence"]] or "").strip().upper()
        if start in (None, "") or not sequence:
            continue
        genotype, subtype = assignment
        group = (genotype, "" if args.group_by == "genotype" else subtype)
        groups[group].add(accession)
        for offset, amino_acid in enumerate(sequence):
            position = int(start) + offset
            if position not in wanted:
                continue
            key = (*group, position)
            if amino_acid != "X":
                calls[key]["NonXCoverage"] += 1
            if amino_acid == "X":
                calls[key]["XCount"] += 1
            elif amino_acid in MIXTURE_CODES:
                calls[key]["MixtureCount"] += 1
                calls[key][f"{amino_acid}Count"] += 1
    workbook.close()

    rows = []
    for (genotype, subtype), accessions in sorted(groups.items()):
        for position in wanted_positions:
            count = calls[(genotype, subtype, position)]
            rows.append(
                {
                    "Genotype": genotype,
                    "Subtype": subtype,
                    "Position": position,
                    "IncludedAccessions": len(accessions),
                    "NonXCoverage": count["NonXCoverage"],
                    "MixtureCount": count["MixtureCount"],
                    "XCount": count["XCount"],
                    "B_DN_Count": count["BCount"],
                    "Z_EQ_Count": count["ZCount"],
                    "J_IL_Count": count["JCount"],
                }
            )
    output_path = Path(args.output_csv)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    return {
        "output_csv": str(output_path.resolve()),
        "group_by": args.group_by,
        "group_count": len(groups),
        "position_count": len(wanted_positions),
        "mixture_codes": {"B": "D/N", "Z": "E/Q", "J": "I/L"},
    }


def main() -> int:
    print(json.dumps(audit(parse_args()), indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
