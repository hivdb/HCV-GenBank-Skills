#!/usr/bin/env python3
"""Copy final workflow artifacts into its output report directory."""

from __future__ import annotations
import argparse
import csv
import json
import math
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook

README_GENE = {"NS3": "NS3", "NS5A": "NS5A_NTD", "NS5B": "NS5B"}
SUBTYPE_PROFILE_LABEL = re.compile(
    r"^GT(?P<genotype>[1-6])_(?P<subtype>\S+)\s+\((?P<accessions>\d+),"
)


def add_distance_notes(path: Path, sequence_type: str) -> None:
    """Add an explanation of the sequence-quality exclusion rule."""
    workbook = load_workbook(path)
    if "Notes" in workbook.sheetnames:
        del workbook["Notes"]
    notes = workbook.create_sheet("Notes")
    notes.append(["Note"])
    notes.append(
        [
            f"Sequences with an ambiguous or missing {sequence_type} call at any "
            "required comparison position are excluded from this distance matrix."
        ]
    )
    if "excluded_sequences" in workbook.sheetnames:
        del workbook["excluded_sequences"]
    notes.column_dimensions["A"].width = 110
    notes.freeze_panes = "A2"
    workbook.save(path)


def format_significant_percent(fraction: float) -> str:
    """Format a proportion with the workflow's one-significant-figure rule."""
    if fraction <= 0:
        return "0%"
    percent = fraction * 100
    magnitude = math.floor(math.log10(percent))
    rounded = round(percent, -magnitude)
    decimal_places = max(0, -math.floor(math.log10(rounded)))
    return f"{rounded:.{decimal_places}f}%"


def step_directory(root: Path, step_name: str) -> Path:
    matches = sorted(root.glob(f"*_{step_name}"))
    if len(matches) != 1:
        raise SystemExit(
            f"Expected one {step_name} directory under {root}, found {len(matches)}"
        )
    return matches[0]


def count_gt1_to_gt6_subtypes_with_minimum_accessions(
    profile_workbook: Path, minimum_accessions: int = 10
) -> tuple[int, int]:
    """Return qualifying and total GT1--GT6 subtype counts from the profile."""
    workbook = load_workbook(profile_workbook, read_only=True, data_only=True)
    worksheet = workbook["Combined_RAS_Profile"]
    subtypes: dict[tuple[str, str], int] = {}
    for (label, *_) in worksheet.iter_rows(min_row=2, values_only=True):
        match = SUBTYPE_PROFILE_LABEL.match(str(label or "").strip())
        if match:
            key = (match["genotype"], match["subtype"].casefold())
            subtypes[key] = int(match["accessions"])
    workbook.close()
    qualifying_count = sum(count >= minimum_accessions for count in subtypes.values())
    return qualifying_count, len(subtypes)


def write_workflow_summary(root: Path, gene: str, destination: Path) -> None:
    """Write the requested QC, RAS, and COMET-assignment summary for the report."""
    qc_dir = step_directory(root, "summarize-qc-mutation-burden")
    with (qc_dir / f"{gene}_QC_Passed_Genotype_Mutation_Burden_Summary.csv").open(
        newline="", encoding="utf-8"
    ) as handle:
        qc_rows = list(csv.DictReader(handle))
    total_accessions = sum(int(row["n"]) for row in qc_rows)
    distribution = sorted(qc_rows, key=lambda row: (-int(row["n"]), row["genotype"]))
    distribution_text = ", ".join(
        f"{row['genotype']} ({format_significant_percent(int(row['n']) / total_accessions)}, {int(row['n'])})"
        for row in distribution
    )

    profile_counts_dir = step_directory(root, "report-profile-input-counts")
    profile_counts_text = (profile_counts_dir / "profile_input_counts.json").read_text(
        encoding="utf-8"
    )
    profile_counts = json.loads(profile_counts_text[profile_counts_text.index("{") :])
    complete_profile_total = int(profile_counts["included_accession_count"])
    complete_profile_distribution = profile_counts.get("included_genotype_distribution")
    if complete_profile_distribution is None:
        qc_pass_dir = step_directory(root, "build-complete-profiles")
        with (qc_pass_dir / f"{gene}_Profile_Accessions_QC_Pass.csv").open(
            newline="", encoding="utf-8"
        ) as handle:
            complete_profile_distribution = {}
            for row in csv.DictReader(handle):
                genotype = (row.get("genotype") or "").strip()
                if genotype:
                    complete_profile_distribution[genotype] = (
                        int(complete_profile_distribution.get(genotype, 0)) + 1
                    )
    complete_profile_distribution_text = ", ".join(
        f"GT{genotype} ({format_significant_percent(int(count) / complete_profile_total)}, {int(count)})"
        for genotype, count in sorted(
            complete_profile_distribution.items(),
            key=lambda item: (-int(item[1]), item[0]),
        )
    )

    combined_dir = step_directory(root, "build-combined-ras-reports")
    with (combined_dir / "last_run_summary.json").open(encoding="utf-8") as handle:
        combined_summary = json.load(handle)
    high_mean_diff = combined_summary.get("mean_diff_at_least_2_5_subtypes", [])
    mean_diff_text = ", ".join(
        f"{str(item['subtype']).split('_', 1)[-1].split(maxsplit=1)[0]} ({item['mean_diff']:.1f})"
        for item in high_mean_diff
    ) or "none"
    coverage_path = combined_dir / f"{gene}_Subtype_RAS_Coverage_Report.xlsx"
    coverage_workbook = load_workbook(coverage_path, read_only=True, data_only=True)
    coverage_subtype_count = coverage_workbook["Subtype_Summary"].max_row - 1
    coverage_workbook.close()

    subtype_profile_path = (
        step_directory(root, "build-subtype-ras-profile")
        / f"{gene}_Subtype_RAS_Profiles.xlsx"
    )
    qualifying_subtype_count, subtype_profile_count = (
        count_gt1_to_gt6_subtypes_with_minimum_accessions(subtype_profile_path)
    )

    unassigned_path = (
        step_directory(root, "prepare-comet-assignments")
        / f"{gene}_Comet_Unassigned_Accession_Count.txt"
    )
    unassigned_count = unassigned_path.read_text(encoding="utf-8").strip()

    summary_path = destination / "Workflow_Summary.md"
    summary_path.write_text(
        "\n".join(
            (
                f"# {gene} COMET workflow summary",
                "",
                "## Complete-profile inputs",
                "",
                f"included_accession_count={complete_profile_total}",
                f"Complete-profile genotype distribution: {complete_profile_distribution_text}",
                "",
                "## Subtype-profile representation",
                "",
                (
                    "GT1-GT6 subtypes with at least 10 sequence accessions: "
                    f"{qualifying_subtype_count} of {subtype_profile_count}"
                ),
                "",
                "## Combined RAS reports",
                "",
                f"Subtypes with MeanDiff >= 2.5: {mean_diff_text}",
                f"Mean MeanDiff across subtypes: {combined_summary.get('mean_subtype_mean_diff', 0):.1f}",
                (
                    "Wrote "
                    f"{coverage_path.relative_to(root.parent)} "
                    f"({coverage_subtype_count} subtypes)"
                ),
                "",
                "## COMET assignments",
                "",
                f"comet_unassigned_accession_count={unassigned_count}",
                "",
            )
        ),
        encoding="utf-8",
    )



def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workflow-output-dir", type=Path, required=True)
    parser.add_argument("--gene", choices=tuple(README_GENE), required=True)
    args = parser.parse_args()
    root = args.workflow_output_dir

    def artifact(steps: tuple[str, ...], filename: str) -> Path:
        for step in steps:
            matches = sorted(root.glob(f"*_{step}/{filename}"))
            if len(matches) == 1:
                return matches[0]
        raise SystemExit(f"Missing {' or '.join(steps)}/{filename} under {root}")

    gene = args.gene

    def artifacts(step: str, pattern: str) -> list[Path]:
        step_dirs = sorted(root.glob(f"*_{step}"))
        if len(step_dirs) != 1:
            raise SystemExit(
                f"Expected one {step} directory under {root}, found {len(step_dirs)}"
            )
        matches = sorted(step_dirs[0].glob(pattern))
        if not matches:
            raise SystemExit(f"Missing {pattern} under {step_dirs[0]}")
        return matches

    sources = (
        (
            Path(__file__).resolve().parents[2] / "profile_accession_aa_call_handling.md",
            "Profile_Accession_AA_Call_Handling.md",
        ),
        (
            artifact(
                ("compare-reference-consensus", "publish-ictv-report"),
                f"README_Subtype_Consensus_Mutations_{README_GENE[gene]}.docx",
            ),
            f"README_Subtype_Consensus_Mutations_{README_GENE[gene]}.docx",
        ),
        (
            artifact(
                ("add-nonconsensus-row",),
                f"{gene}_Combined_RAS_Profiles_Annotated.xlsx",
            ),
            f"{gene}_Combined_RAS_Profiles_Annotated.xlsx",
        ),
        (
            artifact(
                ("build-combined-ras-reports",),
                f"{gene}_Combined_RAS_Profiles.xlsx",
            ),
            f"{gene}_Profile_GE10PerSubtype.xlsx",
        ),
        (
            artifact(
                ("build-subtype-ras-profile",),
                f"{gene}_Subtype_RAS_Profiles.xlsx",
            ),
            f"{gene}_Profile_AllSubtypes.xlsx",
        ),
        (
            artifact(
                ("merge-subtype-complete-profiles",),
                f"{gene}_Subtype_CompleteProfiles_Merged.xlsx",
            ),
            f"{gene}_CompleteProfile.xlsx",
        ),
        (
            artifact(
                ("merge-subtype-complete-profiles",),
                f"{gene}_Profile_Accession_AA_Calls.csv",
            ),
            f"{gene}_Profile_Accession_AA_Calls.csv",
        ),
        (
            artifact(
                ("analyze-genotype-subtype-aa-predictability",),
                f"{gene}_Genotype_Subtype_AA_Predictability.xlsx",
            ),
            f"{gene}_Genotype_Subtype_AA_Predictability.xlsx",
        ),
        *(
            (source, source.name)
            for source in artifacts(
                "build-paired-distance-matrices",
                f"{gene}_*Distance_*.xlsx",
            )
        ),
    )
    destination = root / "report"
    destination.mkdir(parents=True, exist_ok=True)
    for legacy_name in (
        f"{gene}_Combined_RAS_Profiles.xlsx",
        f"{gene}_Subtype_RAS_Profiles.xlsx",
        f"{gene}_Subtype_CompleteProfiles_Merged.xlsx",
    ):
        legacy_path = destination / legacy_name
        if legacy_path.exists():
            legacy_path.unlink()
    for source, destination_name in sources:
        report_path = destination / destination_name
        shutil.copy2(source, report_path)
        if "_Distance_" in destination_name:
            sequence_type = "amino-acid" if "_AA_Distance_" in destination_name else "nucleotide"
            add_distance_notes(report_path, sequence_type)

    write_workflow_summary(root, gene, destination)


if __name__ == "__main__":
    main()
