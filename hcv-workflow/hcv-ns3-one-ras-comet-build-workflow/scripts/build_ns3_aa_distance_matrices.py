#!/usr/bin/env python3
"""Write NS3 RAS-only pairwise-average amino-acid distance matrices."""

from __future__ import annotations
import argparse, csv
from collections import Counter, defaultdict
from pathlib import Path
from statistics import median
from openpyxl import Workbook, load_workbook

RAS = (36, 41, 43, 54, 55, 56, 80, 122, 155, 156, 158, 166, 168, 170, 175)
VALID = set("ACDEFGHIKLMNPQRSTVWY")


def args():
    p = argparse.ArgumentParser()
    p.add_argument("--input-workbook", required=True)
    p.add_argument("--profile-accessions-csv", required=True)
    p.add_argument("--gt-output-xlsx", required=True)
    p.add_argument("--subtype-output-xlsx", required=True)
    p.add_argument("--min-subtype-sequences", type=int, default=10)
    p.add_argument("--start", type=int)
    p.add_argument("--end", type=int)
    return p.parse_args()


def profile_accessions(path):
    with path.open(encoding="utf-8-sig", newline="") as f:
        return {
            r["accession"].strip()
            for r in csv.DictReader(f)
            if r.get("accession", "").strip()
        }


def load(path, allowed):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    h = [str(c.value or "") for c in next(ws.iter_rows(max_row=1))]
    i = {v: n for n, v in enumerate(h)}
    required = {
        "AccessionID",
        "ClosestGT",
        "ClosestSubtype",
        "StartAAPosition",
        "AASequence",
    }
    if missing := required - i.keys():
        raise RuntimeError(f"Missing columns: {', '.join(sorted(missing))}")
    groups = {"gt": defaultdict(list), "subtype": defaultdict(list)}
    excluded = Counter()
    for r in ws.iter_rows(min_row=2, values_only=True):
        accession = str(r[i["AccessionID"]] or "").strip()
        gt = str(r[i["ClosestGT"]] or "").strip()
        st = str(r[i["ClosestSubtype"]] or "").strip().lower()
        start = r[i["StartAAPosition"]]
        seq = str(r[i["AASequence"]] or "").upper()
        if accession not in allowed:
            excluded["not_in_profile_accession_set"] += 1
            continue
        if not gt:
            excluded["missing_genotype_assignment"] += 1
            continue
        if not st:
            excluded["missing_subtype_assignment"] += 1
            continue
        if not start or not seq:
            excluded["missing_AA_sequence_or_start_position"] += 1
            continue
        calls = {int(start) + n: aa for n, aa in enumerate(seq)}
        if any(pos not in calls for pos in RAS):
            excluded["does_not_cover_all_RAS_positions"] += 1
            continue
        if any(calls[pos] not in VALID for pos in RAS):
            excluded["missing_or_ambiguous_AA_at_RAS_position"] += 1
            continue
        values = tuple(calls[pos] for pos in RAS)
        groups["gt"][f"GT{gt}"].append(values)
        groups["subtype"][f"GT{gt}_{st}"].append(values)
    wb.close()
    return groups, excluded


def distance(a, b):
    diff = compared = 0
    for p in range(len(RAS)):
        ca = Counter(x[p] for x in a)
        cb = Counter(x[p] for x in b)
        if a is b:
            n = sum(ca.values())
            total = n * (n - 1) // 2
            same = sum(v * (v - 1) // 2 for v in ca.values())
        else:
            total = len(a) * len(b)
            same = sum(ca[x] * cb[x] for x in VALID)
        compared += total
        diff += total - same
    return diff / compared if compared else None


def exclusions(wb, rows):
    ws = wb.create_sheet("excluded_sequences")
    ws.append(["Reason", "SequenceCount"])
    for k, v in sorted(rows.items()):
        ws.append([k, v])
    ws.append(["total_excluded", sum(rows.values())])


def matrix(wb, title, groups, within_genotype_subtypes=False):
    labels = sorted(groups)
    ws = wb.create_sheet(title)
    ws.append(["Group", *labels])
    same_genotype, different_genotype = [], []
    for row_index, label_a in enumerate(labels):
        row = [label_a]
        for column_index, label_b in enumerate(labels):
            value = distance(groups[label_a], groups[label_b])
            if column_index < row_index:
                row.append(None)
            else:
                row.append(value)
                if value is not None:
                    if within_genotype_subtypes:
                        (same_genotype if row_index == column_index else different_genotype).append(value)
                    elif not within_genotype_subtypes:
                        (same_genotype if row_index == column_index else different_genotype).append(value)
        ws.append(row)
    ws.append([])
    if within_genotype_subtypes:
        ws.append(["Distance comparison", "Mean", "Median", "Pair count"])
        for label, distances in (("Same subtype", same_genotype), ("Different subtype", different_genotype)):
            ws.append([label, sum(distances) / len(distances) if distances else None, median(distances) if distances else None, len(distances)])
    else:
        ws.append(["Distance comparison", "Mean", "Median", "Pair count"])
        for label, distances in (("Same genotype", same_genotype), ("Different genotype", different_genotype)):
            ws.append([label, sum(distances) / len(distances) if distances else None, median(distances) if distances else None, len(distances)])
    for row in ws.iter_rows(min_row=2, min_col=2):
        for c in row:
            c.number_format = "0.0%"
    return labels


def main():
    global RAS
    a = args()
    if (a.start is None) != (a.end is None):
        raise SystemExit("--start and --end must be specified together")
    if a.start is not None:
        if a.start < 1 or a.end < a.start or a.end > 631:
            raise SystemExit("--start/--end must be a valid NS3 amino-acid range")
        RAS = tuple(range(a.start, a.end + 1))
    groups, excluded = load(
        Path(a.input_workbook), profile_accessions(Path(a.profile_accessions_csv))
    )
    for path, kind, minimum in (
        (Path(a.gt_output_xlsx), "gt", 1),
        (Path(a.subtype_output_xlsx), "subtype", a.min_subtype_sequences),
    ):
        wb = Workbook()
        wb.remove(wb.active)
        if kind == "gt":
            matrix(wb, "distance_matrix", groups[kind])
        else:
            bygt = defaultdict(dict)
            for label, seqs in groups[kind].items():
                gt, _, st = label.partition("_")
                bygt[gt][st] = seqs
            for gt in sorted(bygt):
                kept = {st: s for st, s in bygt[gt].items() if len(s) >= minimum}
                if kept:
                    matrix(wb, gt, kept, within_genotype_subtypes=True)
        counts = wb.create_sheet("sequence_counts")
        counts.append(["Group", "SequenceCount"])
        for label, seqs in sorted(groups[kind].items()):
            if kind == "gt" or len(seqs) >= minimum:
                counts.append([label, len(seqs)])
        meta = wb.create_sheet("metadata")
        meta.append(
            [
                "distance_definition",
                "mean pairwise AA differences across complete NS3 RAS positions",
            ]
        )
        meta.append(["aa_positions", ",".join(map(str, RAS))])
        meta.append(["minimum_subtype_sequences", minimum])
        exclusions(wb, excluded)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)


if __name__ == "__main__":
    main()
