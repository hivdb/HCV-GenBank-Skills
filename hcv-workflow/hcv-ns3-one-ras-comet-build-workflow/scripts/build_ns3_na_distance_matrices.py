#!/usr/bin/env python3
"""Write full NS3 nucleotide pairwise-average distance matrices by genotype and subtype."""

from __future__ import annotations
import argparse
import csv
from collections import Counter, defaultdict
from pathlib import Path
from statistics import median
from openpyxl import Workbook, load_workbook

NS3_NT_LENGTH = 631 * 3
BASES = "ACGT"
RAS_AA_POSITIONS = (36, 41, 43, 54, 55, 56, 80, 122, 155, 156, 158, 166, 168, 170, 175)


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--input-workbook", required=True)
    p.add_argument(
        "--profile-accessions-csv",
        required=True,
        help="Accession list written by the profile builder",
    )
    p.add_argument("--gt-output-xlsx", required=True)
    p.add_argument("--subtype-output-xlsx", required=True)
    p.add_argument("--min-subtype-sequences", type=int, default=10)
    p.add_argument(
        "--start", type=int, help="Optional 1-based NS3 amino-acid range start"
    )
    p.add_argument(
        "--end", type=int, help="Optional 1-based NS3 amino-acid range end, inclusive"
    )
    return p.parse_args()


def load_profile_accessions(path):
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return {
            row["accession"].strip()
            for row in csv.DictReader(handle)
            if row.get("accession", "").strip()
        }


def load_groups(path, profile_accessions, positions_nt):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(x.value or "") for x in next(ws.iter_rows(max_row=1))]
    ix = {x: i for i, x in enumerate(header)}
    required = {
        "AccessionID",
        "ClosestGT",
        "ClosestSubtype",
        "StartAAPosition",
        "NASequence",
    }
    if missing := required - ix.keys():
        raise RuntimeError(f"Missing columns: {', '.join(sorted(missing))}")
    groups = {"gt": defaultdict(list), "subtype": defaultdict(list)}
    excluded = Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        accession = str(row[ix["AccessionID"]] or "").strip()
        if accession not in profile_accessions:
            excluded["not_in_profile_accession_set"] += 1
            continue
        gt = str(row[ix["ClosestGT"]] or "").strip()
        subtype = str(row[ix["ClosestSubtype"]] or "").strip().lower()
        start = row[ix["StartAAPosition"]]
        seq = str(row[ix["NASequence"]] or "").upper()
        if not gt:
            excluded["missing_genotype_assignment"] += 1
            continue
        if not subtype:
            excluded["missing_subtype_assignment"] += 1
            continue
        if not start or not seq:
            excluded["missing_NA_sequence_or_start_position"] += 1
            continue
        projected = ["-"] * NS3_NT_LENGTH
        offset = (int(start) - 1) * 3
        for i, base in enumerate(seq[: NS3_NT_LENGTH - offset]):
            projected[offset + i] = base if base in BASES else "-"
        covers_window = offset <= min(positions_nt) and offset + len(seq) > max(
            positions_nt
        )
        if not covers_window:
            excluded["does_not_cover_all_RAS_positions"] += 1
            continue
        if any(projected[position] not in BASES for position in positions_nt):
            excluded["missing_or_ambiguous_base_in_range"] += 1
            continue
        groups["gt"][f"GT{gt}"].append(projected)
        groups["subtype"][f"GT{gt}_{subtype}"].append(projected)
    wb.close()
    return groups, excluded


def add_exclusion_sheet(wb, excluded):
    ws = wb.create_sheet("excluded_sequences")
    ws.append(["Reason", "SequenceCount"])
    for reason, count in sorted(excluded.items()):
        ws.append([reason, count])
    ws.append(["total_excluded", sum(excluded.values())])


def distance(a, b, positions_nt):
    different = 0
    compared = 0
    for pos in positions_nt:
        ca = Counter(s[pos] for s in a)
        cb = Counter(s[pos] for s in b)
        if a is b:
            comparable_pairs = sum(ca[x] for x in BASES)
            comparable_pairs = comparable_pairs * (comparable_pairs - 1) // 2
            matching_pairs = sum(ca[x] * (ca[x] - 1) // 2 for x in BASES)
        else:
            comparable_pairs = sum(ca[x] for x in BASES) * sum(cb[x] for x in BASES)
            matching_pairs = sum(ca[x] * cb[x] for x in BASES)
        compared += comparable_pairs
        different += comparable_pairs - matching_pairs
    return different / compared if compared else None


def write(path, groups, minimum, positions_nt, excluded):
    kept = {k: v for k, v in groups.items() if len(v) >= minimum}
    labels = sorted(kept)
    wb = Workbook()
    ws = wb.active
    ws.title = "distance_matrix"
    ws.append(["Group", *labels])
    same_genotype: list[float] = []
    different_genotype: list[float] = []
    for row_index, label_a in enumerate(labels):
        row = [label_a]
        for column_index, label_b in enumerate(labels):
            value = distance(kept[label_a], kept[label_b], positions_nt)
            if column_index < row_index:
                row.append(None)
            else:
                row.append(value)
                if value is not None:
                    (same_genotype if row_index == column_index else different_genotype).append(value)
        ws.append(row)
    ws.append([])
    ws.append(["Distance comparison", "Mean", "Median", "Pair count"])
    for label, distances in (("Same genotype", same_genotype), ("Different genotype", different_genotype)):
        ws.append([label, sum(distances) / len(distances) if distances else None, median(distances) if distances else None, len(distances)])
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = "0.0%"
    counts = wb.create_sheet("sequence_counts")
    counts.append(["Group", "SequenceCount"])
    for label in labels:
        counts.append([label, len(kept[label])])
    meta = wb.create_sheet("metadata")
    meta.append(
        [
            "distance_definition",
            "aggregate pairwise NA differences / aggregate comparable NA positions; missing bases are excluded",
        ]
    )
    meta.append(["aa_positions", ",".join(map(str, RAS_AA_POSITIONS))])
    meta.append(["na_positions", ",".join(str(pos + 1) for pos in positions_nt)])
    add_exclusion_sheet(wb, excluded)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return len(labels)


def write_subtypes_by_genotype(path, groups, minimum, positions_nt, excluded):
    by_gt = defaultdict(dict)
    for label, sequences in groups.items():
        genotype, _, subtype = label.partition("_")
        by_gt[genotype][subtype] = sequences
    wb = Workbook()
    wb.remove(wb.active)
    sheet_count = 0
    for genotype in sorted(
        by_gt, key=lambda value: (int(value[2:]) if value[2:].isdigit() else 999, value)
    ):
        kept = {
            subtype: sequences
            for subtype, sequences in by_gt[genotype].items()
            if len(sequences) >= minimum
        }
        if not kept:
            continue
        labels = sorted(kept)
        ws = wb.create_sheet(genotype)
        ws.append(["Subtype", *labels])
        same_subtype: list[float] = []
        different_subtype: list[float] = []
        for row_index, subtype_a in enumerate(labels):
            row = [subtype_a]
            for column_index, subtype_b in enumerate(labels):
                value = distance(kept[subtype_a], kept[subtype_b], positions_nt)
                if column_index < row_index:
                    row.append(None)
                else:
                    row.append(value)
                    if value is not None:
                        (same_subtype if row_index == column_index else different_subtype).append(value)
            ws.append(row)
        ws.append([])
        ws.append(["Distance comparison", "Mean", "Median", "Pair count"])
        for label, distances in (("Same subtype", same_subtype), ("Different subtype", different_subtype)):
            ws.append([label, sum(distances) / len(distances) if distances else None, median(distances) if distances else None, len(distances)])
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                cell.number_format = "0.0%"
        counts = wb.create_sheet(f"{genotype}_counts")
        counts.append(["Subtype", "SequenceCount"])
        for subtype in labels:
            counts.append([subtype, len(kept[subtype])])
        sheet_count += 1
    if sheet_count == 0:
        raise RuntimeError(f"No subtype groups had at least {minimum} usable sequences")
    meta = wb.create_sheet("metadata")
    meta.append(
        [
            "distance_definition",
            "aggregate pairwise NA differences / aggregate comparable NA positions; missing bases are excluded",
        ]
    )
    meta.append(["aa_positions", ",".join(map(str, RAS_AA_POSITIONS))])
    meta.append(["na_positions", ",".join(str(pos + 1) for pos in positions_nt)])
    meta.append(["minimum_subtype_sequences", minimum])
    add_exclusion_sheet(wb, excluded)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return sheet_count


def main():
    global RAS_AA_POSITIONS
    a = parse_args()
    if (a.start is None) != (a.end is None):
        raise SystemExit("--start and --end must be specified together")
    if a.start is not None:
        if a.start < 1 or a.end < a.start or a.end > 631:
            raise SystemExit("--start/--end must be a valid NS3 amino-acid range")
        RAS_AA_POSITIONS = tuple(range(a.start, a.end + 1))
    positions_nt = tuple(
        position for aa in RAS_AA_POSITIONS for position in range((aa - 1) * 3, aa * 3)
    )
    groups, excluded = load_groups(
        Path(a.input_workbook),
        load_profile_accessions(Path(a.profile_accessions_csv)),
        positions_nt,
    )
    gt = write(Path(a.gt_output_xlsx), groups["gt"], 1, positions_nt, excluded)
    st = write_subtypes_by_genotype(
        Path(a.subtype_output_xlsx),
        groups["subtype"],
        a.min_subtype_sequences,
        positions_nt,
        excluded,
    )
    print(f"gt_group_count={gt}\nsubtype_genotype_sheet_count={st}")


if __name__ == "__main__":
    main()
