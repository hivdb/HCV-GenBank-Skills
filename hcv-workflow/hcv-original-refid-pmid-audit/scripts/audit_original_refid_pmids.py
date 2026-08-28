#!/usr/bin/env python3
"""Audit Original worksheet PMIDs against blank PMID entries in Ref.csv."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[3]
SOURCE_SHEETS = ("Original_NS3", "Original_NS5A", "Original_NS5B")
COMBINED_COLUMNS = ("SourceSheet", "SourceRow", "RefID", "PMID")
FOUND_COLUMNS = (*COMBINED_COLUMNS, "RefCsvMedlineID")
SUMMARY_COLUMNS = (
    "SourceSheet",
    "Gene",
    "Total rows",
    "Final with PMID",
    "Found PMID",
    "PMID from genbank",
)


def nonblank(value: object) -> str:
    """Return a stripped value, treating None and whitespace as blank."""
    return "" if value is None else str(value).strip()


def read_original_sheets(
    workbook_path: Path,
) -> tuple[dict[str, list[dict[str, str]]], dict[str, dict[str, int]]]:
    """Read eligible RefID/PMID pairs and source-sheet denominators."""
    rows_by_sheet: dict[str, list[dict[str, str]]] = {}
    counts_by_sheet: dict[str, dict[str, int]] = {}
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for sheet_name in SOURCE_SHEETS:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Workbook is missing required sheet: {sheet_name}")

            worksheet = workbook[sheet_name]
            header_cells = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
            )
            headers = {
                nonblank(value): index for index, value in enumerate(header_cells)
            }
            missing = {"RefID", "PMID"} - headers.keys()
            if missing:
                missing_text = ", ".join(sorted(missing))
                raise ValueError(
                    f"{sheet_name} is missing required column(s): {missing_text}"
                )

            sheet_rows: list[dict[str, str]] = []
            total_rows = 0
            final_with_pmid = 0
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                if not any(nonblank(value) for value in row):
                    continue
                total_rows += 1
                refid = nonblank(row[headers["RefID"]])
                pmid = nonblank(row[headers["PMID"]])
                if pmid:
                    final_with_pmid += 1
                if refid and pmid:
                    sheet_rows.append(
                        {
                            "SourceSheet": sheet_name,
                            "SourceRow": str(row_number),
                            "RefID": refid,
                            "PMID": pmid,
                        }
                    )
            rows_by_sheet[sheet_name] = sheet_rows
            counts_by_sheet[sheet_name] = {
                "Total rows": total_rows,
                "Final with PMID": final_with_pmid,
            }
    finally:
        workbook.close()
    return rows_by_sheet, counts_by_sheet


def ref_csv_pmids(ref_csv_path: Path) -> dict[str, str]:
    """Map Ref.csv RefID values to their PMID-bearing MedlineID values."""
    with ref_csv_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ValueError(f"Ref.csv has no header row: {ref_csv_path}")
        missing = {"RefID", "MedlineID"} - set(reader.fieldnames)
        if missing:
            missing_text = ", ".join(sorted(missing))
            raise ValueError(f"Ref.csv is missing required column(s): {missing_text}")

        return {
            nonblank(row["RefID"]): nonblank(row["MedlineID"])
            for row in reader
            if nonblank(row["RefID"])
        }


def write_csv(path: Path, columns: tuple[str, ...], rows: list[dict[str, str]]) -> None:
    """Write a CSV with a stable header and UTF-8 encoding."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=REPO_ROOT / "HCVData/HCV_BlastHists_202604_data.xlsx",
        help="Workbook containing the Original_NS3, Original_NS5A, and Original_NS5B sheets.",
    )
    parser.add_argument(
        "--ref-csv",
        type=Path,
        default=REPO_ROOT / "HCVData/HCV-all-seq-subtype/Ref.csv",
        help="Ref.csv whose MedlineID column is treated as the PMID field.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=REPO_ROOT / "outputs/hcv-original-refid-pmid-audit",
        help="Directory for generated CSV files.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows_by_sheet, counts_by_sheet = read_original_sheets(args.workbook)
    combined_rows = [row for sheet_rows in rows_by_sheet.values() for row in sheet_rows]
    ref_pmids = ref_csv_pmids(args.ref_csv)
    found_rows_by_sheet = {
        sheet_name: [
            {**row, "RefCsvMedlineID": ref_pmids[row["RefID"]]}
            for row in rows_by_sheet[sheet_name]
            if row["RefID"] in ref_pmids and not ref_pmids[row["RefID"]]
        ]
        for sheet_name in SOURCE_SHEETS
    }
    found_rows = [
        row for sheet_rows in found_rows_by_sheet.values() for row in sheet_rows
    ]
    summary_rows: list[dict[str, str]] = []
    for sheet_name in SOURCE_SHEETS:
        counts = counts_by_sheet[sheet_name]
        found_count = len(found_rows_by_sheet[sheet_name])
        report = {
            "SourceSheet": sheet_name,
            "Gene": sheet_name.removeprefix("Original_"),
            "Total rows": str(counts["Total rows"]),
            "Final with PMID": str(counts["Final with PMID"]),
            "Found PMID": str(found_count),
            "PMID from genbank": str(counts["Final with PMID"] - found_count),
        }
        summary_rows.append(report)
        gene_dir = args.output_dir / report["Gene"]
        write_csv(
            gene_dir / "found_refids_with_blank_ref_pmid.csv",
            FOUND_COLUMNS,
            found_rows_by_sheet[sheet_name],
        )
        write_csv(gene_dir / "report.csv", SUMMARY_COLUMNS, [report])

    write_csv(
        args.output_dir / "combined_refid_pmid_pairs.csv",
        COMBINED_COLUMNS,
        combined_rows,
    )
    write_csv(
        args.output_dir / "found_refids_with_blank_ref_pmid.csv",
        FOUND_COLUMNS,
        found_rows,
    )
    write_csv(
        args.output_dir / "PMID_found_summary_by_sheet.csv",
        SUMMARY_COLUMNS,
        summary_rows,
    )

    print(f"combined_refid_pmid_pairs={len(combined_rows)}")
    for report in summary_rows:
        print(
            "{Gene}: total_rows={Total rows}; final_with_pmid={Final with PMID}; "
            "found_pmid={Found PMID}; pmid_from_genbank={PMID from genbank}".format(
                **report
            )
        )
    print(f"found_total={len(found_rows)}")


if __name__ == "__main__":
    main()
