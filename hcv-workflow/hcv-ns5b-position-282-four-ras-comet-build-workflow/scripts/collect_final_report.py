#!/usr/bin/env python3
"""Copy final workflow artifacts into its output report directory."""

from __future__ import annotations
import argparse
import shutil
from pathlib import Path

from openpyxl import load_workbook

README_GENE = {"NS3": "NS3", "NS5A": "NS5A_NTD", "NS5B": "NS5B"}


def add_distance_notes(path: Path) -> None:
    """Add an explanation of the sequence-quality exclusion rule."""
    workbook = load_workbook(path)
    if "Notes" in workbook.sheetnames:
        del workbook["Notes"]
    notes = workbook.create_sheet("Notes")
    notes.append(["Note"])
    notes.append(
        [
            "Sequences with an ambiguous or missing amino-acid call at any "
            "required comparison position are excluded from this distance matrix."
        ]
    )
    if "excluded_sequences" in workbook.sheetnames:
        del workbook["excluded_sequences"]
    notes.column_dimensions["A"].width = 110
    notes.freeze_panes = "A2"
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workflow-output-dir", type=Path, required=True)
    parser.add_argument("--gene", choices=tuple(README_GENE), required=True)
    args = parser.parse_args()
    root = args.workflow_output_dir

    def artifact(steps: tuple[str, ...], filename: str) -> Path:
        for step in steps:
            matches = sorted(root.glob(f"*_{step}/{filename}"))
            if len(matches) == 1:
                return matches[0]
        raise SystemExit(f"Missing {' or '.join(steps)}/{filename} under {root}")

    gene = args.gene

    def artifacts(step: str, pattern: str) -> list[Path]:
        step_dirs = sorted(root.glob(f"*_{step}"))
        if len(step_dirs) != 1:
            raise SystemExit(
                f"Expected one {step} directory under {root}, found {len(step_dirs)}"
            )
        matches = sorted(step_dirs[0].glob(pattern))
        if not matches:
            raise SystemExit(f"Missing {pattern} under {step_dirs[0]}")
        return matches

    sources = (
        (
            artifact(
                ("compare-reference-consensus", "publish-ictv-report"),
                f"README_Subtype_Consensus_Mutations_{README_GENE[gene]}.docx",
            ),
            f"README_Subtype_Consensus_Mutations_{README_GENE[gene]}.docx",
        ),
        (
            artifact(
                ("add-nonconsensus-row",),
                f"{gene}_Combined_RAS_Profiles_Annotated.xlsx",
            ),
            f"{gene}_Combined_RAS_Profiles_Annotated.xlsx",
        ),
        (
            artifact(
                ("build-combined-ras-reports",),
                f"{gene}_Combined_RAS_Profiles.xlsx",
            ),
            "Table1.xlsx",
        ),
        (
            artifact(
                ("build-subtype-ras-profile",),
                f"{gene}_Subtype_RAS_Profiles.xlsx",
            ),
            "S_Table1.xlsx",
        ),
        (
            artifact(
                ("merge-subtype-complete-profiles",),
                f"{gene}_Subtype_CompleteProfiles_Merged.xlsx",
            ),
            "S_file1.xlsx",
        ),
        (
            artifact(
                ("analyze-genotype-subtype-aa-predictability",),
                f"{gene}_Genotype_Subtype_AA_Predictability.xlsx",
            ),
            f"{gene}_Genotype_Subtype_AA_Predictability.xlsx",
        ),
        *(
            (source, source.name)
            for source in artifacts(
                "build-paired-distance-matrices",
                f"{gene}_*AA_Distance_*.xlsx",
            )
        ),
    )
    destination = root / "report"
    destination.mkdir(parents=True, exist_ok=True)
    for legacy_name in (
        f"{gene}_Combined_RAS_Profiles.xlsx",
        f"{gene}_Subtype_RAS_Profiles.xlsx",
        f"{gene}_Subtype_CompleteProfiles_Merged.xlsx",
    ):
        legacy_path = destination / legacy_name
        if legacy_path.exists():
            legacy_path.unlink()
    for source, destination_name in sources:
        report_path = destination / destination_name
        shutil.copy2(source, report_path)
        if "_AA_Distance_" in destination_name:
            add_distance_notes(report_path)


if __name__ == "__main__":
    main()
