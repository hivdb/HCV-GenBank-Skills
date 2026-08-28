#!/usr/bin/env python3
"""Compare NS5B workflow GT7/GT8 subtype calls with local NS5B assignments."""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

GENE = "NS5B"
REFERENCE_SUBTYPES_CSV = (
    Path(__file__).resolve().parents[3]
    / "HCVData/Subtype-Ref/HCV_Subtype_Refs_AA_Accession_Subtype.csv"
)
ALL_COMET_SUBTYPE_CSV = (
    Path(__file__).resolve().parents[3]
    / "HCVData/HCV-all-seq-subtype/all_comet_subtype.csv"
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--gene", help=argparse.SUPPRESS)
    parser.add_argument("--subtype-workbook", required=True)
    parser.add_argument("--coverage-csv", required=True)
    parser.add_argument(
        "--reference-subtypes-csv", type=Path, default=REFERENCE_SUBTYPES_CSV
    )
    parser.add_argument(
        "--all-comet-subtype-csv", type=Path, default=ALL_COMET_SUBTYPE_CSV
    )
    parser.add_argument("--profile-accessions-csv", type=Path)
    parser.add_argument("--output-xlsx", required=True)
    parser.add_argument("--output-csv", required=True)
    return parser.parse_args()


def accession_key(value: object) -> str:
    return str(value or "").strip().split(".", 1)[0].upper()


def normalized_genotype(value: object) -> str:
    return str(value or "").strip().lower().removeprefix("gt")


def normalized_subtype(value: object) -> str:
    return str(value or "").strip().lower().removeprefix("gt")


def read_workflow_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header = [str(value or "") for value in next(worksheet.iter_rows(values_only=True))]
    required = {
        "RefID",
        "RefName",
        "AccessionID",
        "ClosestGT",
        "ClosestSubtype",
        "ClosestSubtypeAssignmentSource",
        "ClosestSubtypeMetadataColumn",
    }
    missing = sorted(required - set(header))
    if missing:
        raise RuntimeError(f"{path} is missing required columns: {', '.join(missing)}")

    rows: list[dict[str, str]] = []
    for values in worksheet.iter_rows(min_row=2, values_only=True):
        row = {
            column: str(value or "").strip() for column, value in zip(header, values)
        }
        if normalized_genotype(row["ClosestGT"]) in {"7", "8"}:
            rows.append(row)
    workbook.close()
    return rows


def read_local_assignments(path: Path) -> dict[str, list[dict[str, str]]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"Accession", "ClosestGenotype", "ClosestSubtype"}
        missing = sorted(required - set(reader.fieldnames or []))
        if missing:
            raise RuntimeError(
                f"{path} is missing required columns: {', '.join(missing)}"
            )
        assignments: dict[str, list[dict[str, str]]] = defaultdict(list)
        for row in reader:
            accession = str(row.get("Accession") or "").strip()
            key = accession_key(accession)
            if key:
                assignments[key].append(
                    {
                        "accession": accession,
                        "genotype": str(row.get("ClosestGenotype") or "").strip(),
                        "subtype": str(row.get("ClosestSubtype") or "").strip(),
                        "subtype_reference_accession": "",
                        "subtype_aligned_nt": "",
                        "subtype_pident": str(
                            row.get("ClosestSubtypePident") or ""
                        ).strip(),
                    }
                )
    return assignments


def read_reference_accessions(path: Path) -> set[str]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames or "Accession" not in reader.fieldnames:
            raise RuntimeError(f"{path} is missing the Accession column")
        return {
            accession_key(row.get("Accession"))
            for row in reader
            if accession_key(row.get("Accession"))
        }


def read_all_comet_subtypes(path: Path) -> dict[str, dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"name", "virus", "subtype", "bootstrap support"}
        missing = sorted(required - set(reader.fieldnames or []))
        if missing:
            raise RuntimeError(
                f"{path} is missing required columns: {', '.join(missing)}"
            )
        return {
            accession_key(row.get("name")): {
                "CometVirus": str(row.get("virus") or "").strip(),
                "CometSubtype": str(row.get("subtype") or "").strip(),
                "CometSubtypeBootstrapSupport": str(
                    row.get("bootstrap support") or ""
                ).strip(),
            }
            for row in reader
            if accession_key(row.get("name"))
        }


def read_profile_accessions(path: Path) -> set[str]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames or "accession" not in reader.fieldnames:
            raise RuntimeError(f"{path} is missing the accession column")
        return {
            accession_key(row.get("accession"))
            for row in reader
            if accession_key(row.get("accession"))
        }


def compare_rows(
    workflow_rows: list[dict[str, str]],
    local_assignments: dict[str, list[dict[str, str]]],
    reference_accessions: set[str],
    all_comet_subtypes: dict[str, dict[str, str]],
    profile_accessions: set[str],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for workflow in workflow_rows:
        accession = workflow["AccessionID"]
        local_rows = local_assignments.get(accession_key(accession), [])
        comet_subtype = all_comet_subtypes.get(accession_key(accession), {})
        output = {
            "RefID": workflow["RefID"],
            "RefName": workflow["RefName"],
            "Accession": accession,
            "IsReferenceSubtypeAccession": "YES"
            if accession_key(accession) in reference_accessions
            else "NO",
            "CometVirus": comet_subtype.get("CometVirus", ""),
            "CometSubtype": comet_subtype.get("CometSubtype", ""),
            "CometSubtypeBootstrapSupport": comet_subtype.get(
                "CometSubtypeBootstrapSupport", ""
            ),
            "in profile": "YES"
            if accession_key(accession) in profile_accessions
            else "NO",
            "WorkflowGenotype": workflow["ClosestGT"],
            "WorkflowSubtype": workflow["ClosestSubtype"],
            "WorkflowAssignmentSource": workflow["ClosestSubtypeAssignmentSource"],
            "WorkflowAssignmentColumn": workflow["ClosestSubtypeMetadataColumn"],
            "LocalAssignmentStatus": "",
            "LocalGenotype": "",
            "LocalSubtype": "",
            "LocalSubtypeReferenceAccession": "",
            "LocalSubtypeAlignedNt": "",
            "LocalSubtypePident": "",
            "GenotypeMatchesLocal": "",
            "SubtypeMatchesLocal": "",
        }
        if not local_rows:
            output.update(
                LocalAssignmentStatus="NOT_FOUND",
                GenotypeMatchesLocal="NOT_FOUND",
                SubtypeMatchesLocal="NOT_FOUND",
            )
        elif len(local_rows) > 1:
            output.update(
                LocalAssignmentStatus="MULTIPLE_LOCAL_ROWS",
                GenotypeMatchesLocal="NOT_COMPARED",
                SubtypeMatchesLocal="NOT_COMPARED",
            )
        else:
            local = local_rows[0]
            local_subtype = local["subtype"]
            output.update(
                LocalAssignmentStatus="FOUND"
                if local_subtype
                else "FOUND_WITHOUT_SUBTYPE",
                LocalGenotype=local["genotype"],
                LocalSubtype=local_subtype,
                LocalSubtypeReferenceAccession=local["subtype_reference_accession"],
                LocalSubtypeAlignedNt=local.get("subtype_aligned_nt", ""),
                LocalSubtypePident=local.get("subtype_pident", ""),
                GenotypeMatchesLocal="YES"
                if normalized_genotype(workflow["ClosestGT"])
                == normalized_genotype(local["genotype"])
                else "NO",
                SubtypeMatchesLocal=(
                    "YES"
                    if normalized_subtype(workflow["ClosestSubtype"])
                    == normalized_subtype(local_subtype)
                    else "NO"
                )
                if local_subtype
                else "NO_LOCAL_SUBTYPE",
            )
        rows.append(output)
    return rows


def write_report(
    rows: list[dict[str, str]], output_xlsx: Path, output_csv: Path
) -> None:
    fields = [
        "Accession",
        "IsReferenceSubtypeAccession",
        "CometSubtype",
        "CometSubtypeBootstrapSupport",
        "in profile",
        "WorkflowGenotype",
        "WorkflowSubtype",
        "LocalSubtypePident",
    ]
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "GT7_GT8_Local_Comparison"
    worksheet.append(fields)
    for row in rows:
        worksheet.append([row[field] for field in fields])
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    summary = workbook.create_sheet("Summary")
    summary.append(["Metric", "Count"])
    summary.append(["GT7/GT8 workflow accessions", len(rows)])
    for field in (
        "LocalAssignmentStatus",
        "GenotypeMatchesLocal",
        "SubtypeMatchesLocal",
    ):
        for value, count in sorted(Counter(row[field] for row in rows).items()):
            summary.append([f"{field}: {value}", count])
    for cell in summary[1]:
        cell.font = Font(bold=True)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)


def main() -> int:
    args = parse_args()
    subtype_workbook = Path(args.subtype_workbook)
    coverage_csv = Path(args.coverage_csv)
    profile_accessions_csv = (
        Path(args.profile_accessions_csv)
        if args.profile_accessions_csv
        else next(
            iter(
                sorted(
                    subtype_workbook.parent.parent.glob(
                        "*_build-complete-profiles/*_Profile_Accessions_QC_Pass.csv"
                    )
                )
            ),
            None,
        )
    )
    if not subtype_workbook.is_file():
        raise SystemExit(f"NS3 subtype workbook was not found: {subtype_workbook}")
    if not coverage_csv.is_file():
        raise SystemExit(f"Non-COMET coverage CSV was not found: {coverage_csv}.")
    if profile_accessions_csv is None or not profile_accessions_csv.is_file():
        raise SystemExit(
            f"Profile RAS-coverage accessions CSV was not found for {subtype_workbook}"
        )
    rows = compare_rows(
        read_workflow_rows(subtype_workbook),
        read_local_assignments(coverage_csv),
        read_reference_accessions(Path(args.reference_subtypes_csv)),
        read_all_comet_subtypes(Path(args.all_comet_subtype_csv)),
        read_profile_accessions(profile_accessions_csv),
    )
    write_report(rows, Path(args.output_xlsx), Path(args.output_csv))
    print(
        json.dumps(
            {
                "gene": GENE,
                "workflow_gt7_gt8_accession_count": len(rows),
                "subtype_match_counts": dict(
                    sorted(Counter(row["SubtypeMatchesLocal"] for row in rows).items())
                ),
                "output_xlsx": str(Path(args.output_xlsx).resolve()),
                "output_csv": str(Path(args.output_csv).resolve()),
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
