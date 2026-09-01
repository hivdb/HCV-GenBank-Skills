"""Export QC-passed RAS amino-acid calls with genotype consensus calls."""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

CANONICAL_AAS = frozenset("ACDEFGHIKLMNPQRSTVWY")
EXPORT_COLUMNS = ["Accession", "Position", "Genotype", "Subtype", "GT_Consensus", "AA"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--profile-input-workbook", required=True)
    parser.add_argument("--profile-accessions-csv", required=True)
    parser.add_argument("--gt-profile-workbook", required=True)
    parser.add_argument(
        "--ras-positions",
        required=True,
        help="Comma-separated amino-acid positions to include in the export.",
    )
    parser.add_argument("--output-csv", required=True)
    return parser.parse_args()


def normalize_gt(value: object) -> str:
    text = str(value or "").strip()
    return text if text.upper().startswith("GT") else f"GT{text}"


def load_profile_accessions(path: Path) -> set[str]:
    with path.open(encoding="utf-8", newline="") as handle:
        return {
            str(row.get("accession") or "").strip()
            for row in csv.DictReader(handle)
            if str(row.get("accession") or "").strip()
        }


def parse_positions(value: str) -> set[int]:
    positions = {int(position.strip()) for position in value.split(",") if position.strip()}
    if not positions:
        raise ValueError("--ras-positions must contain at least one position")
    return positions


def load_gt_consensus(path: Path) -> dict[tuple[str, int], str]:
    """Match the PctWithAA, then lexical-AA, consensus rule of FASTA export."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    calls: dict[tuple[str, int], tuple[float, str]] = {}
    for worksheet in workbook.worksheets:
        header = [
            str(value or "") for value in next(worksheet.iter_rows(values_only=True))
        ]
        index = {name: position for position, name in enumerate(header)}
        position_columns = [
            name
            for name in header
            if name.endswith("Position") and name != "NumSeqsIncludingPosition"
        ]
        if len(position_columns) != 1:
            raise RuntimeError(
                f"Could not identify position column in {worksheet.title}"
            )
        required = [position_columns[0], "AminoAcid", "PctWithAA"]
        missing = [name for name in required if name not in index]
        if missing:
            raise RuntimeError(
                f"Missing columns in {worksheet.title}: {', '.join(missing)}"
            )
        genotype = normalize_gt(worksheet.title)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            position = int(row[index[position_columns[0]]])
            amino_acid = str(row[index["AminoAcid"]] or "").strip().upper()
            candidate = (float(row[index["PctWithAA"]]), amino_acid)
            key = (genotype, position)
            current = calls.get(key)
            if (
                current is None
                or candidate[0] > current[0]
                or (candidate[0] == current[0] and candidate[1] < current[1])
            ):
                calls[key] = candidate
    workbook.close()
    return {key: value[1] for key, value in calls.items()}


def export_calls(args: argparse.Namespace) -> dict[str, object]:
    allowed_accessions = load_profile_accessions(Path(args.profile_accessions_csv))
    ras_positions = parse_positions(args.ras_positions)
    consensus = load_gt_consensus(Path(args.gt_profile_workbook))
    workbook = load_workbook(
        args.profile_input_workbook, read_only=True, data_only=True
    )
    worksheet = workbook[workbook.sheetnames[0]]
    header = [str(value or "") for value in next(worksheet.iter_rows(values_only=True))]
    index = {name: position for position, name in enumerate(header)}
    required = [
        "AccessionID",
        "ClosestGT",
        "ClosestSubtype",
        "StartAAPosition",
        "AASequence",
    ]
    missing = [name for name in required if name not in index]
    if missing:
        raise RuntimeError(f"Missing columns in profile input: {', '.join(missing)}")

    output_path = Path(args.output_csv)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows_written = 0
    skipped_calls: Counter[str] = Counter()
    included_accessions: set[str] = set()
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=EXPORT_COLUMNS)
        writer.writeheader()
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            accession = str(row[index["AccessionID"]] or "").strip()
            if accession not in allowed_accessions:
                continue
            if (
                "AlignmentQCStatus" in index
                and str(row[index["AlignmentQCStatus"]] or "").strip() != "PASS"
            ):
                continue
            start = row[index["StartAAPosition"]]
            sequence = str(row[index["AASequence"]] or "").strip().upper()
            if start in (None, "") or not sequence:
                continue
            genotype = normalize_gt(row[index["ClosestGT"]])
            subtype = str(row[index["ClosestSubtype"]] or "").strip()
            included_accessions.add(accession)
            for offset, amino_acid in enumerate(sequence):
                if amino_acid not in CANONICAL_AAS and amino_acid != "*":
                    skipped_calls[amino_acid or "EMPTY"] += 1
                    continue
                position = int(start) + offset
                if position not in ras_positions:
                    continue
                writer.writerow(
                    {
                        "Accession": accession,
                        "Position": position,
                        "Genotype": genotype,
                        "Subtype": subtype,
                        "GT_Consensus": consensus.get((genotype, position), "X"),
                        "AA": amino_acid,
                    }
                )
                rows_written += 1
    workbook.close()
    return {
        "output_csv": str(output_path.resolve()),
        "rows_written": rows_written,
        "included_accession_count": len(included_accessions),
        "ras_positions": sorted(ras_positions),
        "skipped_nonstandard_aa_calls": dict(sorted(skipped_calls.items())),
    }


def main() -> int:
    print(json.dumps(export_calls(parse_args()), indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
