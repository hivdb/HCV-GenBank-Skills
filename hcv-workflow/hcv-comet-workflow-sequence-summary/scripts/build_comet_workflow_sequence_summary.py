#!/usr/bin/env python3
"""Build the cross-gene HCV COMET workflow sequence-inclusion summary workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


REPO_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_OUTPUT = (
    REPO_ROOT / "outputs/HCV_Comet_Workflow_Sequence_Inclusion_Summary.xlsx"
)
WORKFLOWS = (
    (
        "hcv-ns3-comet-build-workflow",
        "NS3",
        "outputs/comet-NS3",
        "All QC-passed, subtype assigned inputs",
    ),
    (
        "hcv-ns3-one-ras-comet-build-workflow",
        "NS3",
        "outputs/comet-NS3-one-ras",
        "Callable AA at ≥1 NS3 RAS position",
    ),
    (
        "hcv-ns3-all-ras-comet-build-workflow",
        "NS3",
        "outputs/comet-NS3-all-ras",
        "Callable AA at every NS3 RAS position",
    ),
    (
        "hcv-ns5a-comet-build-workflow",
        "NS5A",
        "outputs/comet-NS5A",
        "All QC-passed, subtype assigned inputs",
    ),
    (
        "hcv-ns5a-one-ras-comet-build-workflow",
        "NS5A",
        "outputs/comet-NS5A-one-ras",
        "Callable AA at ≥1 NS5A RAS position",
    ),
    (
        "hcv-ns5a-all-ras-comet-build-workflow",
        "NS5A",
        "outputs/comet-NS5A-all-ras",
        "Callable AA at every NS5A RAS position",
    ),
    (
        "hcv-ns5b-all-ras-comet-build-workflow",
        "NS5B",
        "outputs/comet-NS5B-all-ras",
        "Callable AA at every required RAS position: 150, 159, 206, 282, 316, 320, 321",
    ),
    (
        "hcv-ns5b-position-282-comet-build-workflow",
        "NS5B",
        "outputs/comet-NS5B-position-282",
        "Callable AA at position 282",
    ),
    (
        "hcv-ns5b-position-282-four-ras-comet-build-workflow",
        "NS5B",
        "outputs/comet-NS5B-position-282-four-ras",
        "Callable AA at position 282 plus ≥4 of positions 150, 159, 206, 316, 320, 321",
    ),
)


def included_accession_count(output_directory: str) -> int:
    count_path = (
        REPO_ROOT
        / output_directory
        / "15_report-profile-input-counts/profile_input_counts.json"
    )
    if not count_path.is_file():
        raise FileNotFoundError(f"Missing profile input count file: {count_path}")
    payload = count_path.read_text(encoding="utf-8")
    json_start = payload.find("{")
    if json_start < 0:
        raise ValueError(f"No JSON payload found in {count_path}")
    try:
        data = json.loads(payload[json_start:])
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid JSON payload in {count_path}") from exc
    try:
        return int(data["included_accession_count"])
    except (KeyError, TypeError, ValueError) as exc:
        raise ValueError(f"Invalid included_accession_count in {count_path}") from exc


def build_workbook(output_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Workflow Summary"
    worksheet.append(
        [
            "Workflow name",
            "Gene",
            "Workflow include-sequence method for combined profile",
            "Included accessions",
        ]
    )
    for workflow_name, gene, output_directory, method in WORKFLOWS:
        worksheet.append(
            [workflow_name, gene, method, included_accession_count(output_directory)]
        )

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[3].number_format = "#,##0"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column, width in {"A": 48, "B": 12, "C": 88, "D": 20}.items():
        worksheet.column_dimensions[column].width = width
    for row_index in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = 34
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    build_workbook(args.output)
    print(args.output.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
