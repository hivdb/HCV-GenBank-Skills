#!/usr/bin/env python3
"""Replace combined-profile coverage ranges with non-X amino-acid coverage."""

from __future__ import annotations

import argparse
import csv
import re
import shutil
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


LABEL_RE = re.compile(r"^GT(?P<gt>\d+)(?:_(?P<subtype>\S+))? \((?P<count>\d+), [^)]+\)$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--combined-profile-workbook", required=True)
    parser.add_argument("--profile-input-workbook", required=True)
    parser.add_argument("--profile-accessions-csv", required=True)
    parser.add_argument("--output-workbook", help="Optional destination; copy the input workbook before updating it.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    combined_path = Path(args.combined_profile_workbook)
    if args.output_workbook:
        destination = Path(args.output_workbook)
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(combined_path, destination)
        combined_path = destination
    input_path = Path(args.profile_input_workbook)
    with Path(args.profile_accessions_csv).open(newline="", encoding="utf-8-sig") as handle:
        allowed = {row["accession"].strip() for row in csv.DictReader(handle) if row.get("accession", "").strip()}

    combined = load_workbook(combined_path, rich_text=True)
    worksheet = combined.active
    positions = [int(str(cell.value)[1:]) for cell in worksheet[1] if re.fullmatch(r"P\d+", str(cell.value or ""))]

    profile = load_workbook(input_path, read_only=True, data_only=True)
    source = profile.active
    headers = [str(cell.value or "") for cell in next(source.iter_rows(min_row=1, max_row=1))]
    index = {header: position for position, header in enumerate(headers)}
    non_x_calls: dict[tuple[str, str | None], int] = defaultdict(int)
    for row in source.iter_rows(min_row=2, values_only=True):
        accession = str(row[index["AccessionID"]] or "").strip()
        if accession not in allowed:
            continue
        if "AlignmentQCStatus" in index and str(row[index["AlignmentQCStatus"]] or "").strip() != "PASS":
            continue
        genotype = str(row[index["ClosestGT"]] or "").strip().removeprefix("GT")
        subtype = str(row[index["ClosestSubtype"]] or "").strip().lower()
        start = row[index["StartAAPosition"]]
        sequence = str(row[index["AASequence"]] or "").strip().upper()
        if not genotype or not start or not sequence:
            continue
        start = int(start)
        for position in positions:
            offset = position - start
            if not 0 <= offset < len(sequence):
                continue
            amino_acid = sequence[offset]
            if amino_acid != "X":
                non_x_calls[(genotype, None)] += 1
                non_x_calls[(genotype, subtype)] += 1
    profile.close()

    changed = 0
    for row in worksheet.iter_rows(min_row=2, max_col=1):
        cell = row[0]
        match = LABEL_RE.fullmatch(str(cell.value or ""))
        if not match:
            continue
        genotype, subtype, count = match.group("gt"), match.group("subtype"), int(match.group("count"))
        fraction = non_x_calls[(genotype, subtype)] / (len(positions) * count) if count else 0.0
        cell.value = f"GT{genotype}{'_' + subtype if subtype else ''} ({count}, {fraction:.1%})"
        changed += 1
    combined.save(combined_path)
    print(f"Updated {changed} profile labels in {combined_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
