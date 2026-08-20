#!/usr/bin/env python3
"""Create the NS5A genotype workbook directly from Comet assignments."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--fasta-dir", required=True)
    parser.add_argument("--comet-genotype-csv", required=True)
    parser.add_argument("--noncomet-subtype-workbook", required=True)
    parser.add_argument("--output-dir", required=True)
    return parser.parse_args()


def fasta_accessions(path: Path) -> list[str]:
    return [line[1:].strip().split(maxsplit=1)[0] for line in path.read_text(encoding="utf-8").splitlines() if line.startswith(">")]


def load_comet(path: Path) -> dict[str, str]:
    assignments: dict[str, str] = {}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            accession = (row.get("accession") or "").strip()
            genotype = (row.get("genotype") or "").strip()
            if accession and genotype:
                assignments[accession] = genotype
                assignments.setdefault(accession.split(".", 1)[0], genotype)
    return assignments


def load_noncomet_priority_genotypes(path: Path) -> dict[str, tuple[str, str, str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header = [str(value) if value is not None else "" for value in next(sheet.iter_rows(values_only=True))]
    index = {name: position for position, name in enumerate(header)}
    required = ["RefID", "RefName", "AccessionID", "ClosestGT", "ClosestSubtype"]
    missing = [name for name in required if name not in index]
    if missing:
        raise RuntimeError(f"Columns missing from {path}: {', '.join(missing)}")

    assignments: dict[str, tuple[str, str, str, str]] = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        subtype = str(values[index["ClosestSubtype"]]).strip().lower()
        accession = str(values[index["AccessionID"]]).strip()
        genotype = str(values[index["ClosestGT"]]).strip().lower()
        if accession and (subtype == "1d" or genotype in {"7", "8"} or subtype.startswith(("7", "8"))):
            assignments[accession.split(".", 1)[0]] = (
                str(values[index["RefID"]]).strip(),
                str(values[index["RefName"]]).strip(), accession, genotype,
            )
    workbook.close()
    return assignments


def main() -> int:
    args = parse_args()
    assignments = load_comet(Path(args.comet_genotype_csv))
    noncomet_priority_genotypes = load_noncomet_priority_genotypes(Path(args.noncomet_subtype_workbook))
    rows: list[tuple[str, str, str, str, str]] = []
    seen_accessions: set[str] = set()
    override_count = 0
    for fasta_path in sorted(Path(args.fasta_dir).glob("*.fasta")):
        refid, _, refname = fasta_path.stem.partition("_")
        for accession in fasta_accessions(fasta_path):
            accession_key = accession.split(".", 1)[0]
            priority_assignment = noncomet_priority_genotypes.get(accession_key)
            genotype = priority_assignment[3] if priority_assignment else assignments.get(accession) or assignments.get(accession_key)
            if genotype:
                source = "Non-Comet priority genotype override" if priority_assignment else "Comet"
                rows.append((refid, refname, accession, genotype, source))
                seen_accessions.add(accession_key)
                if priority_assignment:
                    override_count += 1

    addition_count = 0
    for accession_key, (refid, refname, accession, genotype) in noncomet_priority_genotypes.items():
        if accession_key not in seen_accessions:
            rows.append((refid, refname, accession, genotype, "Non-Comet priority genotype addition"))
            addition_count += 1

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "NS5A_GT_AllStudies.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "NS5A_GT_AllStudies"
    sheet.append(["RefID", "RefName", "GenBankAccession", "BestGT", "BestGTAssignmentSource"])
    for row in rows:
        sheet.append(row)
    workbook.save(output_path)
    print(json.dumps({"combined_xlsx": str(output_path.resolve()), "master_row_count": len(rows), "noncomet_priority_genotype_override_count": override_count, "noncomet_priority_genotype_addition_count": addition_count}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
