#!/usr/bin/env python3
"""Summarize per-subtype RAS coverage counts across HCV gene profiles."""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


REPO_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_VARIANTS = {
    "one-ras": {
        "#NS3": REPO_ROOT
        / "outputs/comet-NS3-one-ras/23_build-subtype-ras-profile/NS3_Subtype_RAS_Profiles.xlsx",
        "#NS5A": REPO_ROOT
        / "outputs/comet-NS5A-one-ras/23_build-subtype-ras-profile/NS5A_Subtype_RAS_Profiles.xlsx",
        "#NS5B": REPO_ROOT
        / "outputs/comet-NS5B-position-282-include-or-short/23_build-subtype-ras-profile/NS5B_Subtype_RAS_Profiles.xlsx",
    },
}
DEFAULT_OUTPUT_DIR = REPO_ROOT / "outputs/hcv-profile-subtype-accession-summary"
COUNT_COLUMNS = ("#NS3", "#NS5A", "#NS5B")
DISPLAY_THRESHOLD = 10
ALWAYS_INCLUDE_GENOTYPES = frozenset({"7", "8"})
PROFILE_LABEL = re.compile(r"^GT(?P<genotype>\d+)_(?P<subtype>[^ ]+) \((?P<count>\d+),")


def genotype_sort_key(value: str) -> tuple[int, str]:
    match = re.fullmatch(r"\d+", value)
    return (int(value), "") if match else (10**9, value)


def subtype_sort_key(value: str) -> tuple[int, str]:
    match = re.fullmatch(r"(\d+)(.*)", value)
    return (int(match.group(1)), match.group(2)) if match else (10**9, value)


def load_subtype_counts(path: Path) -> dict[tuple[str, str], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    counts: dict[tuple[str, str], int] = {}
    for label, *_ in worksheet.iter_rows(min_row=2, values_only=True):
        match = PROFILE_LABEL.match(str(label or ""))
        if match:
            counts[(match["genotype"], match["subtype"].lower())] = int(match["count"])
    workbook.close()
    if not counts:
        raise ValueError(f"No subtype profile labels found in {path}")
    return counts


def write_csv(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def write_xlsx(
    path: Path,
    title: str,
    headers: list[str],
    rows: list[list[object]],
    count_columns: tuple[int, ...],
    *,
    suppress_repeated_first_column: bool = False,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    previous_first_value: object = None
    for row in rows:
        output_row = list(row)
        if suppress_repeated_first_column and output_row[0] == previous_first_value:
            output_row[0] = ""
        worksheet.append(output_row)
        previous_first_value = row[0]
    for column in count_columns:
        for cells in worksheet.iter_cols(min_col=column, max_col=column, min_row=2):
            for cell in cells:
                if (
                    isinstance(cell.value, (int, float))
                    and cell.value >= DISPLAY_THRESHOLD
                ):
                    cell.font = Font(color="0000FF")
    for column in range(1, len(headers) + 1):
        worksheet.column_dimensions[chr(64 + column)].width = 14
    workbook.save(path)


def build_summary(
    ns3_ras_profile: Path,
    ns5a_ras_profile: Path,
    ns5b_ras_profile: Path,
    output_dir: Path,
    condition: str,
    include_all_subtypes: bool = False,
) -> dict[str, object]:
    grouped_by_gene = {
        "#NS3": load_subtype_counts(ns3_ras_profile),
        "#NS5A": load_subtype_counts(ns5a_ras_profile),
        "#NS5B": load_subtype_counts(ns5b_ras_profile),
    }
    keys = set().union(*(grouped.keys() for grouped in grouped_by_gene.values()))
    rows = []
    for genotype, subtype in sorted(
        keys, key=lambda key: (genotype_sort_key(key[0]), subtype_sort_key(key[1]))
    ):
        counts = [
            grouped_by_gene[column].get((genotype, subtype), 0)
            for column in COUNT_COLUMNS
        ]
        if (
            not include_all_subtypes
            and genotype not in ALWAYS_INCLUDE_GENOTYPES
            and all(count < DISPLAY_THRESHOLD for count in counts)
        ):
            continue
        rows.append([genotype, subtype, *counts])
    rows.append(
        [
            "Total",
            "",
            *[sum(int(row[column]) for row in rows) for column in range(2, 5)],
        ]
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / "HCV_Profile_Subtype_Accession_Counts.csv"
    xlsx_path = output_dir / f"HCV_Profile_Subtype_Accession_Counts_{condition}.xlsx"
    write_csv(csv_path, ["Genotype", "Subtype", *COUNT_COLUMNS], rows)
    write_xlsx(
        xlsx_path,
        "Subtype_Accession_Counts",
        ["Genotype", "Subtype", *COUNT_COLUMNS],
        rows,
        (3, 4, 5),
        suppress_repeated_first_column=True,
    )
    if condition == "one-ras":
        shutil.copy2(xlsx_path, output_dir / "Table1_Gene_Subtype_Counts.xlsx")

    gene_list_headers = [
        "NS3 subtype",
        "#NS3",
        "NS5A subtype",
        "#NS5A",
        "NS5B subtype",
        "#NS5B",
    ]
    gene_lists = [
        [
            (subtype, count)
            for (genotype, subtype), count in sorted(
                grouped_by_gene[column].items(),
                key=lambda item: (
                    genotype_sort_key(item[0][0]),
                    subtype_sort_key(item[0][1]),
                ),
            )
            if include_all_subtypes
            or genotype in ALWAYS_INCLUDE_GENOTYPES
            or count >= DISPLAY_THRESHOLD
        ]
        for column in COUNT_COLUMNS
    ]
    gene_list_rows = []
    for row_index in range(max(len(values) for values in gene_lists)):
        row: list[object] = []
        for values in gene_lists:
            row.extend(values[row_index] if row_index < len(values) else ("", ""))
        gene_list_rows.append(row)
    total_row: list[object] = []
    for values in gene_lists:
        total_row.extend(["Total", sum(count for _, count in values)])
    gene_list_rows.append(total_row)
    gene_list_csv_path = output_dir / "HCV_Profile_Subtype_Counts_By_Gene.csv"
    gene_list_xlsx_path = (
        output_dir / f"HCV_Profile_Subtype_Counts_By_Gene_{condition}.xlsx"
    )
    write_csv(gene_list_csv_path, gene_list_headers, gene_list_rows)
    write_xlsx(
        gene_list_xlsx_path,
        "Subtype_Counts_By_Gene",
        gene_list_headers,
        gene_list_rows,
        (2, 4, 6),
    )

    return {
        "csv": str(csv_path.resolve()),
        "xlsx": str(xlsx_path.resolve()),
        "subtype_group_count": len(rows) - 1,
        "by_gene_csv": str(gene_list_csv_path.resolve()),
        "by_gene_xlsx": str(gene_list_xlsx_path.resolve()),
        "by_gene_row_count": len(gene_list_rows) - 1,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--ns3-ras-profile",
        type=Path,
        help="Generate one custom summary with this NS3 RAS profile.",
    )
    parser.add_argument(
        "--ns5a-ras-profile",
        type=Path,
        help="Generate one custom summary with this NS5A RAS profile.",
    )
    parser.add_argument(
        "--ns5b-ras-profile",
        type=Path,
        help="Generate one summary with this NS5B RAS profile instead of the default One RAS inputs.",
    )
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()

    custom_profiles = (
        args.ns3_ras_profile,
        args.ns5a_ras_profile,
        args.ns5b_ras_profile,
    )
    if any(custom_profiles):
        if not all(custom_profiles):
            parser.error(
                "--ns3-ras-profile, --ns5a-ras-profile, and --ns5b-ras-profile must be provided together."
            )
        summaries = {
            "custom": build_summary(
                args.ns3_ras_profile,
                args.ns5a_ras_profile,
                args.ns5b_ras_profile,
                args.output_dir,
                "custom",
            )
        }
    else:
        summaries = {}
        for variant, profiles in DEFAULT_VARIANTS.items():
            summaries[variant] = build_summary(
                profiles["#NS3"],
                profiles["#NS5A"],
                profiles["#NS5B"],
                args.output_dir / variant,
                variant,
                include_all_subtypes=True,
            )
    print(json.dumps({"summaries": summaries}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
