#!/usr/bin/env python3
"""Combine NS5B genotype and subtype RAS profile workbooks by genotype."""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
import xlsxwriter


VARIANT_RE = re.compile(r"([A-Z*])(\d+(?:\.\d+)?)")
TOTAL_SEQUENCE_RE = re.compile(r"\(\s*(\d+)\s*,")
MIN_TOTAL_SEQUENCES = 10
FULL_PROFILE_SUMMARY_MIN_TOTAL_SEQUENCES = 10


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Combine NS5B GT and subtype RAS profiles, retaining subtype amino-acid "
            "variants strictly above a frequency threshold."
        )
    )
    parser.add_argument("--gt-ras-profile-workbook", required=True)
    parser.add_argument("--subtype-ras-profile-workbook", required=True)
    parser.add_argument("--output-xlsx", default="outputs/NS5B_Combined_RAS_Profiles.xlsx")
    parser.add_argument("--subtype-frequency-threshold", type=float, default=1.0)
    return parser.parse_args()


def read_profile_workbook(path: Path) -> tuple[list[object], list[list[object]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        raise RuntimeError(f"Profile workbook is empty: {path}")
    return list(rows[0]), [list(row) for row in rows[1:] if row and row[0]]


def genotype_from_label(label: object) -> str | None:
    match = re.match(r"GT(\d+)(?:\s|_)", str(label))
    return match.group(1) if match else None


def has_minimum_total_sequences(label: object) -> bool:
    match = TOTAL_SEQUENCE_RE.search(str(label))
    return match is not None and int(match.group(1)) >= MIN_TOTAL_SEQUENCES


def total_sequences(label: object) -> int | None:
    match = TOTAL_SEQUENCE_RE.search(str(label))
    return int(match.group(1)) if match is not None else None


def filtered_variants(
    value: object, threshold: float | None = None, excluded_amino_acids: set[str] | None = None
) -> list[tuple[str, str]]:
    return [
        (amino_acid, frequency)
        for amino_acid, frequency in VARIANT_RE.findall(str(value or ""))
        if (threshold is None or float(frequency) >= threshold)
        and amino_acid not in (excluded_amino_acids or set())
    ]


def most_frequent_variant(value: object) -> tuple[str, str] | None:
    variants = VARIANT_RE.findall(str(value or ""))
    if not variants:
        return None
    return max(variants, key=lambda variant: float(variant[1]))


def mean_diff(value_rows: list[object], gt_variants: list[tuple[str, str] | None], threshold: float) -> float:
    """Sum displayed non-consensus amino-acid percentages and express as a decimal."""
    displayed_percent = sum(
        float(frequency)
        for value, gt_variant in zip(value_rows, gt_variants)
        for amino_acid, frequency in VARIANT_RE.findall(str(value or ""))
        if float(frequency) >= threshold and (gt_variant is None or amino_acid != gt_variant[0])
    )
    return displayed_percent / 100.0


def write_combined_workbook(
    output_path: Path,
    positions: list[object],
    gt_rows: list[list[object]],
    subtype_rows: list[list[object]],
    threshold: float,
) -> tuple[int, int]:
    gt_by_number = {
        genotype: row
        for row in gt_rows
        if (genotype := genotype_from_label(row[0])) is not None
        and has_minimum_total_sequences(row[0])
    }
    subtypes_by_gt: dict[str, list[list[object]]] = defaultdict(list)
    for row in subtype_rows:
        genotype = genotype_from_label(row[0])
        if genotype is not None and has_minimum_total_sequences(row[0]):
            subtypes_by_gt[genotype].append(row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = xlsxwriter.Workbook(output_path)
    worksheet = workbook.add_worksheet("Combined_RAS_Profile")
    cell_format = workbook.add_format({"align": "center", "valign": "vcenter", "text_wrap": True})
    header_format = workbook.add_format({"bold": True, "bg_color": "#F2F2F2", "align": "center", "valign": "vcenter", "text_wrap": True})
    gt_format = workbook.add_format({"bold": True, "bg_color": "#D9EAF7", "align": "center", "valign": "vcenter", "text_wrap": True})
    mean_diff_format = workbook.add_format({"align": "center", "valign": "vcenter", "text_wrap": True, "num_format": "0.0"})
    superscript_format = workbook.add_format({"font_script": 1})
    red_text_format = workbook.add_format({"font_color": "#FF0000"})
    red_superscript_format = workbook.add_format({"font_color": "#FF0000", "font_script": 1})
    headers = [*positions, "MeanDiff"]
    worksheet.set_column(0, 0, 24)
    worksheet.set_column(1, len(headers) - 1, 12)
    for column, value in enumerate(headers):
        worksheet.write_blank(0, column, None, header_format) if value is None else worksheet.write(0, column, value, header_format)

    def write_variant_cell(
        row: int, column: int, variants: list[tuple[str, str]], cell_style, genotype_consensus_aas: set[str] | None = None
    ) -> None:
        if not variants:
            worksheet.write_blank(row, column, None, cell_style)
            return
        rich_parts: list[object] = []
        for index, (amino_acid, frequency) in enumerate(variants):
            if index and index % 2 == 0:
                amino_acid = f"\n{amino_acid}"
            if genotype_consensus_aas is not None and amino_acid.lstrip("\n") not in genotype_consensus_aas:
                rich_parts.extend((red_text_format, amino_acid, red_superscript_format, frequency))
            else:
                rich_parts.extend((amino_acid, superscript_format, frequency))
        result = worksheet.write_rich_string(row, column, *rich_parts, cell_style)
        if result:
            raise RuntimeError(f"Unable to write rich text at row {row + 1}, column {column + 1}: {result}")

    output_rows = 0
    worksheet_row = 1
    for genotype in sorted(gt_by_number, key=int):
        gt_row = gt_by_number[genotype]
        worksheet.write(worksheet_row, 0, gt_row[0], gt_format)
        for column, value in enumerate(gt_row[1:], start=1):
            variant = most_frequent_variant(value)
            write_variant_cell(worksheet_row, column, [variant] if variant else [], gt_format)
        worksheet.write_blank(worksheet_row, len(headers) - 1, None, gt_format)
        output_rows += 1
        worksheet_row += 1

        gt_amino_acids = [most_frequent_variant(value) for value in gt_row[1:]]
        all_gt_consensus_aas = [
            {
                variant[0]
                for gt_number, row in gt_by_number.items()
                if gt_number in {"1", "2", "3", "4", "5", "6"}
                if (variant := most_frequent_variant(row[position_index + 1])) is not None
            }
            for position_index in range(len(gt_amino_acids))
        ]
        for subtype_row in subtypes_by_gt.get(genotype, []):
            subtype_values = subtype_row[1:]
            worksheet.write(worksheet_row, 0, subtype_row[0], cell_format)
            for column, (value, gt_variant) in enumerate(zip(subtype_values, gt_amino_acids), start=1):
                write_variant_cell(worksheet_row, column, filtered_variants(value, threshold, {gt_variant[0]} if gt_variant is not None else None), cell_format, all_gt_consensus_aas[column - 1])
            worksheet.write_number(worksheet_row, len(headers) - 1, mean_diff(subtype_values, gt_amino_acids, threshold), mean_diff_format)
            output_rows += 1
            worksheet_row += 1
        worksheet_row += 1

    workbook.close()
    return len(gt_by_number), output_rows


def main() -> int:
    args = parse_args()
    gt_path = Path(args.gt_ras_profile_workbook).expanduser()
    subtype_path = Path(args.subtype_ras_profile_workbook).expanduser()
    output_path = Path(args.output_xlsx).expanduser()
    if not gt_path.is_file() or not subtype_path.is_file():
        raise RuntimeError("Both GT and subtype RAS profile workbooks are required")

    positions, gt_rows = read_profile_workbook(gt_path)
    subtype_positions, subtype_rows = read_profile_workbook(subtype_path)
    if positions != subtype_positions:
        raise RuntimeError("GT and subtype RAS profile workbooks have different position rows")
    genotype_count, output_rows = write_combined_workbook(
        output_path, positions, gt_rows, subtype_rows, args.subtype_frequency_threshold
    )
    full_profile_subtype_count = len(subtype_rows)
    full_profile_subtype_at_least_10_count = sum(
        (count := total_sequences(row[0])) is not None
        and count >= FULL_PROFILE_SUMMARY_MIN_TOTAL_SEQUENCES
        for row in subtype_rows
    )
    print(
        json.dumps(
            {
                "output_xlsx": str(output_path.resolve()),
                "genotype_count": genotype_count,
                "profile_row_count": output_rows,
                "full_profile_subtype_count": full_profile_subtype_count,
                "full_profile_subtype_at_least_10_sequence_count": full_profile_subtype_at_least_10_count,
                "subtype_frequency_threshold": args.subtype_frequency_threshold,
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
