#!/usr/bin/env python3
"""Report profile-used accessions whose subtype is prioritized from non-Comet."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import load_workbook


def accession_key(value: object) -> str:
    return str(value or "").strip().split(".", 1)[0]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--profile-accessions-csv", required=True)
    parser.add_argument("--comet-subtype-csv", required=True)
    parser.add_argument("--noncomet-subtype-workbook", required=True)
    parser.add_argument("--output-csv", required=True)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    with Path(args.profile_accessions_csv).open(encoding="utf-8", newline="") as handle:
        profile_accessions = {accession_key(row["accession"]): row["accession"].strip() for row in csv.DictReader(handle) if accession_key(row.get("accession"))}
    with Path(args.comet_subtype_csv).open(encoding="utf-8", newline="") as handle:
        comet_subtypes = {accession_key(row.get("name") or row.get("accession")): str(row.get("subtype") or "").strip().lower() for row in csv.DictReader(handle)}

    workbook = load_workbook(args.noncomet_subtype_workbook, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header = [str(value or "") for value in next(sheet.iter_rows(values_only=True))]
    index = {name: position for position, name in enumerate(header)}
    required = {"AccessionID", "ClosestGT", "ClosestSubtype"}
    if missing := required - index.keys():
        raise RuntimeError(f"Columns missing from {args.noncomet_subtype_workbook}: {', '.join(sorted(missing))}")
    priority_subtypes: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        accession = accession_key(row[index["AccessionID"]])
        genotype = str(row[index["ClosestGT"]] or "").strip().lower()
        subtype = str(row[index["ClosestSubtype"]] or "").strip().lower()
        if accession and (subtype == "1d" or genotype in {"7", "8"} or subtype.startswith(("7", "8"))):
            priority_subtypes[accession] = subtype
    workbook.close()

    rows = [(profile_accessions[key], comet_subtypes.get(key, ""), subtype) for key, subtype in priority_subtypes.items() if key in profile_accessions]
    rows.sort(key=lambda row: row[0])
    output = Path(args.output_csv)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["AccessionID", "CometSubtype", "NonCometSubtype"])
        writer.writerows(rows)
    print(json.dumps({"output_csv": str(output.resolve()), "profile_used_noncomet_priority_accession_count": len(rows)}))


if __name__ == "__main__":
    main()
