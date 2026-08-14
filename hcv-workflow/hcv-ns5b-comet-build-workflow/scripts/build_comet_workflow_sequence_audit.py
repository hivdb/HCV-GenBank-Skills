#!/usr/bin/env python3
"""Create a sequence-count and filtering audit from Ref-selection to COMET profiles."""
from __future__ import annotations
import argparse, csv, re
from collections import Counter
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

QUASISPECIES={'19','31','32','34','70','81','115','262','1044','2043','2071','2129','2139','2175','2195','2212','2216','2225','2324'}
SUBTYPE=re.compile(r'^GT(\d+)_(\S+) \((\d+),')
VALID_CALL=re.compile(r'^[1-8][a-z][a-z0-9]*$',re.I)

def args():
 p=argparse.ArgumentParser(description=__doc__)
 for n in ('gene','selection_workbook','selection_sheet','fasta_dir','metadata_csv','comet_csv','qc_workbook','profile_accessions_csv','combined_profile_workbook','output_xlsx'):p.add_argument('--'+n.replace('_','-'),required=True)
 p.add_argument('--combined-subtype-cutoff',type=int,default=10,help='Subtype sequence-count cutoff used by the combined-profile builder.')
 p.add_argument('--combined-subtype-cutoff-exclusive',action='store_true',help='Exclude subtypes whose count equals the cutoff (for example, retain >5).')
 p.add_argument('--profile-filter-reason',default='QC-passed accession missing from profile builder output',help='Reason to report when the profile builder omits QC-passed accessions.')
 return p.parse_args()
def headers(ws):
 h=[str(c.value or '') for c in next(ws.iter_rows(min_row=1,max_row=1))];return h,{v:i for i,v in enumerate(h)}
def fasta(path):
 return [l[1:].split()[0].split('.')[0] for l in path.read_text(errors='replace').splitlines() if l.startswith('>')]
def main():
 a=args(); sel=load_workbook(a.selection_workbook,read_only=True,data_only=True);ws=sel[a.selection_sheet];h,ix=headers(ws); rows=list(ws.iter_rows(min_row=2,values_only=True));sel.close()
 selected=[]; check_refids=[]; reasons=Counter()
 for r in rows:
  rid=str(r[ix['RefID']] or '').strip(); pts=str(r[ix['Num Pts']] or '').strip()
  if not rid: reasons['Missing RefID']+=1
  elif pts.casefold()=='exclude': reasons['Num Pts = Exclude']+=1
  elif rid in QUASISPECIES: reasons['Known quasispecies RefID']+=1
  elif pts.casefold()=='check': check_refids.append(rid)
  else:selected.append(rid)
 refids=set(selected); check_refids=set(check_refids)-refids; raw=[]; check_raw=[]
 for f in Path(a.fasta_dir).glob('*.fasta'):
  if f.name.split('_',1)[0] in refids:raw+=fasta(f)
  elif f.name.split('_',1)[0] in check_refids:check_raw+=fasta(f)
 raw=set(raw); check_raw=set(check_raw)-raw; raw_before_check=raw|check_raw
 with open(a.metadata_csv,newline='',encoding='utf-8-sig') as f:meta={r['Accession'].strip() for r in csv.DictReader(f) if r.get('Accession','').strip()}
 meta_kept=raw & meta
 with open(a.comet_csv,newline='',encoding='utf-8-sig') as f:
  calls={r['name'].strip().split('.')[0]:r.get('subtype','').strip() for r in csv.DictReader(f) if r.get('name','').strip() and r.get('virus','').strip().upper()=='HCV'}
 assigned={x for x in meta_kept if VALID_CALL.fullmatch(calls.get(x,''))}; missing=meta_kept-set(calls); unassigned=(meta_kept&set(calls))-assigned
 qwb=load_workbook(a.qc_workbook,read_only=True,data_only=True);qws=qwb.active;qh,qix=headers(qws); qrows=list(qws.iter_rows(min_row=2,values_only=True));qwb.close()
 qacc={str(r[qix['AccessionID']] or '').strip() for r in qrows}; q_assigned=assigned&qacc; qcpass={str(r[qix['AccessionID']] or '').strip() for r in qrows if str(r[qix.get('AlignmentQCStatus','')] or '').strip()=='PASS'}
 qreasons=Counter()
 for r in qrows:
  ac=str(r[qix['AccessionID']] or '').strip()
  if ac in assigned and str(r[qix.get('AlignmentQCStatus','')] or '').strip()!='PASS': qreasons[str(r[qix.get('AlignmentQCReasons','')] or 'QC failed')]+=1
 with open(a.profile_accessions_csv,newline='',encoding='utf-8-sig') as f:profile={r['accession'].strip() for r in csv.DictReader(f) if r.get('accession','').strip()}
 with open(a.profile_accessions_csv,newline='',encoding='utf-8-sig') as f:
  profile_groups=Counter((r['genotype'].strip(),r['subtype'].strip().lower()) for r in csv.DictReader(f) if r.get('accession','').strip())
 cutoff_excluded={group:count for group,count in profile_groups.items() if group[0] not in {'7','8'} and (count<=a.combined_subtype_cutoff if a.combined_subtype_cutoff_exclusive else count<a.combined_subtype_cutoff)}
 cutoff_excluded_count=sum(cutoff_excluded.values())
 cutoff_retained_count=len(profile)-cutoff_excluded_count
 cutoff_rule=(f'{a.combined_subtype_cutoff} or fewer' if a.combined_subtype_cutoff_exclusive else f'fewer than {a.combined_subtype_cutoff}')
 cw=load_workbook(a.combined_profile_workbook,read_only=True,data_only=True);cws=cw.active; labels=[str(r[0] or '') for r in cws.iter_rows(min_row=2,max_col=1,values_only=True)];cw.close(); combined=[SUBTYPE.match(x) for x in labels]; combined=[m for m in combined if m]
 out=Workbook();summary=out.active;summary.title='Sequence_Audit';summary.append(['Step','Sequences retained','Sequences filtered out','Filtering applied'])
 profile_missing=(qcpass&assigned)-profile
 priority_profile=profile-(qcpass&assigned)
 entries=[('Ref-selection rows',len(rows),0,'Input rows in selected worksheet'),('Ref-selection eligibility',len(selected)+len(check_refids),len(rows)-len(selected)-len(check_refids),'Exclude Num Pts = Exclude; known quasispecies RefIDs'),('Raw FASTA accessions',len(raw_before_check),0,'FASTA files matching RefIDs eligible before the Check exclusion'),('RefID Check exclusion',len(raw),len(check_raw),'RefID marked Check'),('Metadata match',len(meta_kept),len(raw-meta),'Accession absent from Accessions_metadata.csv'),('COMET CSV match',len(meta_kept)-len(missing),len(missing),'Accession absent from COMET CSV'),('COMET subtype validation',len(assigned),len(unassigned),'COMET unassigned/invalid subtype'),('AA extraction',len(q_assigned),len(assigned-qacc),'No extracted AA row'),('Alignment QC pass',len(qcpass&assigned),len(q_assigned-qcpass),'AlignmentQCStatus is not PASS'),('Non-COMET priority subtype additions',len(qcpass&assigned)+len(priority_profile),0,f'{len(priority_profile)} QC-passed accessions retained by the non-COMET priority subtype rule'),('Profile accessions',len(profile),len(profile_missing),'All eligible QC-passed accessions included; none excluded by profile builder' if not profile_missing else a.profile_filter_reason),('Combined-profile subtype cutoff',cutoff_retained_count,cutoff_excluded_count,f'Exclude subtypes with {cutoff_rule} profile accessions; genotypes 7 and 8 are retained regardless ({len(cutoff_excluded)} subtype groups excluded)'),('Combined profile',sum(int(m.group(3)) for m in combined),0,f'{len(combined)} subtype rows; count is summed RAS-position coverage, not unique accessions')]
 for x in entries:summary.append(x)
 details=out.create_sheet('Filtered_Out_Reasons');details.append(['Stage','Reason','SequenceOrRowCount'])
 for k,v in sorted(reasons.items()):details.append(['Ref-selection',k,v])
 details.append(['RefID Check exclusion','RefID marked Check',len(check_raw)])
 details.append(['Metadata','Accession absent from Accessions_metadata.csv',len(raw-meta)])
 details.append(['COMET','Accession absent from COMET CSV',len(missing)]);details.append(['COMET','COMET unassigned/invalid subtype',len(unassigned)])
 details.append(['AA extraction','No extracted AA row',len(assigned-qacc)])
 for k,v in sorted(qreasons.items()):details.append(['Alignment QC',k,v])
 if profile_missing:details.append(['Profile accessions',a.profile_filter_reason,len(profile_missing)])
 details.append(['Combined profile',f'Subtype has {cutoff_rule} profile accessions',cutoff_excluded_count])
 for s in (summary,details):
  for c in s[1]:c.font=Font(bold=True);c.fill=PatternFill(fill_type='solid',fgColor='D9EAF7')
  s.freeze_panes='A2'
  for col in s.columns:s.column_dimensions[col[0].column_letter].width=min(max(max(len(str(c.value or '')) for c in col)+2,16),100)
 Path(a.output_xlsx).parent.mkdir(parents=True,exist_ok=True);out.save(a.output_xlsx);print(a.output_xlsx)
if __name__=='__main__':main()
