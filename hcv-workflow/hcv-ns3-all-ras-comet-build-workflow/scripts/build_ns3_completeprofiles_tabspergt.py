#!/usr/bin/env python3

from __future__ import annotations

import argparse
import os
import csv
import json
import shutil
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


AA_ORDER = list("ACDEFGHIKLMNPQRSTVWY") + ["*"]
NS3_RAS_POSITIONS = (36, 41, 43, 54, 55, 56, 80, 122, 155, 156, 158, 166, 168, 170, 175)
CALLABLE_AAS = frozenset(AA_ORDER) - {"*"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build GT and subtype NS3 amino-acid profile workbooks.")
    parser.add_argument("--input-workbook", required=True, help="Path to NS3 AA extraction workbook.")
    parser.add_argument("--output-dir", default="outputs", help="Base output directory.")
    parser.add_argument("--profile-accessions-csv", help="CSV listing accessions included in profile construction.")
    parser.add_argument(
        "--report-only",
        action="store_true",
        help="Print profile-input accession counts without creating workbooks.",
    )
    return parser.parse_args()


def script_temp_dir() -> Path:
    path = Path(os.environ.get("NS3_STEP_OUTPUT_DIR", "outputs/comet-NS3-all-ras/temp")) / Path(__file__).stem
    path.mkdir(parents=True, exist_ok=True)
    return path


def sanitize_label(value: str) -> str:
    return "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in value).strip("._-") or "job"


def make_job_dir(base_output_dir: Path, workbook_path: Path) -> Path:
    label = sanitize_label(f"{workbook_path.stem}_ns3_aa_profiles")
    job_dir = base_output_dir / label
    if job_dir.exists():
        shutil.rmtree(job_dir)
    job_dir.mkdir(parents=True, exist_ok=True)
    return job_dir


def has_callable_ras_positions(start: int, aa_sequence: str) -> bool:
    """Return whether every NS3 RAS position has a callable standard amino-acid call."""
    sequence = aa_sequence.upper()
    return all(
        0 <= position - start < len(sequence) and sequence[position - start] in CALLABLE_AAS
        for position in NS3_RAS_POSITIONS
    )


def load_rows(workbook_path: Path) -> tuple[list[dict[str, Any]], dict[str, int]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(v) if v is not None else "" for v in next(ws.iter_rows(values_only=True))]
    index = {name: i for i, name in enumerate(header)}
    required = ["AccessionID", "ClosestGT", "ClosestSubtype", "StartAAPosition", "EndAAPosition", "AASequence"]
    for name in required:
        if name not in index:
            raise RuntimeError(f"Column '{name}' not found in {workbook_path}")
    rows: list[dict[str, Any]] = []
    unassigned_genotype_accessions: set[str] = set()
    unassigned_subtype_accessions: set[str] = set()
    qc_failed_accessions: set[str] = set()
    qc_failed_by_status: dict[str, set[str]] = defaultdict(set)
    incomplete_ras_accessions: set[str] = set()
    for values in ws.iter_rows(min_row=2, values_only=True):
        accession = str(values[index["AccessionID"]]).strip()
        genotype = str(values[index["ClosestGT"]]).strip()
        subtype = str(values[index["ClosestSubtype"]]).strip()
        # When the pre-profile QC workbook is supplied, coordinate failures must
        # not contribute shifted amino-acid calls to any profile.  Older input
        # workbooks without QC columns retain their existing behavior.
        if "AlignmentQCStatus" in index and str(values[index["AlignmentQCStatus"]] or "").strip() != "PASS":
            status = str(values[index["AlignmentQCStatus"]] or "").strip() or "MISSING_QC_STATUS"
            if accession:
                qc_failed_accessions.add(accession)
                qc_failed_by_status[status].add(accession)
            continue
        aa_sequence = values[index["AASequence"]]
        start = values[index["StartAAPosition"]]
        end = values[index["EndAAPosition"]]
        if not aa_sequence or start in (None, "") or end in (None, ""):
            continue
        if genotype.casefold().startswith("unassign"):
            unassigned_genotype_accessions.add(accession)
        if subtype.casefold().startswith("unassign"):
            unassigned_subtype_accessions.add(accession)
        if genotype.casefold().startswith("unassign") or subtype.casefold().startswith("unassign"):
            continue
        sequence = str(aa_sequence).strip()
        if not has_callable_ras_positions(int(start), sequence):
            incomplete_ras_accessions.add(accession)
            continue
        rows.append(
            {
                "AccessionID": accession,
                "ClosestGT": genotype,
                "ClosestSubtype": subtype,
                "StartAAPosition": int(start),
                "EndAAPosition": int(end),
                "AASequence": sequence,
            }
        )
    wb.close()
    counts = {
        "ignored_unassigned_genotype_accession_count": len(unassigned_genotype_accessions),
        "ignored_unassigned_subtype_accession_count": len(unassigned_subtype_accessions),
        "ignored_unassigned_accession_count": len(unassigned_genotype_accessions | unassigned_subtype_accessions),
        "ignored_alignment_qc_accession_count": len(qc_failed_accessions),
        "ignored_incomplete_ras_coverage_accession_count": len(incomplete_ras_accessions),
    }
    for status, accessions in sorted(qc_failed_by_status.items()):
        counts[f"excluded_qc_{status.casefold()}_accession_count"] = len(accessions)
    return rows, counts


def build_position_counts(rows: list[dict[str, Any]]) -> tuple[dict[int, int], dict[int, Counter[str]]]:
    included_counts: dict[int, int] = defaultdict(int)
    aa_counts: dict[int, Counter[str]] = defaultdict(Counter)
    for row in rows:
        start = row["StartAAPosition"]
        aa_sequence = row["AASequence"]
        for offset, aa in enumerate(aa_sequence):
            aa = aa.upper()
            if aa == "X":
                continue
            pos = start + offset
            included_counts[pos] += 1
            aa_counts[pos][aa] += 1
    return included_counts, aa_counts


def accession_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    included = {row["AccessionID"] for row in rows if row["AccessionID"]}
    with_subtype = {
        row["AccessionID"]
        for row in rows
        if row["AccessionID"] and row["ClosestSubtype"]
    }
    return {
        "included_accession_count": len(included),
        "accessions_with_comet_subtype_count": len(with_subtype),
        "accessions_without_comet_subtype_count": len(included - with_subtype),
    }


def write_profile_accessions(path: Path, rows: list[dict[str, Any]]) -> None:
    accessions: dict[str, tuple[str, str]] = {}
    for row in rows:
        accession = row["AccessionID"]
        if accession:
            accessions[accession] = (row["ClosestGT"], row["ClosestSubtype"])
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["accession", "genotype", "subtype"])
        for accession in sorted(accessions):
            genotype, subtype = accessions[accession]
            writer.writerow([accession, genotype, subtype])


def write_gt_workbook(path: Path, rows_by_gt: dict[str, list[dict[str, Any]]]) -> dict[str, int]:
    wb = Workbook()
    wb.remove(wb.active)
    summary: dict[str, int] = {}
    header = [
        "NS3Position",
        "NumSeqsIncludingPosition",
        "AminoAcid",
        "CountWithAA",
        "CountWithAAAlone",
        "PctWithAA",
        "PctWithAAAlone",
    ]
    for gt in sorted(rows_by_gt, key=int):
        ws = wb.create_sheet(f"GT{gt}")
        ws.append(header)
        included_counts, aa_counts = build_position_counts(rows_by_gt[gt])
        summary[gt] = len(rows_by_gt[gt])
        for pos in sorted(included_counts):
            denom = included_counts[pos]
            for aa in AA_ORDER:
                count = aa_counts[pos].get(aa, 0)
                if count == 0:
                    continue
                ws.append([pos, denom, aa, count, count, 100.0 * count / denom, 100.0 * count / denom])
    wb.save(path)
    return summary


def write_subtype_workbook(path: Path, rows_by_gt_subtype: dict[str, dict[str, list[dict[str, Any]]]]) -> dict[str, dict[str, int]]:
    wb = Workbook()
    wb.remove(wb.active)
    summary: dict[str, dict[str, int]] = {}
    header = [
        "Subtype",
        "NS3Position",
        "NumSeqsIncludingPosition",
        "AminoAcid",
        "CountWithAA",
        "CountWithAAAlone",
        "PctWithAA",
        "PctWithAAAlone",
    ]
    for gt in sorted(rows_by_gt_subtype, key=int):
        ws = wb.create_sheet(f"GT{gt}")
        ws.append(header)
        summary[gt] = {}
        for subtype in sorted(rows_by_gt_subtype[gt]):
            subtype_rows = rows_by_gt_subtype[gt][subtype]
            included_counts, aa_counts = build_position_counts(subtype_rows)
            summary[gt][subtype] = len(subtype_rows)
            for pos in sorted(included_counts):
                denom = included_counts[pos]
                for aa in AA_ORDER:
                    count = aa_counts[pos].get(aa, 0)
                    if count == 0:
                        continue
                    pct = 100.0 * count / denom
                    ws.append([subtype, pos, denom, aa, count, count, pct, pct])
    wb.save(path)
    return summary


def main() -> int:
    args = parse_args()
    input_workbook = Path(args.input_workbook).expanduser()
    output_dir = Path(args.output_dir)
    script_temp_dir()

    rows, ignored_counts = load_rows(input_workbook)
    genotype_distribution = Counter(str(row["ClosestGT"]).removeprefix("GT") for row in rows)
    counts = {
        **accession_counts(rows),
        **ignored_counts,
        "included_genotype_distribution": dict(sorted(genotype_distribution.items(), key=lambda item: (0, int(item[0])) if item[0].isdigit() else (1, item[0]))),
    }
    if args.report_only:
        print(json.dumps(counts))
        return 0
    if args.profile_accessions_csv:
        write_profile_accessions(Path(args.profile_accessions_csv), rows)

    rows_by_gt: dict[str, list[dict[str, Any]]] = defaultdict(list)
    rows_by_gt_subtype: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    for row in rows:
        gt = row["ClosestGT"]
        subtype = row["ClosestSubtype"]
        rows_by_gt[gt].append(row)
        rows_by_gt_subtype[gt][subtype].append(row)

    output_dir.mkdir(parents=True, exist_ok=True)
    gt_path = output_dir / "NS3_GT_CompleteProfiles_TabsPerGT.xlsx"
    subtype_path = output_dir / "NS3_Subtype_CompleteProfiles_TabsPerGT.xlsx"

    gt_summary = write_gt_workbook(gt_path, rows_by_gt)
    subtype_summary = write_subtype_workbook(subtype_path, rows_by_gt_subtype)

    summary = {
        "input_workbook": str(input_workbook.resolve()),
        "rows_with_aa": len(rows),
        **counts,
        "gt_workbook": str(gt_path.resolve()),
        "subtype_workbook": str(subtype_path.resolve()),
        "gt_sequence_counts": gt_summary,
        "subtype_group_count": sum(len(v) for v in rows_by_gt_subtype.values()),
        "profile_required_ras_positions": list(NS3_RAS_POSITIONS),
        "profile_input_rule": "Every retained accession has a callable amino-acid at every NS3 RAS position.",
        "note": "CountWithAA and CountWithAAAlone are identical because current AA sequences contain single-letter calls, not explicit mixtures.",
    }
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
