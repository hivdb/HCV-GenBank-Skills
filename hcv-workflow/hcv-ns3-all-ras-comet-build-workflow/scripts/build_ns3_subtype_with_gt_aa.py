#!/usr/bin/env python3

from __future__ import annotations

import argparse
import os
import json
import re
import shutil
import subprocess
import tempfile
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from Bio import Align
from openpyxl import Workbook, load_workbook


BLAST_OUTFMT = "6 qseqid sseqid length mismatch gaps pident evalue bitscore qstart qend sstart send qframe"
FASTA_EXTENSIONS = {".fa", ".fasta", ".fna"}
CODON_TABLE = {
    "TTT": "F", "TTC": "F", "TTA": "L", "TTG": "L",
    "TCT": "S", "TCC": "S", "TCA": "S", "TCG": "S",
    "TAT": "Y", "TAC": "Y", "TAA": "*", "TAG": "*",
    "TGT": "C", "TGC": "C", "TGA": "*", "TGG": "W",
    "CTT": "L", "CTC": "L", "CTA": "L", "CTG": "L",
    "CCT": "P", "CCC": "P", "CCA": "P", "CCG": "P",
    "CAT": "H", "CAC": "H", "CAA": "Q", "CAG": "Q",
    "CGT": "R", "CGC": "R", "CGA": "R", "CGG": "R",
    "ATT": "I", "ATC": "I", "ATA": "I", "ATG": "M",
    "ACT": "T", "ACC": "T", "ACA": "T", "ACG": "T",
    "AAT": "N", "AAC": "N", "AAA": "K", "AAG": "K",
    "AGT": "S", "AGC": "S", "AGA": "R", "AGG": "R",
    "GTT": "V", "GTC": "V", "GTA": "V", "GTG": "V",
    "GCT": "A", "GCC": "A", "GCA": "A", "GCG": "A",
    "GAT": "D", "GAC": "D", "GAA": "E", "GAG": "E",
    "GGT": "G", "GGC": "G", "GGA": "G", "GGG": "G",
}
COMPLEMENT = str.maketrans("ACGTRYSWKMBDHVNacgtryswkmbdhvn", "TGCAYRSWMKVHDBNtgcayrswmkvhdbn")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Add GT-guided NS3 amino-acid extraction columns to the subtype workbook."
    )
    parser.add_argument("--subtype-workbook", required=True)
    parser.add_argument("--fasta-dir", required=True)
    parser.add_argument("--gt-aa-json", required=True)
    parser.add_argument("--output-dir", default="outputs")
    parser.add_argument("--output-workbook")
    parser.add_argument("--min-aa-overlap", type=int, default=80)
    return parser.parse_args()


def sanitize_label(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return text.strip("._-") or "job"


def script_temp_dir() -> Path:
    path = Path(os.environ.get("NS3_STEP_OUTPUT_DIR", "outputs/comet-NS3-all-ras/temp")) / Path(__file__).stem
    path.mkdir(parents=True, exist_ok=True)
    return path


def make_job_dir(base_output_dir: Path, workbook_path: Path) -> Path:
    label = sanitize_label(f"{workbook_path.stem}_ns3_gt_aa_extraction")
    return Path(tempfile.mkdtemp(prefix=f"{label}_", dir=script_temp_dir()))


def make_temp_output_path(filename: str) -> Path:
    handle = tempfile.NamedTemporaryFile(prefix="tmp_", suffix=f"_{filename}", dir=script_temp_dir(), delete=False)
    handle.close()
    return Path(handle.name)


def parse_fasta(path: Path) -> list[tuple[str, str]]:
    records: list[tuple[str, str]] = []
    header: str | None = None
    chunks: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line:
            continue
        if line.startswith(">"):
            if header is not None:
                records.append((header, "".join(chunks).upper()))
            header = line[1:].strip()
            chunks = []
        else:
            chunks.append(re.sub(r"\s+", "", line))
    if header is not None:
        records.append((header, "".join(chunks).upper()))
    return records


def accession_from_header(header: str) -> str:
    return header.split()[0]


def normalize_nt(sequence: str) -> str:
    return re.sub(r"[^ACGTRYSWKMBDHVN]", "N", sequence.upper())


def reverse_complement(sequence: str) -> str:
    return sequence.translate(COMPLEMENT)[::-1]


def translate_nt(sequence: str) -> str:
    aa: list[str] = []
    for i in range(0, len(sequence) - 2, 3):
        codon = sequence[i : i + 3]
        if any(base not in "ACGT" for base in codon):
            aa.append("X")
        else:
            aa.append(CODON_TABLE.get(codon, "X"))
    return "".join(aa)


def build_aligner() -> Align.PairwiseAligner:
    aligner = Align.PairwiseAligner(mode="global")
    aligner.match_score = 2.0
    aligner.mismatch_score = -1.0
    aligner.open_gap_score = -5.0
    aligner.extend_gap_score = -1.0
    return aligner


def alignment_strings(reference: str, query: str, coordinates: Any) -> tuple[str, str]:
    ref_chunks: list[str] = []
    query_chunks: list[str] = []
    ref_points = coordinates[0]
    query_points = coordinates[1]
    for idx in range(len(ref_points) - 1):
        ref_start = int(ref_points[idx])
        ref_end = int(ref_points[idx + 1])
        query_start = int(query_points[idx])
        query_end = int(query_points[idx + 1])
        ref_span = ref_end - ref_start
        query_span = query_end - query_start
        if ref_span > 0 and query_span > 0:
            if ref_span != query_span:
                raise RuntimeError("Unexpected alignment segment with unequal spans")
            ref_chunks.append(reference[ref_start:ref_end])
            query_chunks.append(query[query_start:query_end])
        elif ref_span > 0:
            ref_chunks.append(reference[ref_start:ref_end])
            query_chunks.append("-" * ref_span)
        elif query_span > 0:
            ref_chunks.append("-" * query_span)
            query_chunks.append(query[query_start:query_end])
    return "".join(ref_chunks), "".join(query_chunks)


def load_gt_refs(json_path: Path) -> dict[str, str]:
    rows = json.loads(json_path.read_text(encoding="utf-8"))
    refs: dict[str, str] = {}
    for row in rows:
        name = str(row.get("name", ""))
        match = re.fullmatch(r"HCV([1-8])NS3", name)
        if not match:
            continue
        refs[match.group(1)] = str(row.get("refSequence", "")).strip().upper()
    missing = [gt for gt in map(str, range(1, 9)) if gt not in refs]
    if missing:
        raise RuntimeError(f"Missing NS3 AA refs for GTs: {', '.join(missing)}")
    return refs


def load_subtype_rows(workbook_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(v) if v is not None else "" for v in next(ws.iter_rows(values_only=True))]
    index = {name: i for i, name in enumerate(header)}
    required = ["RefID", "RefName", "AccessionID", "ClosestGT"]
    for name in required:
        if name not in index:
            raise RuntimeError(f"Column '{name}' not found in {workbook_path}")
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {name: values[i] for name, i in index.items()}
        refid = str(row["RefID"]).strip()
        if not refid:
            continue
        row["RefID"] = refid
        row["AccessionID"] = str(row["AccessionID"]).strip()
        row["ClosestGT"] = str(row["ClosestGT"]).strip()
        rows.append(row)
    wb.close()
    return header, rows


def build_refid_to_fasta(fasta_dir: Path) -> dict[str, Path]:
    mapping: dict[str, Path] = {}
    for path in sorted(fasta_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in FASTA_EXTENSIONS:
            continue
        if "_" not in path.name:
            continue
        refid = path.name.split("_", 1)[0]
        if refid not in mapping:
            mapping[refid] = path
    return mapping


def write_fasta(path: Path, entries: list[tuple[str, str]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for header, sequence in entries:
            handle.write(f">{header}\n")
            for start in range(0, len(sequence), 70):
                handle.write(sequence[start : start + 70] + "\n")


def build_gt_db(job_dir: Path, gt: str, aa_sequence: str) -> Path:
    ref_fasta = job_dir / f"gt{gt}_ns3_aa.fasta"
    write_fasta(ref_fasta, [(f"GT{gt}_NS3", aa_sequence)])
    db_prefix = job_dir / f"gt{gt}_ns3_aa_db"
    subprocess.run(
        ["makeblastdb", "-in", str(ref_fasta), "-dbtype", "prot", "-out", str(db_prefix), "-parse_seqids"],
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        text=True,
    )
    return db_prefix


def run_blastx(query_fasta: Path, db_prefix: Path, out_path: Path) -> list[dict[str, Any]]:
    subprocess.run(
        [
            "blastx",
            "-query",
            str(query_fasta),
            "-db",
            str(db_prefix),
            "-seg",
            "no",
            "-evalue",
            "1e-6",
            "-max_hsps",
            "1",
            "-max_target_seqs",
            "1",
            "-outfmt",
            BLAST_OUTFMT,
            "-out",
            str(out_path),
        ],
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        text=True,
    )
    hits: list[dict[str, Any]] = []
    for line in out_path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        p = line.split("\t")
        hits.append(
            {
                "qseqid": p[0],
                "length": int(p[2]),
                "mismatch": int(p[3]),
                "gaps": int(p[4]),
                "bitscore": float(p[7]),
                "qstart": int(p[8]),
                "qend": int(p[9]),
                "sstart": int(p[10]),
                "send": int(p[11]),
                "qframe": int(p[12]),
            }
        )
    try:
        out_path.unlink()
    except OSError:
        pass
    return hits


def choose_best_hits(hits: list[dict[str, Any]], min_aa_overlap: int) -> dict[str, dict[str, Any]]:
    best: dict[str, dict[str, Any]] = {}
    for hit in hits:
        if hit["length"] < min_aa_overlap:
            continue
        current = best.get(hit["qseqid"])
        if current is None or (
            -hit["bitscore"], -hit["length"], hit["mismatch"] + hit["gaps"]
        ) < (
            -current["bitscore"], -current["length"], current["mismatch"] + current["gaps"]
        ):
            best[hit["qseqid"]] = hit
    return best


def extract_aa(sequence: str, hit: dict[str, Any], reference_aa: str) -> tuple[int, int, str, str]:
    qstart = min(hit["qstart"], hit["qend"])
    qend = max(hit["qstart"], hit["qend"])
    frame = hit["qframe"]
    if frame > 0:
        nt_window = normalize_nt(sequence[qstart - 1 : qend])
    else:
        nt_window = normalize_nt(reverse_complement(sequence[qstart - 1 : qend]))
    nt_window = nt_window[: len(nt_window) - (len(nt_window) % 3)]
    query_aa = translate_nt(nt_window)
    if not query_aa:
        return 0, 0, "", ""

    global_alignment = build_aligner().align(reference_aa, query_aa)[0]
    aligned_reference_aa, aligned_query_aa = alignment_strings(reference_aa, query_aa, global_alignment.coordinates)

    ref_pos = 0
    query_chars: list[str] = []
    start_aa: int | None = None
    end_aa: int | None = None
    for ref_char, query_char in zip(aligned_reference_aa, aligned_query_aa):
        if ref_char != "-":
            ref_pos += 1
            if query_char != "-":
                if start_aa is None:
                    start_aa = ref_pos
                end_aa = ref_pos
                query_chars.append(query_char)

    if start_aa is None or end_aa is None:
        return 0, 0, "", ""
    return start_aa, end_aa, "".join(query_chars), nt_window


def write_workbook(path: Path, header: list[str], rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "NS3_Subtype_Distance"
    base_header = [col for col in header if col not in {"StartAAPosition", "EndAAPosition", "AASequence", "NASequence"}]
    full_header = base_header + ["StartAAPosition", "EndAAPosition", "AASequence", "NASequence"]
    ws.append(full_header)
    for row in rows:
        ws.append([row.get(col, "") for col in full_header])
    wb.save(path)


def cleanup_job_dir(job_dir: Path) -> None:
    for path in job_dir.iterdir():
        if path.name.startswith("gt") and path.suffix in {".fasta", ".phr", ".pin", ".pog", ".psq", ".pto", ".ptf", ".pot"}:
            try:
                path.unlink()
            except OSError:
                pass
        if path.name.startswith("queries_gt") and path.suffix == ".fasta":
            try:
                path.unlink()
            except OSError:
                pass


def main() -> int:
    args = parse_args()
    subtype_workbook = Path(args.subtype_workbook).expanduser()
    fasta_dir = Path(args.fasta_dir).expanduser()
    gt_aa_json = Path(args.gt_aa_json).expanduser()
    output_dir = Path(args.output_dir)

    header, rows = load_subtype_rows(subtype_workbook)
    gt_refs = load_gt_refs(gt_aa_json)
    refid_to_fasta = build_refid_to_fasta(fasta_dir)
    job_dir = make_job_dir(output_dir, subtype_workbook)

    rows_by_gt: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        rows_by_gt[row["ClosestGT"]].append(row)

    fasta_cache: dict[str, dict[str, str]] = {}
    results: dict[tuple[str, str], tuple[int, int, str, str]] = {}
    summary: dict[str, Any] = {"input_rows": len(rows), "by_gt": {}}

    for gt, group_rows in sorted(rows_by_gt.items(), key=lambda item: int(item[0])):
        print(f"extract_aa_genotype=GT{gt}", flush=True)
        query_entries: list[tuple[str, str]] = []
        seq_by_qseqid: dict[str, str] = {}
        for row in group_rows:
            fasta_path = refid_to_fasta.get(row["RefID"])
            if fasta_path is None:
                continue
            fasta_map = fasta_cache.get(row["RefID"])
            if fasta_map is None:
                fasta_map = {accession_from_header(h): s for h, s in parse_fasta(fasta_path)}
                fasta_cache[row["RefID"]] = fasta_map
            seq = fasta_map.get(row["AccessionID"])
            if not seq:
                continue
            qseqid = f"{row['RefID']}|{row['AccessionID']}"
            query_entries.append((qseqid, seq))
            seq_by_qseqid[qseqid] = seq

        query_fasta = job_dir / f"queries_gt{gt}.fasta"
        write_fasta(query_fasta, query_entries)
        db_prefix = build_gt_db(job_dir, gt, gt_refs[gt])
        hits = run_blastx(query_fasta, db_prefix, job_dir / f"queries_gt{gt}.blast.tsv")
        best_hits = choose_best_hits(hits, args.min_aa_overlap)

        extracted = 0
        reference_aa = gt_refs[gt]
        for qseqid, hit in best_hits.items():
            start_aa, end_aa, aa_sequence, na_sequence = extract_aa(seq_by_qseqid[qseqid], hit, reference_aa)
            results[tuple(qseqid.split("|", 1))] = (start_aa, end_aa, aa_sequence, na_sequence)
            extracted += 1
        summary["by_gt"][gt] = {"queries": len(query_entries), "extracted": extracted}

    for row in rows:
        key = (row["RefID"], row["AccessionID"])
        value = results.get(key)
        if value is None:
            row["StartAAPosition"] = ""
            row["EndAAPosition"] = ""
            row["AASequence"] = ""
            row["NASequence"] = ""
        else:
            row["StartAAPosition"], row["EndAAPosition"], row["AASequence"], row["NASequence"] = value

    out_path = (
        Path(args.output_workbook).expanduser()
        if args.output_workbook
        else make_temp_output_path("NS3_Subtype_Alignments_with_GT_AA_Extraction.xlsx")
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    write_workbook(out_path, header, rows)
    summary["output_workbook"] = str(out_path.resolve())
    summary["rows_with_aa"] = sum(1 for row in rows if row.get("AASequence"))
    cleanup_job_dir(job_dir)
    shutil.rmtree(job_dir, ignore_errors=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
