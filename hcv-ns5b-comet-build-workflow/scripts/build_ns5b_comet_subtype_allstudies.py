#!/usr/bin/env python3
"""Create the NS5B subtype workbook directly from Comet assignments."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--genotype-workbook", required=True)
    parser.add_argument("--comet-subtype-csv", required=True)
    parser.add_argument("--output-dir", required=True)
    return parser.parse_args()


def load_subtypes(path: Path) -> dict[str, str]:
    assignments: dict[str, str] = {}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            accession = (row.get("accession") or "").strip()
            subtype = (row.get("subtype") or "").strip().lower()
            if accession and subtype:
                assignments[accession] = subtype
                assignments.setdefault(accession.split(".", 1)[0], subtype)
    return assignments


def main() -> int:
    args = parse_args()
    assignments = load_subtypes(Path(args.comet_subtype_csv))
    workbook = load_workbook(args.genotype_workbook, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header = [str(value) if value is not None else "" for value in next(sheet.iter_rows(values_only=True))]
    index = {name: position for position, name in enumerate(header)}
    required = ["RefID", "RefName", "GenBankAccession", "BestGT"]
    missing = [name for name in required if name not in index]
    if missing:
        raise RuntimeError(f"Columns missing from {args.genotype_workbook}: {', '.join(missing)}")

    rows: list[tuple[str, str, str, str, str]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        accession = str(values[index["GenBankAccession"]]).strip()
        subtype = assignments.get(accession) or assignments.get(accession.split(".", 1)[0])
        if subtype:
            rows.append((
                str(values[index["RefID"]]).strip(),
                str(values[index["RefName"]]).strip(),
                accession,
                str(values[index["BestGT"]]).strip(),
                subtype,
            ))
    workbook.close()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "NS5B_Subtype_AllStudies_WSeqs.xlsx"
    output = Workbook()
    sheet = output.active
    sheet.title = "NS5B_Subtype_Comet"
    sheet.append([
        "RefID", "RefName", "AccessionID", "ClosestGT", "ClosestSubtype",
        "ClosestSubtypeAssignmentSource", "ClosestSubtypeMetadataColumn",
    ])
    for refid, refname, accession, genotype, subtype in rows:
        sheet.append([refid, refname, accession, genotype, subtype, "Comet", "Comet NS5B"])
    output.save(output_path)
    print(json.dumps({"output_workbook": str(output_path.resolve()), "row_count": len(rows), "comet_subtype_count": len(rows)}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
