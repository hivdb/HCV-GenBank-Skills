#!/usr/bin/env python3
"""Compare NS3 profile-derived subtype consensus to matching subtype references."""
from __future__ import annotations
import argparse,json,re
from pathlib import Path
from Bio import Align
from openpyxl import Workbook,load_workbook
CODONS={'TTT':'F','TTC':'F','TTA':'L','TTG':'L','TCT':'S','TCC':'S','TCA':'S','TCG':'S','TAT':'Y','TAC':'Y','TAA':'*','TAG':'*','TGT':'C','TGC':'C','TGA':'*','TGG':'W','CTT':'L','CTC':'L','CTA':'L','CTG':'L','CCT':'P','CCC':'P','CCA':'P','CCG':'P','CAT':'H','CAC':'H','CAA':'Q','CAG':'Q','CGT':'R','CGC':'R','CGA':'R','CGG':'R','ATT':'I','ATC':'I','ATA':'I','ATG':'M','ACT':'T','ACC':'T','ACA':'A','ACG':'T','AAT':'N','AAC':'N','AAA':'K','AAG':'K','AGT':'S','AGC':'S','AGA':'R','AGG':'R','GTT':'V','GTC':'V','GTA':'V','GTG':'V','GCT':'A','GCC':'A','GCA':'A','GCG':'A','GAT':'D','GAC':'D','GAA':'E','GAG':'E','GGT':'G','GGC':'G','GGA':'G','GGG':'G'}
VALID=set('ACDEFGHIKLMNPQRSTVWY*')
RAS_POSITIONS=(36,41,43,54,55,56,80,122,155,156,158,166,168,170,175)
def aligned_strings(reference,query,coordinates):
 out=[[],[]]
 for n in range(len(coordinates[0])-1):
  rs,re,qs,qe=(int(coordinates[0][n]),int(coordinates[0][n+1]),int(coordinates[1][n]),int(coordinates[1][n+1]))
  if re-rs and qe-qs: out[0].append(reference[rs:re]); out[1].append(query[qs:qe])
  elif re-rs: out[0].append(reference[rs:re]); out[1].append('-'*(re-rs))
  else: out[0].append('-'*(qe-qs)); out[1].append(query[qs:qe])
 return ''.join(out[0]),''.join(out[1])
def main():
 p=argparse.ArgumentParser(); p.add_argument('--subtype-profile-workbook',required=True); p.add_argument('--subtype-json',required=True); p.add_argument('--output-xlsx',required=True); p.add_argument('--positions',default=','.join(map(str,RAS_POSITIONS))); a=p.parse_args()
 positions=tuple(sorted({int(pos) for pos in a.positions.split(',') if pos.strip()}))
 refs={}
 for r in json.loads(Path(a.subtype_json).read_text()):
  m=re.search(r'Genotype\s*(\d+[A-Za-z0-9]*)',str(r.get('genotypeName','')))
  if m and m.group(1).lower() not in refs:
   nt=str(r.get('sequence','')).upper(); refs[m.group(1).lower()]=(str(r.get('accession','')),''.join(CODONS.get(nt[i:i+3],'X') for i in range(0,min(len(nt),1893)-2,3)))
 wb=load_workbook(a.subtype_profile_workbook,read_only=True,data_only=True); out=Workbook(); wsout=out.active; wsout.title='subtype_reference_distance'; wsout.append(['Genotype','Subtype','ReferenceAccession','ComparedAA','Differences','DifferencePositions','Distance','Status'])
 for sheet in wb.sheetnames:
  gt=sheet.removeprefix('GT'); ws=wb[sheet]; h=[str(c.value or '') for c in next(ws.iter_rows(max_row=1))]; ix={v:i for i,v in enumerate(h)}; seqs={}
  for row in ws.iter_rows(min_row=2,values_only=True):
   st=str(row[ix['Subtype']] or '').strip().lower(); pos=row[ix[[x for x in h if x.endswith('Position')][0]]]; aa=str(row[ix['AminoAcid']] or '').upper(); pct=row[ix['PctWithAA']]
   if st and pos and pct is not None:
    calls=seqs.setdefault(st,{}); current=calls.get(int(pos))
    if current is None or float(pct)>current[0] or (float(pct)==current[0] and aa<current[1]): calls[int(pos)]=(float(pct),aa)
  for st,calls in sorted(seqs.items()):
   ref=refs.get(st); key=f'{gt}{st[len(gt):]}' if st.startswith(gt) else st
   if ref is None: wsout.append([gt,st,'','','','','','reference_not_found']); continue
   acc,refaa=ref; consensus=''.join(calls.get(pos,(0,'X'))[1] for pos in range(1,len(refaa)+1)); aligner=Align.PairwiseAligner(mode='global'); aligner.match_score=2; aligner.mismatch_score=-1; aligner.open_gap_score=-5; aligner.extend_gap_score=-1; alignment=aligner.align(refaa,consensus)[0]; aligned_ref,aligned_consensus=aligned_strings(refaa,consensus,alignment.coordinates); refpos=0; pairs=[]
   for refaa_char,consensus_char in zip(aligned_ref,aligned_consensus):
    if refaa_char!='-': refpos+=1
    if refaa_char!='-' and refpos in positions and refaa_char in VALID and consensus_char in VALID: pairs.append((refpos,refaa_char,consensus_char))
   mismatches=[f'{pos}:{refaa_char}>{consensus_char}' for pos,refaa_char,consensus_char in pairs if refaa_char!=consensus_char]; diff=len(mismatches); wsout.append([gt,st,acc,len(pairs),diff,';'.join(mismatches),diff/len(pairs) if pairs else '', 'ok' if pairs else 'no_comparable_positions'])
 wb.close(); meta=out.create_sheet('metadata'); meta.append(['aa_positions',','.join(map(str,positions))]); meta.append(['alignment','global AA alignment of profile consensus to translated subtype genome reference']); meta.append(['distance_definition','AA differences / comparable aligned NS3 RAS positions']); Path(a.output_xlsx).parent.mkdir(parents=True,exist_ok=True); out.save(a.output_xlsx)
if __name__=='__main__': main()
