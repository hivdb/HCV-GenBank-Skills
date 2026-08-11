#!/usr/bin/env python3
"""Summarize subtype RAS AA differences from genotype consensus for combined profiles."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from statistics import median

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


VALID_AAS = set("ACDEFGHIKLMNPQRSTVWY")
POSITION_HEADER_RE = re.compile(r"^P(\d+)$")
SUBTYPE_LABEL_RE = re.compile(r"^GT(?P<genotype>\d+)_(?P<subtype>\S+)\s+\((?P<count>\d+),")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--combined-profile-workbook", required=True)
    parser.add_argument("--profile-input-workbook", required=True)
    parser.add_argument("--profile-accessions-csv", required=True)
    parser.add_argument("--gt-aa-json", required=True)
    parser.add_argument("--gene", default="NS3", help="Gene token in HCV<GT><GENE> reference names.")
    parser.add_argument("--consensus-gene", help="Reference JSON gene token; defaults to --gene.")
    parser.add_argument(
        "--output-xlsx",
        default="outputs/NS3_Subtype_RAS_Consensus_Difference_Summary.xlsx",
    )
    parser.add_argument("--output-png", help="Optional PNG path for the subtype trend charts.")
    return parser.parse_args()


def load_combined_subtypes(path: Path) -> tuple[dict[tuple[str, str], int], tuple[int, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    header = [str(value or "") for value in next(worksheet.iter_rows(values_only=True))]
    positions = tuple(
        int(match.group(1))
        for value in header[1:]
        if (match := POSITION_HEADER_RE.fullmatch(value))
    )
    if not positions:
        raise RuntimeError(f"No RAS position columns (P<number>) found in {path}")

    subtypes: dict[tuple[str, str], int] = {}
    for (label, *_) in worksheet.iter_rows(min_row=2, values_only=True):
        match = SUBTYPE_LABEL_RE.match(str(label or ""))
        if match:
            key = (match.group("genotype"), match.group("subtype").lower())
            subtypes[key] = int(match.group("count"))
    workbook.close()
    if not subtypes:
        raise RuntimeError(f"No subtype rows found in combined profile workbook: {path}")
    return subtypes, positions


def load_profile_accessions(path: Path) -> set[str]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return {
            row["accession"].strip()
            for row in csv.DictReader(handle)
            if row.get("accession", "").strip()
        }


def load_consensus_by_genotype(path: Path, gene: str) -> dict[str, str]:
    rows = json.loads(path.read_text(encoding="utf-8"))
    consensus: dict[str, str] = {}
    for row in rows:
        match = re.fullmatch(rf"HCV([1-8]){re.escape(gene)}", str(row.get("name", "")))
        if match:
            consensus[match.group(1)] = str(row.get("refSequence", "")).strip().upper()
    if not consensus:
        raise RuntimeError(f"No HCV<GT>{gene} consensus sequences found in {path}")
    return consensus


def load_sequence_differences(
    workbook_path: Path,
    profile_accessions: set[str],
    included_subtypes: dict[tuple[str, str], int],
    consensus_by_genotype: dict[str, str],
    positions: tuple[int, ...],
) -> tuple[
    dict[tuple[str, str], list[int]],
    dict[tuple[str, str], list[int]],
    Counter[str],
    Counter[tuple[str, str]],
]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook.active
    header = [str(value or "") for value in next(worksheet.iter_rows(values_only=True))]
    index = {name: position for position, name in enumerate(header)}
    required = {"AccessionID", "ClosestGT", "ClosestSubtype", "StartAAPosition", "AASequence"}
    missing = required - index.keys()
    if missing:
        raise RuntimeError(f"Missing columns in {workbook_path}: {', '.join(sorted(missing))}")

    differences: dict[tuple[str, str], list[int]] = defaultdict(list)
    range_differences: dict[tuple[str, str], list[int]] = defaultdict(list)
    exclusions: Counter[str] = Counter()
    found_records: Counter[tuple[str, str]] = Counter()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        accession = str(row[index["AccessionID"]] or "").strip()
        if accession not in profile_accessions:
            continue
        genotype = str(row[index["ClosestGT"]] or "").strip().removeprefix("GT")
        subtype = str(row[index["ClosestSubtype"]] or "").strip().lower()
        key = (genotype, subtype)
        if key not in included_subtypes:
            continue
        found_records[key] += 1
        start = row[index["StartAAPosition"]]
        sequence = str(row[index["AASequence"]] or "").strip().upper()
        if not start or not sequence:
            exclusions["missing_AA_sequence_or_start_position"] += 1
            continue
        consensus = consensus_by_genotype.get(genotype)
        if consensus is None:
            exclusions["missing_genotype_consensus"] += 1
            continue
        calls = {int(start) + offset: amino_acid for offset, amino_acid in enumerate(sequence)}
        if any(position not in calls or calls[position] not in VALID_AAS for position in positions):
            exclusions["missing_or_ambiguous_AA_at_RAS_position"] += 1
            continue
        if any(position > len(consensus) or consensus[position - 1] not in VALID_AAS for position in positions):
            exclusions["missing_or_ambiguous_consensus_AA_at_RAS_position"] += 1
            continue
        differences[key].append(sum(calls[position] != consensus[position - 1] for position in positions))
        position_range = range(min(positions), max(positions) + 1)
        if all(
            calls.get(position) in VALID_AAS
            and position <= len(consensus)
            and consensus[position - 1] in VALID_AAS
            for position in position_range
        ):
            range_differences[key].append(
                sum(calls[position] != consensus[position - 1] for position in position_range)
            )
        else:
            exclusions["missing_or_ambiguous_AA_in_RAS_position_range"] += 1
    workbook.close()
    return differences, range_differences, exclusions, found_records


def write_workbook(
    path: Path,
    gene: str,
    included_subtypes: dict[tuple[str, str], int],
    positions: tuple[int, ...],
    differences: dict[tuple[str, str], list[int]],
    range_differences: dict[tuple[str, str], list[int]],
    exclusions: Counter[str],
    found_records: Counter[tuple[str, str]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Subtype_RAS_Differences"
    worksheet.append([
        "Genotype", "Subtype", "ProfileSequencesFound", "SequencesCompared",
        "SequencesExcluded", "RASPositionCount", "TotalAADifferences",
        "MeanAADifferencesPerSequence", "MedianAADifferencesPerSequence",
        "RangeSequencesCompared", f"MeanAADifferencesPerSequence_Pos{min(positions)}_{max(positions)}",
        f"MedianAADifferencesPerSequence_Pos{min(positions)}_{max(positions)}",
    ])
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")

    for key in sorted(included_subtypes, key=lambda item: (int(item[0]), item[1])):
        scores = differences.get(key, [])
        range_scores = range_differences.get(key, [])
        found = found_records[key]
        worksheet.append([
            f"GT{key[0]}", key[1], found, len(scores), found - len(scores), len(positions),
            sum(scores), sum(scores) / len(scores) if scores else None,
            median(scores) if scores else None,
            len(range_scores), sum(range_scores) / len(range_scores) if range_scores else None,
            median(range_scores) if range_scores else None,
        ])
    for row in worksheet.iter_rows(min_row=2, min_col=8, max_col=12):
        for cell in row:
            cell.number_format = "0.00"
    for column, width in {"A": 12, "B": 12, "C": 22, "D": 20, "E": 18, "F": 18, "G": 20, "H": 31, "I": 33, "J": 24, "K": 37, "L": 39}.items():
        worksheet.column_dimensions[column].width = width

    metadata = workbook.create_sheet("Metadata")
    metadata.append(["Metric", "Value"])
    metadata.append(["RAS positions", ",".join(map(str, positions))])
    metadata.append(["Comparison", f"Each retained profile sequence versus its HCV<GT>{gene} AA consensus"])
    metadata.append(["Mean", "Total RAS AA differences / sequences compared"])
    metadata.append(["Median", "Median per-sequence RAS AA difference count"])
    metadata.append(["Range mean", f"Mean per-sequence AA differences across positions {min(positions)}-{max(positions)}"])
    metadata.append(["Range median", f"Median per-sequence AA differences across positions {min(positions)}-{max(positions)}"])
    metadata.append(["Range inclusion", f"Sequence must have unambiguous AA calls at every position from {min(positions)} through {max(positions)}"])
    metadata.append(["Missing or ambiguous RAS AA", "Sequence excluded from both mean and median"])
    exclusion_sheet = workbook.create_sheet("Excluded_Sequences")
    exclusion_sheet.append(["Reason", "SequenceCount"])
    for reason, count in sorted(exclusions.items()):
        exclusion_sheet.append([reason, count])
    exclusion_sheet.append(["total_excluded", sum(exclusions.values())])
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for column in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = max(
                sheet.column_dimensions[get_column_letter(column)].width or 0, 16
            )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_trend_png(
    path: Path,
    gene: str,
    included_subtypes: dict[tuple[str, str], int],
    positions: tuple[int, ...],
    differences: dict[tuple[str, str], list[int]],
    range_differences: dict[tuple[str, str], list[int]],
) -> None:
    keys = sorted(included_subtypes, key=lambda item: (int(item[0]), item[1]))
    labels = [f"GT{genotype}_{subtype}" for genotype, subtype in keys]
    ras_means = [sum(differences[key]) / len(differences[key]) if differences[key] else float("nan") for key in keys]
    ras_medians = [median(differences[key]) if differences[key] else float("nan") for key in keys]
    range_means = [sum(range_differences[key]) / len(range_differences[key]) if range_differences[key] else float("nan") for key in keys]
    range_medians = [median(range_differences[key]) if range_differences[key] else float("nan") for key in keys]

    figure, axes = plt.subplots(2, 1, figsize=(max(14, len(keys) * 0.48), 10), sharex=True, constrained_layout=True)
    x_values = range(len(keys))
    for axis, title, mean_values, median_values in (
        (axes[0], f"{gene} RAS positions", ras_means, ras_medians),
        (axes[1], f"{gene} AA positions {min(positions)}-{max(positions)}", range_means, range_medians),
    ):
        axis.plot(x_values, mean_values, marker="o", linewidth=1.5, label="Mean difference")
        axis.plot(x_values, median_values, marker="s", linewidth=1.5, label="Median difference")
        axis.set_title(title)
        axis.set_ylabel("AA differences per sequence")
        axis.grid(axis="y", alpha=0.3)
        axis.legend()
    axes[1].set_xticks(list(x_values), labels, rotation=65, ha="right")
    path.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(path, dpi=200, bbox_inches="tight")
    plt.close(figure)


def main() -> int:
    args = parse_args()
    combined_path = Path(args.combined_profile_workbook).expanduser()
    input_path = Path(args.profile_input_workbook).expanduser()
    accessions_path = Path(args.profile_accessions_csv).expanduser()
    consensus_path = Path(args.gt_aa_json).expanduser()
    output_path = Path(args.output_xlsx).expanduser()
    png_path = Path(args.output_png).expanduser() if args.output_png else output_path.with_suffix(".png")
    for path in (combined_path, input_path, accessions_path, consensus_path):
        if not path.is_file():
            raise RuntimeError(f"Required input file not found: {path}")
    included_subtypes, positions = load_combined_subtypes(combined_path)
    differences, range_differences, exclusions, found_records = load_sequence_differences(
        input_path, load_profile_accessions(accessions_path), included_subtypes,
        load_consensus_by_genotype(consensus_path, args.consensus_gene or args.gene), positions,
    )
    write_workbook(output_path, args.gene, included_subtypes, positions, differences, range_differences, exclusions, found_records)
    write_trend_png(png_path, args.gene, included_subtypes, positions, differences, range_differences)
    print(json.dumps({
        "output_xlsx": str(output_path.resolve()),
        "output_png": str(png_path.resolve()),
        "subtype_count": len(included_subtypes),
        "sequences_compared": sum(len(values) for values in differences.values()),
        "sequences_excluded": sum(exclusions.values()),
        "ras_positions": list(positions),
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
