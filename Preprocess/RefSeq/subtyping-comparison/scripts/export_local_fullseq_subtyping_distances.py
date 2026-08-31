#!/usr/bin/env python3
"""Export full-sequence distance rows for Step 11 disagreement accessions."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[4]
STEP11_DIR = REPO_ROOT / "HCVData" / "subtyping-comparison" / "11-report-subtype-agreement"
SOURCE_XLSX = REPO_ROOT / "HCVData" / "nonComet-Full-genome" / "Subtyping_Distances.xlsx"
GENES = ("NS3", "NS5A", "NS5B")
SPECIAL_COMET_FULLSEQ_SUBTYPES = ("1L", "3B", "3H", "4R")


def accession_key(value: object) -> str:
    return str(value or "").strip().split(".", 1)[0].upper()


def selected_accessions(
    gene: str,
) -> tuple[set[str], dict[str, set[str]], dict[str, str]]:
    selected: set[str] = set()
    selected_by_subtype = {subtype: set() for subtype in SPECIAL_COMET_FULLSEQ_SUBTYPES}
    comet_fullseq_subtypes: dict[str, str] = {}
    path = STEP11_DIR / f"{gene}_Subtype_Agreement_By_Accession.csv"
    with path.open(encoding="utf-8-sig", newline="") as source:
        reader = csv.DictReader(source)
        required = {"Accession", "CometFullSeqSubtype"}
        if not reader.fieldnames or (missing := required - set(reader.fieldnames)):
            raise ValueError(f"{path} is missing columns: {', '.join(sorted(missing))}")
        for row in reader:
            accession = accession_key(row.get("Accession", ""))
            if not accession:
                continue
            selected.add(accession)
            subtype = str(row.get("CometFullSeqSubtype", "")).strip().upper()
            comet_fullseq_subtypes[accession] = subtype
            if subtype in selected_by_subtype:
                selected_by_subtype[subtype].add(accession)
    return selected, selected_by_subtype, comet_fullseq_subtypes


def export_gene(gene: str) -> None:
    selected, selected_by_subtype, comet_fullseq_subtypes = selected_accessions(gene)
    output_paths = {
        "all": STEP11_DIR / f"{gene}_Subtyping_Distances.xlsx",
        **{
            subtype: STEP11_DIR
            / f"{gene}_CometFullSeqSubtype_{subtype}_Subtyping_Distances.xlsx"
            for subtype in SPECIAL_COMET_FULLSEQ_SUBTYPES
        },
    }
    selected_sets = {"all": selected, **selected_by_subtype}
    legacy_outputs = [
        STEP11_DIR / f"LocalFullSeq_{gene}_Subtyping_Distances.xlsx",
        STEP11_DIR / f"{gene}_LocalFullSeq_Subtyping_Distances.xlsx",
        *[
            STEP11_DIR
            / f"{gene}_LocalFullSeq_CometFullSeqSubtype_{subtype}_Subtyping_Distances.xlsx"
            for subtype in SPECIAL_COMET_FULLSEQ_SUBTYPES
        ],
    ]
    for legacy_output in legacy_outputs:
        if legacy_output.exists():
            legacy_output.unlink()
    source = load_workbook(SOURCE_XLSX, read_only=True, data_only=False)
    destinations = {label: Workbook(write_only=True) for label in output_paths}
    matched_genotype_accessions = {label: set() for label in output_paths}
    try:
        for source_sheet in source.worksheets:
            target_sheets = {
                label: workbook.create_sheet(source_sheet.title)
                for label, workbook in destinations.items()
            }
            rows = source_sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                continue
            output_header = (
                header[0],
                "CometFullSeqSubtype",
                *header[1:],
            )
            for target_sheet in target_sheets.values():
                target_sheet.append(output_header)
            matched_rows = {label: 0 for label in output_paths}
            for row in rows:
                accession = accession_key(row[0] if row else "")
                for label, selected_set in selected_sets.items():
                    if accession not in selected_set:
                        continue
                    target_sheets[label].append(
                        (
                            row[0],
                            comet_fullseq_subtypes.get(accession, ""),
                            *row[1:],
                        )
                    )
                    matched_rows[label] += 1
                    if source_sheet.title == "Genotype":
                        matched_genotype_accessions[label].add(accession)
            print(
                f"{gene} {source_sheet.title}: "
                + ", ".join(f"{label}={count:,}" for label, count in matched_rows.items())
            )
        for label, workbook in destinations.items():
            output_paths[label].parent.mkdir(parents=True, exist_ok=True)
            workbook.save(output_paths[label])
    finally:
        source.close()

    print(f"{gene} Step 11 unique accessions: {len(selected):,}")
    for label, output_xlsx in output_paths.items():
        selected_set = selected_sets[label]
        matched = matched_genotype_accessions[label]
        print(f"{gene} {label} selected accessions: {len(selected_set):,}")
        print(f"{gene} {label} missing from Genotype sheet: {len(selected_set - matched):,}")
        print(f"{gene} {label} output: {output_xlsx}")


def main() -> None:
    obsolete_output = STEP11_DIR / "LocalFullSeq_Subtyping_Distances.xlsx"
    if obsolete_output.exists():
        obsolete_output.unlink()
    for gene in GENES:
        export_gene(gene)


if __name__ == "__main__":
    main()
