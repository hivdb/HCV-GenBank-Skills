#!/usr/bin/env python3
"""Add COMET calls and selected non-COMET priority subtypes to each Ref.csv row."""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


PRIORITY_SUBTYPES = {"1d", "7a", "7b", "8a"}
COVERAGE_COLUMNS = {
    "NS3": "IncludeNS3Pos36_175",
    "NS5A": "IncludeNS5APos26_93",
    "NS5B": "Includes S282 + all positions",
}
LEGACY_COVERAGE_COLUMNS = {
    "Includes S282 + all positions": "IncludeNS5BPos150_321",
}
NS5B_S282_COLUMN = "Includes S282"
NS5B_S282_PLUS_FOUR_COLUMN = "Includes S282 + 4 other RAS positions"
NS5B_S282_PLUS_FIVE_COLUMN = "Includes S282 + 5 other RAS positions"
NS5B_RAS_POSITIONS = (150, 159, 206, 282, 316, 320, 321)
NS5B_OTHER_RAS_POSITIONS = (150, 159, 206, 316, 320, 321)
NS5B_ANNOTATION_COLUMNS = (
    NS5B_S282_COLUMN,
    NS5B_S282_PLUS_FOUR_COLUMN,
    NS5B_S282_PLUS_FIVE_COLUMN,
    COVERAGE_COLUMNS["NS5B"],
)
ANNOTATION_COLUMNS = (
    COVERAGE_COLUMNS["NS3"],
    COVERAGE_COLUMNS["NS5A"],
    *NS5B_ANNOTATION_COLUMNS,
)
OBSOLETE_NS5B_COLUMNS = ("Includes 4 other RAS positions",)
REPO_ROOT = Path(__file__).resolve().parents[3]
DATA_DIR = REPO_ROOT / "HCVData" / "HCV-all-seq-subtype"
NONCOMET_COVERAGE_DIR = REPO_ROOT / "HCVData" / "nonComet-Full-genome"


def accession_key(value: str | None) -> str:
    return (value or "").strip().split(".", 1)[0]


def reference_overlap_interval(value: str | None) -> tuple[int, int] | None:
    match = re.fullmatch(r"\s*(\d+)\s*-\s*(\d+)\s*", value or "")
    if match is None:
        return None
    return tuple(sorted((int(match.group(1)), int(match.group(2)))))


def reorder_columns(
    worksheet, headers: dict[object, int], columns: tuple[str, ...]
) -> None:
    """Reorder existing worksheet columns without changing their data or formatting."""
    if not all(column in headers for column in columns):
        return
    target_columns = sorted(headers[column] for column in columns)
    snapshots = [
        [
            (cell.value, copy(cell._style), cell.comment, cell.hyperlink)
            for cell in worksheet.iter_cols(
                min_col=headers[column], max_col=headers[column]
            )
            for cell in cell
        ]
        for column in columns
    ]
    for target_column, snapshot in zip(target_columns, snapshots):
        for row_index, (value, style, comment, hyperlink) in enumerate(
            snapshot, start=1
        ):
            cell = worksheet.cell(row=row_index, column=target_column)
            cell.value, cell._style, cell.comment, cell._hyperlink = (
                value,
                style,
                comment,
                hyperlink,
            )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--ref-csv", type=Path, default=DATA_DIR / "Ref.csv")
    parser.add_argument(
        "--accessions-csv", type=Path, default=DATA_DIR / "Accessions.csv"
    )
    parser.add_argument(
        "--comet-csv", type=Path, default=DATA_DIR / "all_comet_subtype.csv"
    )
    parser.add_argument(
        "--coverage-csv",
        type=Path,
        action="append",
        help="Coverage CSV to supply priority non-COMET subtypes; repeat as needed.",
    )
    parser.add_argument(
        "--output-csv", type=Path, default=DATA_DIR / "Ref_with_CometSubtypes.csv"
    )
    parser.add_argument(
        "--blast-hists-xlsx",
        type=Path,
        default=DATA_DIR / "HCV_BlastHists_202604_data_Aug19.xlsx",
        help="Workbook whose Original_NS3, Original_NS5A, and Original_NS5B sheets are annotated.",
    )
    args = parser.parse_args()

    with args.comet_csv.open(newline="", encoding="utf-8-sig") as handle:
        comet_subtypes = {
            accession_key(row.get("name")): (row.get("subtype") or "").strip()
            for row in csv.DictReader(handle)
            if accession_key(row.get("name"))
        }

    subtypes_by_refid: dict[str, set[str]] = defaultdict(set)
    refid_by_accession: dict[str, str] = {}
    with args.accessions_csv.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            refid = (row.get("RefID") or "").strip()
            accession = accession_key(row.get("Accession"))
            if accession and refid:
                refid_by_accession[accession] = refid
            subtype = comet_subtypes.get(accession, "")
            if refid and subtype and "unassigned" not in subtype.casefold():
                subtypes_by_refid[refid].add(subtype)

    coverage_paths = args.coverage_csv or [
        NONCOMET_COVERAGE_DIR / "NS3_AllSeq_NonComet_Coverage.csv",
        NONCOMET_COVERAGE_DIR / "NS5A_AllSeq_NonComet_Coverage.csv",
        NONCOMET_COVERAGE_DIR / "NS5B_AllSeq_NonComet_Coverage.csv",
    ]
    priority_added = 0
    coverage_by_refid: dict[str, dict[str, set[str]]] = defaultdict(
        lambda: defaultdict(set)
    )
    for coverage_path in coverage_paths:
        gene = next(
            (
                value
                for value in COVERAGE_COLUMNS
                if coverage_path.name.startswith(f"{value}_")
            ),
            None,
        )
        if gene is None:
            raise ValueError(
                f"Cannot infer gene from coverage filename: {coverage_path}"
            )
        with coverage_path.open(newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                accession = accession_key(row.get("Accession"))
                subtype = (row.get("ClosestSubtype") or "").strip()
                refid = refid_by_accession.get(accession, "")
                if refid and subtype.casefold() in PRIORITY_SUBTYPES:
                    before = len(subtypes_by_refid[refid])
                    subtypes_by_refid[refid].add(subtype)
                    priority_added += len(subtypes_by_refid[refid]) - before
                coverage_column = COVERAGE_COLUMNS[gene]
                if gene == "NS5B" and refid:
                    interval = reference_overlap_interval(row.get("ReferenceOverlapAA"))
                    if interval is None:
                        continue
                    start, end = interval
                    includes_s282 = start <= 282 <= end
                    includes_all_ras = all(
                        start <= position <= end for position in NS5B_RAS_POSITIONS
                    )
                    other_ras_count = sum(
                        start <= position <= end
                        for position in NS5B_OTHER_RAS_POSITIONS
                    )
                    if includes_all_ras:
                        coverage_by_refid[refid][coverage_column].add(accession)
                    if includes_s282:
                        coverage_by_refid[refid][NS5B_S282_COLUMN].add(accession)
                    if includes_s282 and other_ras_count >= 4:
                        coverage_by_refid[refid][NS5B_S282_PLUS_FOUR_COLUMN].add(
                            accession
                        )
                    if includes_s282 and other_ras_count >= 5:
                        coverage_by_refid[refid][NS5B_S282_PLUS_FIVE_COLUMN].add(
                            accession
                        )
                elif (
                    refid and (row.get("FullyCover") or "").strip().casefold() == "yes"
                ):
                    coverage_by_refid[refid][coverage_column].add(accession)

    annotations_by_refid: dict[str, dict[str, str | int]] = {}
    with args.ref_csv.open(newline="", encoding="utf-8-sig") as source:
        reader = csv.DictReader(source)
        if not reader.fieldnames or "RefID" not in reader.fieldnames:
            raise ValueError(f"{args.ref_csv} must contain a RefID column")
        fields = [*reader.fieldnames, "CometSubtypes", *ANNOTATION_COLUMNS]
        with args.output_csv.open("w", newline="", encoding="utf-8") as destination:
            writer = csv.DictWriter(destination, fieldnames=fields, lineterminator="\n")
            writer.writeheader()
            for row in reader:
                refid = (row.get("RefID") or "").strip()
                row["CometSubtypes"] = "; ".join(
                    sorted(subtypes_by_refid.get(refid, set()))
                )
                for column in ANNOTATION_COLUMNS:
                    row[column] = len(coverage_by_refid[refid][column])
                writer.writerow(row)
                if refid:
                    annotations_by_refid[refid] = {
                        "CometSubtypes": row["CometSubtypes"] or "(unassigned)",
                        **{column: row[column] for column in ANNOTATION_COLUMNS},
                    }

    workbook = load_workbook(args.blast_hists_xlsx)
    for gene, coverage_column in COVERAGE_COLUMNS.items():
        worksheet = workbook[f"Original_{gene}"]
        headers = {cell.value: cell.column for cell in worksheet[1]}
        if "RefID" not in headers:
            raise ValueError(f"{worksheet.title} must contain a RefID column")
        if gene == "NS5B":
            for column in sorted(
                (headers[name] for name in OBSOLETE_NS5B_COLUMNS if name in headers),
                reverse=True,
            ):
                worksheet.delete_cols(column)
            headers = {cell.value: cell.column for cell in worksheet[1]}
        annotation_columns = [coverage_column]
        if gene == "NS5B":
            annotation_columns = list(NS5B_ANNOTATION_COLUMNS)
        for column in ("CometSubtypes", *annotation_columns):
            legacy_column = LEGACY_COVERAGE_COLUMNS.get(column)
            if column not in headers and legacy_column in headers:
                headers[column] = headers.pop(legacy_column)
                worksheet.cell(row=1, column=headers[column], value=column)
            if column not in headers:
                headers[column] = worksheet.max_column + 1
                worksheet.cell(row=1, column=headers[column], value=column)
        if gene == "NS5B":
            reorder_columns(worksheet, headers, NS5B_ANNOTATION_COLUMNS)
            headers = {cell.value: cell.column for cell in worksheet[1]}
        for row_index in range(2, worksheet.max_row + 1):
            refid = str(
                worksheet.cell(row=row_index, column=headers["RefID"]).value or ""
            ).strip()
            if not refid or refid not in annotations_by_refid:
                continue
            annotation = annotations_by_refid[refid]
            worksheet.cell(
                row=row_index,
                column=headers["CometSubtypes"],
                value=annotation["CometSubtypes"],
            )
            for column in annotation_columns:
                worksheet.cell(
                    row=row_index, column=headers[column], value=annotation[column]
                )
    workbook.save(args.blast_hists_xlsx)

    print(f"{args.output_csv} ({priority_added} RefID/subtype priority additions)")
    print(
        f"{args.blast_hists_xlsx} (Original_NS3, Original_NS5A, Original_NS5B updated)"
    )


if __name__ == "__main__":
    main()
