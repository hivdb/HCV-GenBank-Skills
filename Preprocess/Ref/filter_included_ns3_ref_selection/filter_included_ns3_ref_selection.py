#!/usr/bin/env python3
"""Export Original_NS3 rows whose Status begins with 'include'."""

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook


REPO_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_INPUT = REPO_ROOT / "HCVData" / "HCV_BlastHists_202604_data.xlsx"
DEFAULT_OUTPUT = (
    REPO_ROOT / "HCVData" / "Ref-selection" / "IncludedNS3Refs_StatusInclude.xlsx"
)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-xlsx", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output-xlsx", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    source_workbook = load_workbook(args.input_xlsx, data_only=False)
    source_sheet = source_workbook["Original_NS3"]
    headers = {cell.value: cell.column for cell in source_sheet[1]}
    if "Status" not in headers:
        raise ValueError("Original_NS3 must contain a Status column")

    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "Original_NS3_Included"
    for row in source_sheet.iter_rows():
        destination_row = output_sheet.max_row + 1 if row[0].row != 1 else 1
        if row[0].row != 1:
            status = str(row[headers["Status"] - 1].value or "").strip()
            if not status.casefold().startswith("include"):
                continue
        for cell in row:
            copied = output_sheet.cell(
                row=destination_row, column=cell.column, value=cell.value
            )
            if cell.has_style:
                copied._style = copy(cell._style)
            if cell.number_format:
                copied.number_format = cell.number_format
            if cell.alignment:
                copied.alignment = copy(cell.alignment)
            if cell.fill:
                copied.fill = copy(cell.fill)
            if cell.font:
                copied.font = copy(cell.font)
            if cell.border:
                copied.border = copy(cell.border)
        if row[0].row == 1:
            for column_letter, dimension in source_sheet.column_dimensions.items():
                output_sheet.column_dimensions[column_letter].width = dimension.width

    args.output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    output_workbook.save(args.output_xlsx)
    print(f"{args.output_xlsx} ({output_sheet.max_row - 1} included rows)")


if __name__ == "__main__":
    main()
